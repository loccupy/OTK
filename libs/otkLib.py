# Библиотека ПП для программ otk


def getAboutOtkLib():
    version = "29.05.2024 14:41"
    descript = "Библиотека ПП для программ otk"
    return [version, descript]


import os
import re   #для работы с регулярными выражениями
import sys
import time
import datetime
import msvcrt
import shutil
import subprocess
import psutil       #pip install psutil
import PIL
import win32gui
import pyautogui as pyautogui
import pyscreeze


import webbrowser
import segno #pip install segno для печати qr-кода
import pyperclip #pip install pyperclip для записи данных в буфер Windows
import docxtpl #pip install docxtpl для обработки *.docx-файлов
from openpyxl import load_workbook #pip install openpyxl для работы с Excell файлами (*.xlsx)
import win32print #pip install pywin32 для печати паспорта напрямую в принтер
import win32api #pip install pywin32 для печати паспорта напрямую в принтер
from win32gui import GetWindowText, EnumWindows, GetForegroundWindow, \
    SetForegroundWindow, SetWindowText, ShowWindow, IsWindowVisible
from win32con import SW_HIDE, SW_SHOWMINIMIZED, SW_MINIMIZE, SW_MAXIMIZE, \
    SW_SHOWNOACTIVATE, SW_SHOW, SW_SHOWNA, SW_RESTORE, SW_SHOWDEFAULT
from win32process import GetWindowThreadProcessId
                            

import json #для сохранения значений переменных по умолчанию в файле в формате json
from gtts import gTTS  # pip install gTTS для синтеза речи с пом.Google через интернет
from playsound import playsound  # pip install playsound==1.2.2 для воспроизведения звукового файла
from rich.console import Console    # pip install rich  для вывода таблиц
from rich.table import Table, Column
from rich.style import Style

from cryptography.fernet import Fernet  #pip install cryptography==41.0.7 библиотека шифрования


from gurux_dlms.objects import GXDLMSClock, GXDLMSData, GXDLMSRegister, GXDLMSDisconnectControl

from gurux_dlms.objects import GXDLMSDisconnectControl, GXDLMSRegister, GXDLMSProfileGeneric
from datetime import datetime, timedelta
from gurux_dlms.objects import GXDLMSClock
from gurux_dlms.enums import DataType, ObjectType
from gurux_dlms import GXDLMSClient, GXTime, GXDateTime
from gurux_dlms.GXDLMSException import *
from gurux_serial import GXSerial
from gurux_net import GXNet
from libs.GXSettings import GXSettings
from libs.GXDLMSReaderHide import GXDLMSReader
from gurux_dlms import GXDLMSClient
from docxtpl import DocxTemplate, InlineImage, RichText
from pathlib import Path
import socket       # для получения локального ip-адреса компьютера
import threading    # для созания потока внутри процесса (запись файла, когда сетевая папка не доступна)
from libs.sutpLib import savetToSUTP2, getNameEmployee, findNameMeterStatus, getInfoAboutDevice, \
    getUrlMeterConfigFile

import smtplib
from email.mime.text import MIMEText
from email.header import Header
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
from email.utils import formatdate

from tzlocal import get_localzone # $ pip install tzlocal


class bcolors:
    HEADER = '\033[95m' #пурпурный
    OKBLUE = '\033[94m' #синий
    ATTENTIONBLUE = '\033[37m\033[44m'  #белый текст на синем фоне
    ATTENTIONWARNING = '\033[30m\033[103m'  # черный текст на оранжевом фоне
    OKCYAN = '\033[96m' #цвет морской волны
    OKGREEN = '\033[92m'
    ATTENTIONGREEN = '\033[30m\033[102m'  # черный текст на зеленом фоне
    OKRESULT='\033[30m\033[42m' #для вывода результата проверки: черный текст на зеленом фоне
    WARNING = '\033[93m' #оранжевый
    FAILRED = '\033[91m' #красный цвет
    FAIL = '\033[37m\033[41m' #белый текст на красном фоне
    ENDC = '\033[0m' #код сброса
    BOLD = '\033[1m' #жирный шрифт
    WHITE = '\033[97m'
    MAGENTA = '\033[95m' #пурпурный
    UNDERLINE = '\033[4m'

global bcolors_list     #список цветов
bcolors_list=[bcolors.HEADER, bcolors.OKBLUE, bcolors.ATTENTIONBLUE,
              bcolors.OKCYAN, bcolors.OKGREEN, bcolors.OKRESULT,
              bcolors.WARNING, bcolors.FAILRED, bcolors.FAIL,
              bcolors.ENDC, bcolors.BOLD, bcolors.WHITE,
              bcolors.MAGENTA, bcolors.UNDERLINE]

from colorama import init

init()

# я закомментил
# def resource_path(relative_path):
#     try:
#         base_path = sys._MEIPASS
#     except Exception:
#         base_path = os.path.abspath(".")
#
#     return os.path.join(base_path, relative_path)


def getWorkDirLink():
    """
    Функция формирует и записывает в json путь к рабочему каталогу.
    Возвращает список с:
     1)флагом 0 или 1,
     2)текстом ошибки или сообщения об успехе,
     3)путь или пустую строку.
    """
    dirname = os.path.dirname(__file__)
    file_name_link = os.path.join(dirname,
        "link_current_folder.json")
    try:
        with open(file_name_link, "r", errors="ignore", encoding='utf-8') as file:
            content = json.load(file)
            dirname_cur= content['link_current_folder']
    except Exception as e:
        txt1=f"Не удалось получить ссылку на рабочий каталог из ф. {file_name_link}: {e.args[1]}"
        printFAIL(f"{txt1}.") 
        return["0",txt1,""]
    return["1","Ссылка на рабочий каталог сформирована",dirname_cur]


def getUserFilePath(file_name: str, only_dir="0", workmode="эксплуатация", create_folder="1"):
    """
    Функция принимает имя файла и флаги,
    возвращает путь к или файлу, или каталогу, или создает каталог, или ошибку.
    Возвращаемый результат в виде листа с флагом, сообщением и путем к файлу.
    """
    res=getWorkDirLink()
    if res[0]=="0":
        return["0",res[1],""]
    dirname_cur=res[2]

    file_name_j = os.path.join(dirname_cur,
        "1userData\\otk_config\\user_file_path.json")


    file_dic={}
    res=readWriteFile(file_name_j, "r-json", "")

    file_dic = res[2]

    
    if file_name in file_dic:
        area,folder = file_dic[file_name]

        if only_dir!="1":
            file_name_1 = os.path.join(folder, file_name)
        else:
            file_name_1=f"{folder}"
        
        if area=="local":
            file_path = os.path.join(dirname_cur, file_name_1)

            a_dir_path=os.path.split(file_path)[0]
            if only_dir=="1":
                a_dir_path=file_path
            if (not os.path.isdir(a_dir_path)) and create_folder=="1":
                res=createFolder(a_dir_path, "1")
                if res[0]=="0":
                    return ["0", res[1]]
        else:
            if "тест" in workmode:
                base_path_LAN = os.path.join(dirname_cur, "1test")
                file_path = f"{base_path_LAN}{file_name_1}"
                file_path = file_path.replace("\\\\","\\")
            
            else:
                file_path = file_name_1

        return ["1", "Имя файла/папки найдено.", file_path]

    txt1_1 = f"{bcolors.FAIL}ПП getUserFilePath(): не найдено имя файла " \
        f"{file_name}.\n{bcolors.OKBLUE}Нажмите любую клавишу."
    questionOneKey(colortxt="", txt=txt1_1)
    return ["0", f"Ошибка в ПП getUserFilePath(): имя файла/папки " \
        f"{file_name} не найдено.", ""]


def createFolder(destination_folder: str, err_msg_print="1"):
    """
    Функция принимает путь к каталогу и создает по указанному адресу такой каталог.
    Возвращает массив с флагом и сообщением
    """
    folders_list = destination_folder.split("\\", -1)
    dir_name = ""
    drive_name = ""
    i = 0
    if destination_folder[0:2] == "\\\\":
        folder = folders_list[2]
        drive_name = os.path.join("\\\\", folder)
        i = 3
    while i < len(folders_list):
        folder = folders_list[i]
        if folder == "":
            break
        if i == 0:
            drive_name = folder
            folder = folders_list[1]
            if folder == "":
                return ["0", "Неверный путь к папке-получателю."]
            folder = f"\\{folder}"
            i += 1
        dir_name = os.path.join(dir_name, folder)
        dir_name_1 = os.path.join(drive_name, dir_name)
        if not os.path.isdir(dir_name_1):
            try:
                os.makedirs(dir_name_1)
            except Exception as e:
                txt1 = f"Ошибка при создании папки " \
                    f"{dir_name_1}: {e.args[0]}"
                if err_msg_print == "1":
                    print(f"{bcolors.WARNING}{txt1}{bcolors.ENDC}")
                return ["0", txt1]
        i += 1
    return ["1", "Папка создана."]


# Закомментил, хз надо ли
# def changeDuplicateFileName(file_dest_full: str, err_msg_print="1",
#     info_msg_print="0"):
#
#     dest_folder=os.path.split(file_dest_full)[0]
#
#     file_name=os.path.split(file_dest_full)[1]
#
#     file_name_old=file_name
#
#     if os.path.exists(file_dest_full):
#         res=getFileModifedTime(file_dest_full, err_msg_print)
#         if res[0]=="0":
#             return ["0", res[1]]
#
#         dt_modify_txt=datetime.fromtimestamp(res[2]).strftime('%d%m%y_%H%M%S')
#         pos_dote=file_name.rfind(".")
#         if pos_dote==-1:
#             file_name=f"{file_name}_{dt_modify_txt}"
#         else:
#             file_name=f"{file_name[0:pos_dote]}_{dt_modify_txt}" \
#                 f"{file_name[pos_dote:len(file_name)]}"
#
#         index_name_file=0
#
#         while True:
#             index_name_file+=1
#
#             file_new_name_full=os.path.join(dest_folder, file_name)
#
#             if not os.path.exists(file_new_name_full):
#                 try:
#                     os.rename(file_dest_full, file_new_name_full)
#
#                     if info_msg_print=="1":
#                         a_txt=f"В папке '{dest_folder}'\n" \
#                             f"файл с именем '{file_name_old}' " \
#                             f"переименован в '{file_name}'."
#                         print(a_txt)
#
#                     return ["1", "Операция выполнена успешно."]
#
#                 except Exception:
#                     txt1 = f"Ошибка при изменении имени файла {file_dest_full}."
#                     if err_msg_print == "1":
#                         printWARNING(txt1)
#                     return ["0", txt1]
#
#             pos_dote=file_name.rfind(".")
#             if pos_dote==-1:
#                 file_name=f"{file_name}_{str(index_name_file)}"
#             else:
#                 if index_name_file==1:
#                     file_name=f"{file_name[0:pos_dote]}_{str(index_name_file)}" \
#                         f"{file_name[pos_dote:len(file_name)]}"
#
#                 else:
#                     a_len_suf=len(str(index_name_file-1))
#                     pos_start_index=pos_dote-a_len_suf-1
#                     file_name=f"{file_name[0:pos_start_index]}_{str(index_name_file)}" \
#                         f"{file_name[pos_dote:len(file_name)]}"
#
#
#     return ["1", "Операция выполнена успешно."]


def copyFile(file_name_full: str, file_dest_full: str,
             save_to_archive="0",err_msg_print="1"):

    destination_folder=os.path.split(file_dest_full)[0]
    if not os.path.isdir(destination_folder):
        res = createFolder(destination_folder=destination_folder,
                           err_msg_print=err_msg_print)
        if res[0] == "0":
            txt1 = f"Ошибка при создании папки " \
                f"{destination_folder}"
            return ["0", txt1]

    if save_to_archive=="1":
        file_name=os.path.split(file_dest_full)[1]
        if os.path.exists(file_dest_full)==True:
            dest_folder_archive=f"{destination_folder}\\ArchiveProgr"
            if not os.path.isdir(dest_folder_archive):
                res = createFolder(destination_folder=dest_folder_archive,
                                err_msg_print=err_msg_print)
                if res[0] == "0":
                    txt1 = f"Ошибка при создании папки " \
                        f"{dest_folder_archive}"
                    return ["0", txt1]
            try:
                shutil.copy2(file_dest_full, dest_folder_archive)
            except Exception:
                txt1 = f"{bcolors.WARNING}Ошибка при копировании файла " \
                    f"{file_dest_full}{bcolors.ENDC}"
                if err_msg_print == "1":
                    print(txt1)
                return ["0", txt1]
            file_archive_full=os.path.join(dest_folder_archive,file_name)
            file_name_archive=file_name
            now_time_str=toformatNow()[4]
            pos_dote=file_name_archive.rfind(".")
            if pos_dote==-1:
                file_name_archive=f"{file_name_archive}_{now_time_str}"
            else:
                file_name_archive=f"{file_name_archive[0:pos_dote]}_{now_time_str}" \
                    f"{file_name_archive[pos_dote:len(file_name_archive)]}"
            file_new_name_archive_full=os.path.join(dest_folder_archive,file_name_archive)
            try:
                os.rename(file_archive_full,file_new_name_archive_full)
            except Exception:
                txt1 = f"{bcolors.WARNING}Ошибка при изменении имени файла " \
                    f"{file_archive_full}{bcolors.ENDC}"
                if err_msg_print == "1":
                    print(txt1)
                return ["0", txt1]

    try:
        shutil.copy2(file_name_full, file_dest_full)
    except Exception:
        txt1 = f"{bcolors.WARNING}Ошибка при копировании файла " \
            f"{file_name_full}{bcolors.ENDC}"
        if err_msg_print == "1":
            print(txt1)
        return ["0", txt1]
    return ["1", "Файл успешно скопирован."]



def getVersUpdProgrFiles(workmode="эксплуатация"):

    auto_update_dir=""
    _,txt1, auto_update_dir = getUserFilePath('auto_update_dir',only_dir="1", 
        workmode=workmode)
    if auto_update_dir=="":
        return ["0",txt1,0,0,"","",""]

    file_name="update_progr_local.json"
    res=readGonfigValue(file_name_in=file_name,workmode=workmode)
    if res[0]!="1":
        txt1=res[1]
        return ["0",txt1,0,0,"","",""]
    config_dic=res[2]
    prog_last_upd_num=config_dic.get("progLastUpdNum",0)
    prog_last_upd_moment=config_dic.get("progLastUpdMoment","")

    prog_actual_upd_num = 0
    prog_actual_upd_moment = ""
    file_name="update_progr_public.json"
    res=readGonfigValue(file_name_in=file_name,workmode=workmode)
    if res[0]!="1":
        txt1=res[1]
        return ["0", txt1, prog_last_upd_num, prog_actual_upd_num, 
            auto_update_dir, prog_last_upd_moment, prog_actual_upd_moment]

    config_dic=res[2]
    prog_actual_upd_num=config_dic.get("progActualUpdNum",0)
    prog_actual_upd_moment=config_dic.get("progActualUpdMoment","")
    
    return ["1","Данные о номерах обновлений успешно получены.",
            prog_last_upd_num, prog_actual_upd_num, auto_update_dir,
            prog_last_upd_moment, prog_actual_upd_moment]



def updateFilesFromList(upd_file_name: str, force_on=False, 
    workmode="эксплуатация"):


    res=readGonfigValue(upd_file_name, [], {}, workmode)
    if res[0]=="0":
        return ["0", "Ошибка при формировании списка обновляемых файлов."]
    
    upd_file_list=res[2]["upd_file_list"]

    update_mode="изменен"

    ret="1"
    ret_txt="Файлы успешно обновлены."

    for i in range(0, len(upd_file_list)):
        file_dic=upd_file_list[i]
        if not file_dic["updateOn"]:
            continue

        fileSource=file_dic["fileSource"]
        file_name=file_dic["fileDest"]
        update_interval=file_dic["updateInterval"]
        last_update=file_dic["lastUpdate"]

        time_sec=toformatNow()[3]
        if last_update+update_interval*60 < time_sec or force_on:
            _, _, file_dir_source = getUserFilePath(fileSource,
                only_dir="1", workmode=workmode)
            if file_dir_source == "":
                txt_err = f"Ошибка при формировании пути до эталонного " \
                    f"ф.{file_name}."
                print(f"{bcolors.WARNING}{txt_err}{bcolors.ENDC}")
                ret="0"
                ret_txt=f"{ret_txt}, {txt_err}"

            file_path_source = os.path.join(file_dir_source, file_name)

            _, _, file_path_dest = getUserFilePath(file_name,
                only_dir="0", workmode=workmode)
            if file_path_dest == "":
                txt_err = f"Ошибка при формировании пути до эталонного " \
                    f"ф.{file_name}."
                print(f"{bcolors.WARNING}{txt_err}{bcolors.ENDC}")
                ret="0"
                ret_txt=f"{ret_txt}, {txt_err}"

            a_list=[{"fileSourcePath": file_path_source,
                  "fileDestPath": file_path_dest,
                  "updateMode": update_mode}]
            
            res = updateFiles(file_list=a_list, err_msg_print="1")
            if res[0] == "0":
                txt_err = f"Ошибка при копировании эталонного ф.{file_name}."
                print(f"{bcolors.WARNING}{txt_err}{bcolors.ENDC}")
                ret="0"
                ret_txt=f"{ret_txt}, {txt_err}"

            last_update = toformatNow()[3]
            file_dic['lastUpdate'] = last_update
            upd_file_list[i]=file_dic

    
    a_dic={"upd_file_list":upd_file_list}
    res = saveConfigValue(upd_file_name, a_dic)
    if res[0] == "0":
        txt_err="Ошибка при обновлении записи в конфигурационном файле."
        printWARNING(f"\n{txt_err}")
        ret="0"
        ret_txt=f"{ret_txt}, {txt_err}"
            
    if ret!="1":
        questionSpecifiedKey(bcolors.WARNING, "Нажмите Enter.",
            ["\r"], "", 1)
        
    return [ret, ret_txt]




def updateFiles(file_list=[], err_msg_print="1"):

    for file_cur in file_list:
        time_file_source = 0
        time_file_dest = 0
        err_txt = f"Ошибка обновления файла: "

        file_name_source=file_cur.get("fileSourcePath","")
        file_name_dest=file_cur.get("fileDestPath","")
        update_mode = file_cur.get("updateMode", "")

        if update_mode=="откл":
            continue

        if file_name_source=="" or os.path.exists(file_name_source) == False:
            err_txt = f"{err_txt}нет доступа к эталонному файлу {file_name_source}"
            if err_msg_print == "1":
                print(f"{bcolors.WARNING}{err_txt}{bcolors.ENDC}")
            return ["0", err_txt]

        if update_mode=="изменен":
            res = getFileModifedTime(file_name_source, err_msg_print)
            if res[0] == "0":
                err_txt = f"{err_txt}{file_name_source}"
                return ["0", err_txt]
            time_file_source = res[2]

            if file_name_dest!="" and os.path.exists(file_name_dest):
                res = getFileModifedTime(file_name_dest, err_msg_print)
                if res[0] == "0":
                    err_txt = f"{err_txt}{file_name_dest}"
                    return ["0", err_txt]
                time_file_dest = res[2]

        if update_mode=="замена" or time_file_dest == 0 or \
            time_file_source > time_file_dest:
            fn_1=os.path.split(file_name_dest)[1]
            print(f"Обновляю файл {fn_1}.")
            res=copyFile(file_name_source, file_name_dest,save_to_archive="1",
                    err_msg_print=err_msg_print)
            if res[0]=="1":
                print(f"{bcolors.OKGREEN}Файл {fn_1} успешно обновлен.{bcolors.ENDC}")
            else:
                print(f"{bcolors.WARNING}Не удалось обновить файл {fn_1}.{bcolors.ENDC}")
        elif update_mode=="запуск .py":
            fn_1=os.path.split(file_name_source)[1]
            print(f"Запускаю на исполнение файл {fn_1}.")
            code_exit = os.system(f"python {file_name_source}")
        elif update_mode=="запуск":
            fn_1=os.path.split(file_name_source)[1]
            print(f"Запускаю на исполнение файл {fn_1}.")
            code_exit = os.system(file_name_source)

    return ["1", "Список с файлами обработан."]



def getFileModifedTime(file_name: str, err_msg_print="1"):
    
    try:
        time_file = os.path.getmtime(file_name)
        return ["1","Время изменения файла получено.",time_file]
    except:
        txt1 = f"Ошибка при получении информации " \
            f"о времени изменении файла " \
            f"{file_name}."
        if err_msg_print == "1":
            print(f"{bcolors.WARNING}{txt1}{bcolors.ENDC}")
        return ["0", txt1,0]
    


def readGonfigValue(file_name_in: str, var_name_list=[],
    default_value_dict={},workmode="эксплуатация", msg_err_print="1"):
 
    var_config_dict={}

    var_config_dict=default_value_dict.copy()

    if len(default_value_dict)>0 and len(var_name_list)>0:
        for var_name in var_name_list:
            if var_name in default_value_dict:
                var_config_dict[var_name]=default_value_dict[var_name]

    _,_,file_name = getUserFilePath(file_name_in,workmode=workmode)
    if file_name=="":
        return ["0","Ошибка при формировании пути до файла.",{}]

    res=readWriteFile(file_name,operation="r-json",content="",
        encoding="utf-8",err_msg_set=msg_err_print)
    if res[0]=="0":
        return ["2","Ошибка при чтении json файла конфигурации.",{}]
    content_dict = res[2]
    
    var_content_key_list=list(content_dict.keys())
    for key in var_content_key_list:
        var_config_dict[key]=content_dict[key]

    return ["1", "Словарь со значениями переменных " 
        "сформирован.", var_config_dict]


def saveConfigValue(file_name_in: str, var_config_dict: dict,
    workmode="эксплуатация", write_mode="заменить часть"):
     
    _,_,file_name = getUserFilePath(file_name_in,workmode=workmode)
    if file_name=="":
        return ["0","Ошибка при формировании пути до файла.",{}]
    
    
    var_config_file_dict=var_config_dict.copy()

    if os.path.exists(file_name) and write_mode=="заменить часть":
        res=readWriteFile(file_name,operation="r-json",content="",
            encoding="utf-8",err_msg_set="1")
        if res[0]=="0":
            return ["0","Ошибка при чтении файла конфигурации.",{}]
        var_config_file_dict = res[2]
        var_config_key_list=list(var_config_dict.keys())
        for key in var_config_key_list:
            var_config_file_dict[key]=var_config_dict[key]
        
    res=readWriteFile(file_name,operation="w-json",
        content=var_config_file_dict,
        encoding="utf-8",err_msg_set="1")
    if res[0]=="0":
        return ["0","Ошибка при записи файла конфигурации.",{}]
    return ["1","Конфигурационный словарь успешно сохранен в файле."]



def getLocalIP():
    hostname = socket.gethostname()
    local_ip = socket.gethostbyname(hostname)
    return local_ip



def toformatNow():
    pc_time = datetime.now()
    pc_time1 = str(pc_time)
    date1 = f"{pc_time1[8:10]}.{pc_time1[5:7]}.{pc_time1[0:4]}"
    date1_revers=f"{pc_time1[0:4]}.{pc_time1[5:7]}.{pc_time1[8:10]}"
    now_full = f"{date1} {pc_time1[11:19]}"
    now_full_1=f"{pc_time1[8:10]}{pc_time1[5:7]}{pc_time1[0:4]}_" \
        f"{pc_time1[11:13]}{pc_time1[14:16]}{pc_time1[17:19]}"
    second = pc_time.timestamp()
    time_full=f"{pc_time1[11:19]}"
    tz = get_localzone() # local timezone
    d = datetime.now(tz) # or some other local date 
    utc_offset = d.utcoffset().total_seconds()

    return [now_full, date1, pc_time, second, now_full_1, time_full,
            date1_revers, tz, utc_offset]



def toConvertDate(date=None):

    if date==None or date=="":
        date=str(datetime.now())
        date=f"{date[8:10]}.{date[5:7]}.{date[0:4]}"

    date0=f"{date[6:10]}.{date[3:5]}.{date[0:2]}"
    date1=f"{date[0:2]}_{date[3:5]}_{date[6:10]}"
    date2=f"{date[0:2]}{date[3:5]}{date[6:10]}"
    date3=f"{date[0:2]}{date[3:5]}{date[8:10]}"
    date4=f"{date[6:10]}-{date[3:5]}-{date[0:2]}T00:00:00.000Z"

    return[date0, date1, date2, date3, date4]



def correctDateTime(dt: str, dt_format_in='%Y-%m-%dT%H:%M:%S',
        dt_format_out='%d.%m.%Y %H:%M',dont_to_correct="откл",
        weeks=0,days=0,hours=0,minutes=0,seconds=0):

    dt_datetime = datetime.strptime(dt, dt_format_in)
    dt_datetime_correct = dt_datetime
    if dont_to_correct=="откл":
        dt_datetime_correct = dt_datetime+ \
            timedelta(weeks=weeks, days=days,
                      hours=hours, minutes=minutes, 
                      seconds=seconds)
    date_time_txt = dt_datetime_correct.strftime(dt_format_out)
    return date_time_txt


def replaceTitleWindow(title_search="", title_new=""):
    
    hwnd=0

    if title_search!=None and title_search!="":
        res = searchTitleWindow(title_search)
        if res[0] in ["0", "2"]:
            return ["0", res[1], hwnd]

        hwnd = res[2]
    
    else:
        hwnd = GetForegroundWindow()

    SetWindowText(hwnd, title_new)
    
    return ["1", "Заголовок окна заменен.", hwnd]


def getPidProcess(filname:str, time_wait_exec=2):

    cicl_find_max=2
    for i in range(0, cicl_find_max):
        for process in psutil.process_iter():
            if process.name() == filname:
                return process.pid
        print(f"Ожидаю запуск программы {filname} " 
            f"({i+1}/{cicl_find_max}).")
        pause_ui(time_wait_exec)

    return None



def getHwndPid(pid):


    def innerEnumWindowCallback(hwnd, pid ): 
        tid, current_pid = GetWindowThreadProcessId(hwnd)
        if pid == current_pid and IsWindowVisible(hwnd): 
            hwnd_windows.append(hwnd)

    hwnd_windows = []

    try:
        EnumWindows(innerEnumWindowCallback, pid)
    except Exception as e:
        a_err_txt = "При запросе списка окон процесса " \
            "возникла ошибка."
        print(f"{bcolors.WARNING}{a_err_txt}{bcolors.ENDC}")
        return ["0", a_err_txt]

    hwnd1=0
    if len(hwnd_windows)>0:
        hwnd1=hwnd_windows[0]
        return ["1", "Искомый процесс найден.", hwnd1, 
                hwnd_windows]

    return ["2", "Окна для процесса не найдены.", hwnd1, 
                hwnd_windows]


def moveResizeWindow(title_window_list: list, x: int, y: int, w: int, h: int, print_err_msg="1"):
    """
    Перемещает и изменяет размер окна в Windows по заданному заголовку
    (или использует активное окно, если заголовок не найден). Разберём его детально.
    """
    hwnd = None
    
    if len(title_window_list)>0:
        for title_window in title_window_list:
            res = searchTitleWindow(title_window)
            if res[0] == "1":
                hwnd = res[2]
                break

        else:
            title_window_str=", ".join(title_window_list)
            a_err_msg=f"Заголовок окна: {title_window_str} " \
                "не найден."
            
            if len(title_window_list)>1:
                a_err_msg=f"Заголовки окон: {title_window_str} " \
                    "не найдены."
            if print_err_msg=="1":
                print(a_err_msg)

            return["2", "Заголовок окна не найден."]

    if hwnd==None:
        hwnd = win32gui.GetForegroundWindow()

    x_1, y_1, w_1, h_1 = win32gui.GetWindowRect(hwnd)

    if x==-1:
        x=x_1

    if y==-1:
        y=y_1

    if w==-1:
        w=w_1

    if h==-1:
        h=h_1

    win32gui.MoveWindow(hwnd, x, y, w, h, True)

    return ["1", "Операция с окном проведена успешно."]



def searchTitleWindow(title_window=""):

    
    def innerWindowEnumerationHandler(hwnd, top_windows):
        title = GetWindowText(hwnd)
        tid, pid = GetWindowThreadProcessId(hwnd)
        top_windows.append((hwnd, pid, title))
    
    
    top_windows = []
    hwnd=None
    pid=None

    try:
        EnumWindows(innerWindowEnumerationHandler, top_windows)
    except Exception as e:
        a_err_txt="При запросе списка запущенных процессов " \
            "возникла ошибка."
        print (f"{bcolors.WARNING}{a_err_txt}{bcolors.ENDC}")
        return ["0", a_err_txt, hwnd, pid, top_windows]
    
    ret_list=[]
    first_window=True

    if title_window!="" and title_window!=None:
        for window in top_windows:
            if title_window in window[2]:
                ret_list.append(window)
                if first_window:
                    hwnd, pid,_=window
                    first_window=False
        
        if len(ret_list)>0:
            return ["1","Искомый заголовок найден.",hwnd, 
                pid, ret_list ]
        
        return ["2","Искомый заголовок отсутствует.", None, 
            None, top_windows]

    return ["3","Список процессов заполнен.", None, 
        None, top_windows]



def actionsSelectedtWindow(title_window_list: list, hwnd=None,
    actions="закрыть", print_err_msg="1"):
    """
    Принимает дескриптор, список заголовков, команду к действию,
    совершает указанное действие,
    возвращает флаг, текст состояния и дескриптор окна.
    """

    show_window_dic = {"скрыть": [SW_HIDE],
        "свернуть+активировать": [SW_SHOWMINIMIZED],
        "максимизировать+активировать": [SW_MAXIMIZE],
        "показать2": [SW_SHOWNOACTIVATE],
        "свернуть": [SW_MINIMIZE],
        "активировать": ["activate"],
        "показать": [SW_RESTORE],
        "показать+активировать": [SW_RESTORE, "activate"],
        "закрыть": ['']}


    var_show_window = show_window_dic.get(actions, None)
    if var_show_window==None:
        a_err_txt="Неизвестная операция над окном: " \
            f"{actions}."
        printMsgWait(a_err_txt,bcolors.WARNING, print_err_msg)
        return ["0", a_err_txt]

    top_windows = []
    ret_hwnd=0
    if hwnd!=None:
        ret_hwnd=hwnd

    if hwnd==None or hwnd==0:
        if len(title_window_list)>0:
            for title_window in title_window_list:
                res = searchTitleWindow(title_window)
                if res[0]=="0":
                    return ["0", res[1], ret_hwnd]

                elif res[0]=="2":
                    continue    

                select_windows_list=res[4]
                hwnd = res[2]
                ret_hwnd = hwnd
                break

        else:
            a_err_txt="Ошибка при операциях с окнами: отсутствует " \
                "дескриптор и заголовок окна."
            
            printMsgWait(a_err_txt,bcolors.WARNING, print_err_msg)

            return ["0", a_err_txt]
        

    if actions=="закрыть":
        if hwnd!=None and hwnd!=0:
            try:
                tid, pid = GetWindowThreadProcessId(hwnd)#  — вызов WinAPI для получения идентификаторов потока (TID) и процесса (PID), которым принадлежит указанное окно
                os.system(f'taskkill /f /t /pid {pid}')
            except Exception:
                a_err_txt="Возникла ошибка при закрытии окна."
                printMsgWait(a_err_txt,bcolors.WARNING, 
                                print_err_msg)
                return ["0", a_err_txt]

        else:
            for window in select_windows_list:
                try:
                    tid, pid = GetWindowThreadProcessId(window[0])
                    os.system(f'taskkill /f /t /pid {pid}')
                except Exception:
                    a_err_txt="Возникла ошибка при закрытии окна."
                    printMsgWait(a_err_txt,bcolors.WARNING, print_err_msg)
                    return ["0", a_err_txt]
    
    else:
        for a_cur in var_show_window:
            if a_cur=="activate":
                try:
                    SetForegroundWindow(hwnd)#  — это вызов WinAPI, который делает указанное окно текущим активным (foreground) окном и передаёт ему фокус ввода
                except Exception:
                    a_err_txt="Возникла ошибка при активации окна."
                    printMsgWait(a_err_txt,bcolors.WARNING, print_err_msg)
                    return ["0", a_err_txt]

            else:
                try:
                    ShowWindow(hwnd, a_cur)#  — вызов WinAPI для управления видимостью и состоянием окна по его дескриптору (hwnd).
                except Exception:
                    a_err_txt="Возникла ошибка при операции над окном."
                    printMsgWait(a_err_txt,bcolors.WARNING, print_err_msg)
                    return ["0", a_err_txt]

    return["1", "Операция выполнена.", ret_hwnd]







def comPortList(print_list="1"):
    import serial.tools.list_ports
    if print_list=="1":
        print("Запрашиваю у системы о наличии COM-портов")
    ports = serial.tools.list_ports.comports()
    port_list = []
    for port, desc, hwid in sorted(ports):
        port_list.append(port)
        if print_list=="1":
            print(port)
    return port_list


def settingOpt(password,serial_num, comport, msg_print="1", 
               authentication="High"):

    global reader, settings

    try:
        import pkg_resources
    except Exception:
        print("pkg_resources not found")
        return ["0"]

    try:
        if msg_print=="1":
            print("gurux_dlms version: " +
                pkg_resources.get_distribution("gurux_dlms").version)
            print("gurux_net version: " +
                pkg_resources.get_distribution("gurux_net").version)
            print("gurux_serial version: " +
                pkg_resources.get_distribution("gurux_serial").version)
        else:
            pkg_resources.get_distribution("gurux_dlms").version
            pkg_resources.get_distribution("gurux_net").version
            pkg_resources.get_distribution("gurux_serial").version
    except Exception:
        print("pkg_resources not found")
        return ["0"]



    opto_baudRate = 9600
    serverAddress = 127
    if serial_num!="":
        serverAddress=int(serial_num)+16
    hight_password = password
    deltatime = 3600  # в секундах

    clientAddress_dic={"High":48, "Low":4}
    clientAddress=clientAddress_dic.get(authentication,48)
    reader = None
    settings = GXSettings()

    settings.getParameters("COM", f"{comport}", password=hight_password, 
        authentication=authentication, serverAddress=serverAddress,
        logicalAddress=1, clientAddress=clientAddress, baudRate=opto_baudRate)

    if not isinstance(settings.media, (GXSerial, GXNet)):
        raise Exception("Unknown media type.")
    reader = GXDLMSReader(settings.client, settings.media,
                          settings.trace, settings.invocationCounter)

    while True:
        try:
            settings.media.open()
            return ["1",reader,settings]
        except Exception as e:
            err_msg=e.args[0]
            if "WinError 5" in err_msg:
                txt1_1 = f"\n{bcolors.FAIL}Отказано в доступе. " \
                    f"{comport} занят другой программой.{bcolors.ENDC}\n" \
                    f"{bcolors.OKBLUE}Закройте другую программу, " \
                    f"использующую оптопорт/преобразователь RS-485.{bcolors.ENDC}\n" \
                    f"{bcolors.OKBLUE}Если такой программы нет, то " \
                    f"попробуйте отключить оптопорт/преобразователь RS-485 от COM-порта " \
                    f"и через 5 сек. подключить его заново.{bcolors.ENDC}"
            elif ("WinError 2" in err_msg) or ("Serial port is not selected." in err_msg) :
                txt1_1 = f"\n{bcolors.FAIL}Не найден COM-порт оптопорта/преобразователя RS-485 " \
                    f"({comport}). " \
                    f"Вероятно оптопорт/преобразователь RS-485 не подключен.{bcolors.ENDC}\n" \
                    f"{bcolors.OKBLUE}Подключите оптопорт/преобразователь RS-485 к компьютеру." \
                    f"{bcolors.ENDC}"
            else:
                txt1_1=err_msg
                
            txt1_1=f"{txt1_1}\n{bcolors.OKBLUE}Для повторной попытки " \
                f"подключения нажмите 1.{bcolors.ENDC}\n" \
                f"{bcolors.OKBLUE}Чтобы прекратить попытки подключения " \
                f"нажмите 0.{bcolors.ENDC}"
            spec_keys=["0","1"]
            oo=questionSpecifiedKey("",txt=txt1_1,specified_keys_in=spec_keys, \
               file_name_mp3="",specified_keys_only=1)
            if oo=="0":
                return ["0"]
        

def checkComPort(com_opto: str, print_msg="all", interface_name=""):
    
    txt_com_opto=f"{bcolors.FAIL}Порт {com_opto} отсутствует.{bcolors.ENDC}"
    if interface_name!=None and interface_name!="":
        txt_com_opto=f"{bcolors.FAIL}Для {interface_name} отсутствует " \
            f"порт {com_opto}.{bcolors.ENDC}"

    ret="0"

    com_opto_set=False

    if com_opto!="":
        com_port_list=comPortList(print_list="0")
        if com_opto in com_port_list:
            ret="1"
            if print_msg=="all":
                txt_com_opto=f"-> {bcolors.OKGREEN}{com_opto}{bcolors.ENDC}"
                print(txt_com_opto)
            com_opto_set=True
    
    if com_opto_set==False and print_msg!="no":
        print(txt_com_opto)
    
    return ret



def checkComPortList(com_port_list_in: list, print_msg="no", interface_name=""):

    com_port_list=com_port_list_in.copy()

    comports_txt_list=[]
    a_com_recurring_list=[]
    ret_id="1"

    for com_cur in com_port_list:
        if com_cur in a_com_recurring_list:
            if ret_id=="1":
                ret_id="4"
            continue

        if com_cur!="":
            a_com_recurring_list.append(com_cur)
            res=checkComPort(com_opto=com_cur,print_msg=print_msg)
            if res=="0":
                comports_txt_list.append(com_cur)
                ret_id="2"
        
        else:
            return ["3", "В списке имеются пустые значения для COM-порта.", []]

    a_ret_dic={"1":"Все COM-порты из списка найдены и не повторяются.",
               "2":"Не найдены все или некоторые COM-порты",
               "4":"Некоторые COM-порты повторяются"}
    
    return [ret_id, a_ret_dic[ret_id], comports_txt_list]



def toInitConnectOpto(prot_filename="", employees="", source="otk_opto",
    com_opto=""):

    global connection_initialized

    cicl = True
    connection_initialized=False
    while cicl:
        try:
            if connection_initialized==False:
                reader.initializeConnection()
                connection_initialized=True
            return 1
        except Exception as e:
            if e.args[0] == "Connection is permanently rejected\r\nAuthentication failure.":
                print(f"{bcolors.FAIL}ПУ отказал в доступе. Проверьте пароль.{bcolors.ENDC}")
                toCloseConnectOpto()
                return 3
            elif "Serial port is not open" in e.args[0]:
                res = checkComPort(com_opto=com_opto, print_msg="")
                if res == "1":
                    return 5
                elif res == "0":
                    print(f"{bcolors.FAIL}COM-порт оптопорта/преобразователя RS-485 " \
                        f"не найден.{bcolors.ENDC}\n" \
                        f"{bcolors.FAIL}Выполните его повторное подключение.{bcolors.ENDC}")
                    return 4
            oo = communicationTimoutError(
                "Инициализация связи с ПУ:", e.args[0], prot_filename, employees)
            if oo == "0" or oo == "-1":
                return 2


def ExchangeBetweenPrograms(operation: str, recipient: str, content={},
    workmode="эксплуатация"):

    exchange_dict = {}

    _, _, file_path = getUserFilePath("exch_programs.json",workmode=workmode)
    if file_path == "":
        return ["0","Ошибка при формировании пути до файла.",""]

    try:
        with open(file_path, "r", errors="ignore", encoding='utf-8') as file:
            exchange_dict = json.load(file)
    except Exception as e:
        if e.args[0] != 2:
            return ["0", f"{e.args[0]}"]
    rec = {}
    if operation == "add":
        exchange_dict[recipient] = content
        msg_txt = "Значение записано"
    elif operation == "read":
        val = exchange_dict.get(recipient, "")
        if val != "":
            return ["1", "Значение найдено", val]
        return ["0", "Значение отсутствует", val]
    elif operation == "del":
        val = exchange_dict.get(recipient, "")
        if val != "":
            exchange_dict.pop(recipient)
        msg_txt = "Запись удалена"
    elif operation == "clear":
        exchange_list = {}
        msg_txt = "Словарь очищен"
    else:
        return ["0", "Такой операции нет"]
    try:
        with open(file_path, "w", errors="ignore", encoding='utf-8') as file:
            json.dump(exchange_dict, file)
            return ["1", msg_txt]
    except Exception as e:
        return ["0", f"{e.args[0]}"]


def toCloseConnectOpto():
    global connection_initialized   #признак инициализации канала связи
    cicl = True
    connection_initialized = False
    while cicl:
        try:
            reader.disconnect()
            return 1
        except Exception as e:
            return 2


def toReadDataFromMeter(adr, item, err_txt, decode="", prot_filename="", employees=""):
    ret=[]
    obj_data=GXDLMSData(adr)
    cicl = True
    while cicl:
        try:
            if decode!="":
                data = reader.read(obj_data, item).decode("utf-8")
            else:
                data = reader.read(obj_data, item)
            ret=[1,data]
            break
        except Exception as e:
            oo = communicationTimoutError(err_txt, e.args[0], prot_filename, employees)
            if oo == "0" or oo == "-1":
                ret=[0,0]
                break
    return ret


def communicationTimoutError(err_operation, err_txt, prot_filename="", employees=""):
    global connection_initialized
    try:
        reader.disconnect()
    except Exception as e:
        pass
    if err_txt == "Failed to receive reply from the device in given time.":
        connection_initialized = False
        txt1 = err_operation + \
            ": ПУ не отвечает на запрос.\nПроверьте состояние ПУ, положение оптопорта " \
                "или состояние преобразователя RS-485.\n"
        txt1 = txt1+"Повторить запрос? 0- нет, 1- да"
        oo = questionSpecifiedKey(bcolors.WARNING, txt1, [
                                  "0", "1"], "NoAnswerFromMeter")
        print()
        if oo == "1":
            cicl = True
            while cicl:
                try:
                    reader.initializeConnection()
                    connection_initialized = True
                    return "1"
                except Exception as e:
                    if e.args[0] == "Failed to receive reply from the device in given time.":
                        txt2 = "Инициализация связи: ПУ не отвечает на запрос.\n" \
                            "Проверьте соостояние ПУ, положение оптопорта или состояние " \
                            "преобразователя RS-485."
                        print(f"{bcolors.WARNING}{txt2}{bcolors.ENDC}")
                        txt3 = txt2+"\nПовторить запрос? 0- нет, 1- да"
                        oo = questionSpecifiedKey(bcolors.WARNING, txt3, [
                                                  "0", "1"], "NoAnswerFromMeter")
                        print()
                        if oo == "0":
                            testBreak("", "Нет связи со счетчиком",
                                      prot_filename, employees)
                            return "0"
                    else:
                        print(f"{bcolors.FAIL}{err_operation}{err_txt}")
                        testBreak(
                            "", "Прочие ошибки при получении данных от счетчика", prot_filename, employees)
                        return "-1"
        testBreak("", "Нет связи со счетчиком", prot_filename, employees)
        return oo
    else:
        print(f"{bcolors.FAIL}{err_operation}{err_txt}{bcolors.ENDC}")
        testBreak("", "Прочие ошибки при получении данных от счетчика",prot_filename, employees)
        return "-1"


def testBreak(current_test_txt, reason_txt, prot_filename="", employees=""):
    txt1_1="Проверка ПУ прервана."
    if employees=="":
        txt1_1="Опрос ПУ прерван."
    print(f"{bcolors.WARNING}\n{txt1_1}{bcolors.ENDC}")
    toCloseConnectOpto()
    if prot_filename != "" and employees!="":
        txt1 = "\n"+current_test_txt+"\n"+reason_txt + \
            "\nПроверку проводил:"+employees + "\n"
        fileWriter(prot_filename, "a", "", txt1,
                "Запись в отчет информации в связи с прерыванием теста")


def meterPasswordInfo(password: str, err_msg_set="on"):
    meter_pass_default_dict={
        "Стандартный высокого уровня": "1234567898765432",
        "Карелия 12.23 высокого уровня": "EkkGsmAdmin2021i",
        "БЭСК 01.24 высокого уровня": "besk000000000000"
        }   
    res=readGonfigValue("meter_pass.json",[],meter_pass_default_dict)
    if res[0]!="1":
            return ["0","Не удалось прочитать данные " \
                "о пароле из файла","",""]
    meter_hight_password_dict=res[2]
    keys_list=list(meter_hight_password_dict.keys())
    for key in keys_list:
        if meter_hight_password_dict[key]==password:
            return ["1","Успешно",password,key]  
    if err_msg_set=="on":
        txt1=f"{bcolors.FAIL}Пароль '{password}' в списке паролей " \
            f" не найден.{bcolors.ENDC}\n" \
            f"{bcolors.OKBLUE}Нажмите Enter.{bcolors.ENDC}"
        spec_keys=["\r"]
        questionSpecifiedKey(
            colortxt="", txt=txt1, specified_keys_in=spec_keys, 
            file_name_mp3="", specified_keys_only=1)
    return ["0","Пароль в списке не найден","",""]

        

def waitOneKey():
    norm_micro = 500000  # нормируемое зн-е 0,5 сек
    pc_time1 = datetime.now() #время, когда появился запрос на экране
    oo1 = msvcrt.getwch()
    cicl=True
    while cicl:
        pc_time2 = datetime.now()
        delta_micro = int(abs(pc_time2 - pc_time1).microseconds)
        delta_sec = int(abs(pc_time2 - pc_time1).seconds)
        if delta_sec>0 or delta_micro>norm_micro:
            break  
        pc_time1 = datetime.now() #время, когда очистили буфер ввода
        oo1 = msvcrt.getwch()
    return oo1


def questionOneKey(colortxt: str,txt: str):
    ret=''
    if colortxt=='':
        print(f"{txt}")
    else:
        print(f"{colortxt}{txt}{bcolors.ENDC}")
    oo1 = waitOneKey()
    ret=oo1
    return ret



def questionSpecifiedKey(colortxt, txt, specified_keys_in: list, file_name_mp3="", specified_keys_only=0):
    specified_keys=listCopy(specified_keys_in)

    if colortxt != None and colortxt != '':
        printColor(txt, colortxt)
    else:
        printColor(txt)
    if file_name_mp3!="":
        dirname = os.path.dirname(__file__)
        file_name_mp3_full=os.path.join(dirname, f'speech\\\\{file_name_mp3}.mp3')
        if os.path.exists(file_name_mp3_full)==False:
            try:
                tts = gTTS(txt, lang="ru", slow=False)
                tts.save(file_name_mp3_full)
            except Exception:
                pass
    cicl=True
    while cicl:
        oo1 = waitOneKey()
        if specified_keys_only ==1:
            for i in specified_keys:
                if i == oo1:
                    if oo1=="\r":
                        print()
                    elif colortxt == '':
                        print(oo1,end="")
                    else:
                        print(f"{colortxt}{oo1} {bcolors.ENDC}", end="")
                    return oo1
        else:
            if oo1=="\r":
                print()
            elif colortxt == '':
                print(oo1, end="")
            else:
                print(f"{colortxt}{oo1} {bcolors.ENDC}", end="")
            for i in specified_keys:
                if oo1 == i:
                    return oo1



def keystrokeEnter(txt=""):
    if txt=="" or txt==None:
        txt="Нажмите Enter."
        
    questionSpecifiedKey(bcolors.OKBLUE, txt, ["\r"], "", 1)



def inputSpecifiedKey(colortxt, txt, err_txt, len_input_list=[], 
    specified_keys_list=[], specified_keys_only=0, edit_str="",
    print_number_big_font="откл", workmode="эксплуатация"):

    
    def innerClrBigFont():
        a_dic={"input_text": ""}
        saveConfigValue("print_big_font_line.json", a_dic,
            workmode, "заменить часть")
        
        return
        
    
    win_clipboard_on=False
    win_clipboard_pos=-1
    
    err_txt1=err_txt
    cursor_select_dic={"P":"#вниз", "H":"#вверх", 
        "K":"#влево", "M":"#вправо"}
    
    min_max_descript=""
    for a_spec_key in specified_keys_list:
        if "#num_fract_limit" in a_spec_key:
            min_max_descript="Число должно быть "
            a_range=a_spec_key[(len("num_fract_limit")+1):]
            a_1=a_range.find(":")
            a_min=a_range[1:a_1]
            a_max=a_range[a_1+1:(len(a_range)-1)]
            if a_min!="#" and a_range[0]=="[":
                min_max_descript=f"{min_max_descript}>= {a_min}"
                min_range=float(a_min)
            
            elif a_min!="#" and a_range[0]=="(":
                min_max_descript=f"{min_max_descript}> {a_min}"
                min_range=float(a_min)

            if a_max!="#" and a_range[-1]=="]":
                if a_min!="#":
                    min_max_descript=f"{min_max_descript} и"

                min_max_descript=f"{min_max_descript} <= {a_max}"
                max_range=float(a_max)

            elif a_max!="#" and a_range[-1]==")":
                if a_min!="#":
                    min_max_descript=f"{min_max_descript} и"

                min_max_descript=f"{min_max_descript} < {a_max}"
                max_range=float(a_max)

            min_max_descript=f"{min_max_descript}."
            break
    

    cicl = True
    while cicl:
        input_str=""
        spec_err=False
        
        a_txt=txt
        if min_max_descript!="":
            if a_txt!="":
                a_txt=f"{a_txt}\n{min_max_descript}"
            else:
                a_txt=min_max_descript
        print(f"{colortxt}{a_txt}{bcolors.ENDC}")
        if edit_str!="":
            print(edit_str, end="",flush=True)
            input_str=str(edit_str)
        if print_number_big_font=="окно":
            innerClrBigFont()
            
            closer_func_big_font=getOneSimbFile("input_text")
            oo1=closer_func_big_font()
        
        else:
            oo1 = waitOneKey()

        for i in specified_keys_list:
            if i==oo1:
                k = 0
                for j in specified_keys_list:
                    a = j.find(oo1, 0)
                    if j != i and a == 0:
                        k = k+1
                if k == 0:
                    print (oo1,end="",flush=True)
                    return oo1

        cicl_first_on=True
        cicl1=True
        while cicl1:

            win_clipboard= pyperclip.paste()

            if ord(oo1)==22 and len(win_clipboard)!=0 and not win_clipboard_on:
                win_clipboard_on=True
                win_clipboard_pos=0

            if win_clipboard_on:
                oo1=win_clipboard[win_clipboard_pos]
                win_clipboard_pos+=1
                
                
            if cicl_first_on:
                cicl_first_on=False
            else:
                if print_number_big_font=="окно":
                    oo1=closer_func_big_font()
                
                else:
                    if not win_clipboard_on: 
                        oo1 = msvcrt.getwch()

                    else:
                        if win_clipboard_pos==len(win_clipboard):
                            win_clipboard_on=False



            if ascii(oo1)=="'\\xe0'":
                if print_number_big_font=="окно":
                    oo1=closer_func_big_font()
                
                else:
                    oo1 = msvcrt.getwch()
                    
                if oo1 in cursor_select_dic and \
                    cursor_select_dic[oo1] in specified_keys_list:
                    return cursor_select_dic[oo1]
                else:
                    continue
            elif oo1=="\r":
                if input_str=="" and ("\r" not in specified_keys_list):
                    continue 
                if (specified_keys_only==1 and input_str in specified_keys_list) or \
                    specified_keys_only!=1:
                    print()
                    break
                else:
                    continue 
            elif oo1=="\b":
                l = len(input_str)
                if l>0:
                    input_str=input_str[0:l-1:1]
                    print(oo1,end="")
                    print(" ",end="")
                    print(oo1, end="",flush=True)
                    continue
            elif oo1.isalnum() or oo1==" " or \
                all(ch in '/*-+.,<>()&^%$#!\`~;:?[]{}=_|' for ch in oo1):
                input_str1 = input_str+oo1
                for i in specified_keys_list:
                    if i == input_str1:
                        k=0
                        for j in specified_keys_list:
                            a = j.find(input_str1,0)
                            if j!=i and a==0:
                                k=k+1
                        if k==0:
                            print (oo1,end="",flush=True)
                            return input_str1
                if specified_keys_only==1:
                    for j in specified_keys_list:
                        a = j.find(input_str1,0)
                        if a==0:
                            input_str = input_str1
                            print(oo1, end="",flush=True)
                            break
                else:
                    input_str = input_str1
                    print(oo1,end="",flush=True)

        oo1=input_str
        oo1=oo1.lstrip()
        if len_input_list[0] == 0 and oo1 == "":
            if err_txt == "":
                err_txt1 = f"Минимальная длина строки 1 символ."
            print(f"{bcolors.WARNING}{err_txt1}{bcolors.ENDC}")
            
            if print_number_big_font=="окно":
                a_dic={"text_3": "ERR  LEN"}
                saveConfigValue("print_big_font_line.json", a_dic,
                    workmode, "заменить часть")
                
        elif specified_keys_list!="":
            for i in range(len(specified_keys_list)):
                if specified_keys_list[i]=="#date":
                    res=checkCorrectDate(oo1)
                    if res[0]=="1":
                        oo=res[2]
                        return oo
                    else:
                        spec_err=True
                        break
                
                elif "#num_fract" in specified_keys_list[i]:
                    a_spec_key=specified_keys_list[i]
                    
                    oo1=oo1.replace(",", ".")
                    oo1=oo1.rstrip(".")
                    try:
                        a_oo1=float(oo1)
                        if not "num_fract_limit" in a_spec_key:
                            if a_oo1==int(a_oo1):
                                a_oo1=int(a_oo1)
                            return a_oo1
                        
                        a_ok=True
                        if "[" in a_range and a_min!="#" and a_oo1<min_range:
                            a_ok=False

                        if  "]" in a_range and a_max!="#" and a_oo1>max_range:
                            a_ok=False

                        if "(" in a_range and a_min!="#" and a_oo1<=min_range:
                            a_ok=False
                                
                        if ")" in a_range and a_max!="#" and a_oo1>=max_range:
                            a_ok=False
                        
                        if a_ok:
                            if a_oo1==int(a_oo1):
                                a_oo1=int(a_oo1)
                            return a_oo1
                            
                    except Exception:
                        pass

                    printWARNING(min_max_descript)

                    if print_number_big_font=="окно":
                        a_dic={"text_3": "ERR  NUMBER"}
                        saveConfigValue("print_big_font_line.json", a_dic,
                            workmode, "заменить часть")
                        
                    spec_err=True
                    break
                    
                elif specified_keys_list[i] == "#num":
                    oo1=oo1.replace(",", ".")
                    oo1=oo1.rstrip(".")
                    try:
                        a_oo1=float(oo1)
                        if a_oo1>=0 and str(int(a_oo1))==oo1:
                            return oo1
                    except Exception:
                        pass

                    if err_txt == "":
                        err_txt1 = "Должно быть введено целое положительное " \
                            "число включая 0."
                    print(f"{bcolors.WARNING}{err_txt1}{bcolors.ENDC}")

                    if print_number_big_font=="окно":
                        a_dic={"text_3": "ERR  NUMBER"}
                        saveConfigValue("print_big_font_line.json", a_dic,
                            workmode, "заменить часть")
                        
                    spec_err=True
                    break
                
                elif specified_keys_list[i] == "#num>0":
                    oo1=oo1.replace(",", ".")
                    oo1=oo1.rstrip(".")
                    try:
                        a_oo1=float(oo1)
                        if a_oo1>0 and str(int(a_oo1))==oo1:
                            return oo1
                    except Exception:
                        pass

                    if err_txt == "":
                        err_txt1 = "Должно быть введено целое положительное число > 0."
                    print(f"{bcolors.WARNING}{err_txt1}{bcolors.ENDC}")

                    if print_number_big_font=="окно":
                        a_dic={"text_3": "ERR  NUMBER"}
                        saveConfigValue("print_big_font_line.json", a_dic,
                            workmode, "заменить часть")
                        
                    spec_err=True
                    break
                
                elif specified_keys_list[i] == "#alpha":
                    if re.search(r'[^\W\d]',oo1) is not None:
                        return oo1
                    if err_txt == "":
                        err_txt1 = "Должна быть введена строка из букв, а не только цифры."
                    print(f"{bcolors.WARNING}{err_txt1}{bcolors.ENDC}")

                    if print_number_big_font=="окно":
                        a_dic={"text_3": "ERR  STRING"}
                        saveConfigValue("print_big_font_line.json", a_dic,
                            workmode, "заменить часть")

                    spec_err = True
                    break
                elif oo1 == specified_keys_list[i]:
                    return oo1
        
        if spec_err==False and specified_keys_only!=1:
            l1=len(oo1)
            for len1 in len_input_list:    
                if (len1 != 0 and len(oo1) == len1) or (len1 == 0 and oo1 != ""):
                    return oo1
                
            if err_txt == "":
                k = 0
                for i in len_input_list:
                    if k == 0:
                        txt1_1 = str(i)
                    elif k==1:
                        txt1_1 = txt1_1+" или "+str(i)
                    else:
                        txt1_1 = txt1_1+", или "+str(i)
                    k += 1
                err_txt1 = f"Длина строки должна быть {txt1_1} симв."
            print(f"{bcolors.WARNING}{err_txt1}{bcolors.ENDC}")

            if print_number_big_font=="окно":
                a_dic={"text_3": "ERR  LEN"}
                saveConfigValue("print_big_font_line.json", a_dic,
                    workmode, "заменить часть")



def getDataFromXlsx(file_name_full: str, sheet_name:str,
    filter_eqv_dic={}, date_actual_rec_filter="", row_start=None):
    
    import warnings
    warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
    wb = load_workbook(file_name_full)

    warnings.filterwarnings('default', category=UserWarning, module='openpyxl')
    sheet = wb[sheet_name]

    file_name=os.path.split(file_name_full)[1]

    if row_start==None:
        dat = sheet.cell(6, 1).value
        if dat == None or dat == "None" or dat == "":
            txt1 = "Не указана стартовая строка на листе" \
                "'{sheet_name}' файла '{file_name}'."
            print(f"{bcolors.FAIL}{txt1}{bcolors.ENDC}")
            txt2 = "Нажмите любую клавишу."
            questionOneKey(bcolors.OKBLUE, txt2)
            return ["0", txt1, []]
        row_start = int(dat)

    
    row_counts = 0
    i = row_start+1
    continue_counting = True
    while continue_counting:
        dat = sheet.cell(row=i, column=1).value
        if dat == None or dat == "" or dat == "None":
            continue_counting = False
        else:
            row_counts += 1
            i += 1
    if row_counts == 0:
        return ["2", "Запрошенные данные на листе " \
                f"'{sheet_name}' файла '{file_name}' " 
                f"отсутствуют.", {}]

    key_list = []
    col_counts = 0
    i = 1
    continue_counting = True
    while continue_counting:
        dat = sheet.cell(row=row_start, column=i).value
        if dat == None or dat == "" or dat == "None":
            continue_counting = False
        else:
            col_counts += 1
            key_list.append(dat)
            i += 1
    if row_counts == 0:
        return ["2", "Запрошенные данные на листе " \
                f"'{sheet_name}' файла '{file_name}' " 
                f"отсутствуют.", {}]

    dat_list = []
    for row in range(row_start+1, row_start+row_counts+1):
        dat_row_dic = {}
        for col in range(1, col_counts+1):
            dat = sheet.cell(row=row, column=col).value
            dat_row_dic[key_list[col-1]] = dat
        dat_row_ok = True
        if len(filter_eqv_dic) > 0:
            for key in key_list:
                if key in filter_eqv_dic:
                    key_filter_list = filter_eqv_dic[key]
                    for filter_cur in key_filter_list:
                        if str(dat_row_dic[key]) == filter_cur:
                            break
                    else:
                        dat_row_ok = False
                        break
        if date_actual_rec_filter != "":
            _, _, _, date1 = checkCorrectDate(date_actual_rec_filter)
            date_from = dat_row_dic["recActualFrom"]
            date_to = dat_row_dic["recActualTo"]
            if date_from > date1 or date_to < date1:
                dat_row_ok = False
        if dat_row_ok:
            dat_list.append(dat_row_dic)

    if len(dat_list)==0:
        return ["2", "Запрошенные данные на листе " \
                f"'{sheet_name}' файла '{file_name}' " 
                f"отсутствуют.", dat_list]

    return ["1", "Словарь с данными сформирован.", dat_list]


def printDicInList(content_list):
    print()
    print(f"{bcolors.WARNING}В списке {len(content_list)} элементов:{bcolors.ENDC}")
    num=0
    for content in content_list:
        num+=1
        print (f"{bcolors.WARNING}{num}.{bcolors.ENDC} {content}")
        print("----")
    return




def addGroupingValueToData(content_list, grouping_values_list,
    filter_list=[],key_name_merged_list=[],key_name_group_list=[]):
    

    content_new_list=[]
    other_column_list=[]
    end_index = len(content_list)
    i = 0
    while i < end_index:
        rec = content_list[i]

        val_ok=True# если имеются значения для фильтрации данных
        if len(filter_list)>0:
            for filter_dic in filter_list:
                key_name_filter=filter_dic["key_name"]
                for item in filter_dic["item"]:
                    if rec[key_name_filter]==item:
                        break
                else:
                    val_ok=False
                    break
        if val_ok==True:
            for grouping_values in grouping_values_list:
                key_name_search=grouping_values['keyNameSearch']
                key_name_new = grouping_values['keyNameNew']
                val_search = rec[key_name_search]
                rec[key_name_new] = val_search
                if val_search in grouping_values:
                    val_search = grouping_values[val_search]
                    rec[key_name_new] = val_search
                else:
                    keys_list=list(grouping_values.keys())
                    for key in keys_list:
                        if grouping_values[key]=="*":
                            rec[key_name_new] = key
                            if val_search not in other_column_list:
                                other_column_list.append(val_search)
                            break
            content_new_list.append(rec)

        
        
        i += 1
    


    if len(content_new_list) == 0:
        return ["2","Список пуст.",[],[]]
    

    group_list=[]
    if len(key_name_group_list)>0:
        for rec in content_new_list:
            rec_new={}
            for key_name in key_name_group_list:
                rec_new[key_name]=rec[key_name]
            group_list.append(rec_new)
    
            
            
    group_merge_list=[]

    end_index=len(group_list)

    if len(key_name_merged_list)>0:
        group_merge_list.append(group_list[0])
        index_group_list = 1
        while index_group_list < end_index:
            rec = group_list[index_group_list]
            keys_rec_list=list(rec.keys())
            for i in range(0,len(group_merge_list)):
                rec_merge=group_merge_list[i]
                val_rec_ok=True
                for key_rec in keys_rec_list:
                    if key_rec in key_name_merged_list:
                        continue
                    if rec_merge[key_rec]==rec[key_rec]:
                        continue
                    else:
                        val_rec_ok=False
                        break
                else:
                    for key_add in key_name_merged_list:
                        rec_merge[key_add]=rec_merge[key_add]+ \
                            rec[key_add]
                    group_merge_list[i] = rec_merge
                    break
            if val_rec_ok==False:
                group_merge_list.append(rec)
            index_group_list+=1


    return ["1", "Данные подготовлены.",
            content_new_list, group_merge_list, other_column_list]



def formatToTable(content_list=[],
    column_name_list=[], filter_dict={}, 
    hide_zero_column=False, hide_zero_row=False,
    subtotal=False, total=True):
    
    content_out=[]

    if len(filter_dict)>0:
        keys_list=list(filter_dict.keys())
        for content in content_list:
            filter_content_ok=True
            for key in keys_list:
                if key not in content:
                    continue
                if content[key]!=filter_dict[key]:
                    filter_content_ok=False
                    break
            if filter_content_ok:
                content_out.append(content)
    else:
        content_out=content_list

    column_header_list=[]

    table_row_shablon_dict={}
    column_header_first=""
    for column_name in column_name_list:
        column_header=column_name["columnName"]
        column_header_list.append(column_header)
        if len(table_row_shablon_dict)==0:
            column_header_first = column_header
        table_row_shablon_dict[column_header]=None
    
    table_type_value_dict={}
    table_type_value_dict=table_row_shablon_dict.copy()
    for column_name in column_name_list:
        column_header=column_name["columnName"]
        type_value=column_name.get("typeValue",None)
        table_type_value_dict[column_header]=type_value

    table_row_total_dict=table_row_shablon_dict.copy()
    table_row_total_dict[column_header_first]="ВСЕГО:"

    table_row_subtotal_shablon_dict=table_row_shablon_dict.copy()
    table_row_subtotal_shablon_dict[column_header_first] = "Итого:"
    table_dash_dict = table_row_shablon_dict.copy()
    for column_header in column_header_list:
        table_dash_dict[column_header]="---"



    
    table_list=[]
    column_sort_dict={}
    for content in content_out:
        table_row_dict=table_row_shablon_dict.copy()
        table_row_is_ok=False
        for column_name in column_name_list:
            column_header=column_name["columnName"]
            sorting=column_name.get("sorting",None)
            if sorting!=None:
                column_sort_dict[column_header] = sorting
            column_filter_dict=column_name.get("filter",{})
            content_filter_is_ok=True
            filter_key_list=list(column_filter_dict.keys())
            if len(filter_key_list)>0:
                content_filter_is_ok = False
                for filter_key in filter_key_list:
                    if filter_key not in content:
                        break
                    filter_value_list=column_filter_dict[filter_key]
                    for filter_value in filter_value_list:
                        if content[filter_key] == filter_value:
                            content_filter_is_ok=True
                            break
            if content_filter_is_ok:
                value_key_list = column_name["valueKeyList"]
                for key in value_key_list:
                    if key in content:
                        content_value = content[key]
                        cell_value=table_row_dict[column_header]
                        if cell_value==None:
                            cell_value=content_value
                        else:
                            type_value = table_type_value_dict.get(
                                column_header, None)
                            if type_value==None:
                                type_value="str"
                                if isinstance(content_value,(float,int)):
                                    type_value="num"
                            if type_value=="num":
                                cell_value=cell_value+content_value
                            else:
                                cell_value=f"{cell_value}; {content_value}"
                        table_row_dict[column_header]=cell_value
                        table_row_is_ok=True
                        if "totalColumn" in column_name and \
                            column_name["totalColumn"]=="yes":
                            total_value = table_row_total_dict[column_header]
                            try:
                                float(cell_value)
                                if total_value=="" or total_value==None:
                                    table_row_total_dict[column_header] =cell_value
                                elif total_value!="-":
                                    table_row_total_dict[column_header] = total_value+cell_value        
                            except ValueError:
                                table_row_total_dict[column_header]="-"                            
                            
        if table_row_is_ok:
            table_list.append(table_row_dict)
    

    table_list_merge = []

    while len(table_list)>0:
        table_row_dict_control = table_list[0]
        del table_list[0]
        i=0
        while i<len(table_list):
            table_row_dict = table_list[i]
            cell_ok = True
            for column_header in column_header_list:
                cell_value = table_row_dict[column_header]
                table_type_value = table_type_value_dict.get(
                    column_header, None)
                if table_type_value == "str" or table_type_value == None:
                    if table_row_dict_control[column_header] != cell_value:
                        cell_ok = False
                        break
            if cell_ok == True:
                for column_header in column_header_list:
                    cell_value = table_row_dict[column_header]
                    table_type_value = table_type_value_dict.get(
                        column_header, None)
                    cell_value_control = table_row_dict_control[column_header]
                    if table_type_value == "num":
                        if cell_value_control==None:
                            cell_value_control=0
                        if cell_value==None:
                            cell_value=0
                        table_row_dict_control[column_header] = \
                            cell_value_control+cell_value
                del table_list[i]
                i-=1
            i+=1
        table_list_merge.append(table_row_dict_control)
    if len(table_list) > 0:
        table_list_merge.extend(table_list)
    table_list = table_list_merge


    if hide_zero_row:
        key_list = column_header_list.copy()
        for table_row_dict in table_list:
            row_is_zero = True
            for key in key_list:
                if table_row_dict[key] != None:
                    row_is_zero = False
                    break
            if row_is_zero:
                table_list.remove(table_row_dict)


    table_list.append(table_row_total_dict)


    if hide_zero_column:
        key_list=column_header_list.copy()
        for key in key_list:
            column_is_zero=True
            for table_row_dict in table_list:
                if table_row_dict[key]!=None:
                    column_is_zero=False
                    break
            if column_is_zero:
                for table_row_dict in table_list:
                    table_row_dict.pop(key,0)


    for table_row_dict in table_list:
        key_list = column_header_list.copy()
        for key in key_list:
            if table_row_dict[key]==None:
                for column_name in column_name_list:
                    if column_name["columnName"]==key:
                        type_value=column_name.get("typeValue","str")
                        if type_value=="num":
                            table_row_dict[key]=0
                            break
                        else:
                            table_row_dict[key]=""
                            break

    
    table_row_total_dict=table_list[len(table_list)-1]
    del table_list[len(table_list)-1]

    column_sort_counts = len(column_sort_dict)
    val_sort_list=[]
    if column_sort_counts > 0:
        keys_list = list(column_sort_dict.keys())
        column_sorting_list = []
        for key in keys_list:
            sorting = column_sort_dict[key]
            reverse = False
            if sorting == "up":
                reverse = True
            column_sorting_list.append(reverse)
        if column_sort_counts == 1:
            val_sort_list = sorted(table_list,key=lambda x: \
            (x[keys_list[0]]), reverse=column_sorting_list[0] )
        elif column_sort_counts == 2:
            val_sort_list = sorted(table_list,key=lambda x: \
                (x[keys_list[0]], sorted(table_list,key=lambda x: \
                (x[keys_list[1]]), reverse=column_sorting_list[1])), \
                reverse=column_sorting_list[0])
        elif column_sort_counts >= 3:
            val_sort_list = sorted(table_list,key=lambda x: \
                (x[keys_list[0]], sorted(table_list,key=lambda x: \
                (x[keys_list[1]], sorted(table_list,key=lambda x: \
                (x[keys_list[2]]), reverse=column_sorting_list[2])), \
                    reverse=column_sorting_list[1])), \
                    reverse=column_sorting_list[0])

        table_list=val_sort_list

    if total:
        table_list.append(table_dash_dict)
        table_list.append(table_row_total_dict)

    
    column_subtotal_list=[]
    if subtotal:
        table_with_subtotal_list=[]
        for column_name in column_name_list:
            column_header = column_name["columnName"]
            total_column = column_name.get("totalColumn", None)
            if total_column=="yes":
                table_row_subtotal_shablon_dict[column_header]=0
                column_subtotal_list.append(column_header)
    
    
    for column_header in column_header_list:
        if table_row_subtotal_shablon_dict[column_header]==None:
            table_row_subtotal_shablon_dict[column_header]=""
    

    if len(column_subtotal_list)>0: 
        while len(table_list)>0:
            table_row_subtotal_dict=table_row_subtotal_shablon_dict.copy()
            table_row_control_dict=table_list[0]
            table_with_subtotal_list.append(table_row_control_dict)
            del table_list[0]
            if 'ВСЕГО:' in table_row_control_dict[column_header_first] or \
                '---' in table_row_control_dict[column_header_first]:
                continue          
            for column_subtotal in column_subtotal_list:
                table_row_subtotal_dict[column_subtotal]= \
                    table_row_control_dict[column_subtotal]
            row_counts=0
            i=0
            while i<len(table_list):
                table_row_dict=table_list[i]
                if table_row_dict[column_header_first] == \
                        table_row_control_dict[column_header_first]:
                    for column_subtotal in column_subtotal_list:
                        val=table_row_subtotal_dict[column_subtotal]
                        table_row_subtotal_dict[column_subtotal]= \
                           val+table_row_dict[column_subtotal]
                    table_with_subtotal_list.append(table_row_dict)
                    del table_list[i]
                    row_counts+=1
                    i-=1
                i+=1
            if row_counts>0:
                table_with_subtotal_list.append(table_row_subtotal_dict)
                table_with_subtotal_list.append(table_dash_dict)
            else:
                table_with_subtotal_list.append(table_dash_dict) 
        if len(table_list) > 0:
            table_with_subtotal_list.extend(table_list)
        table_list = table_with_subtotal_list

    
    dash_on=False
    i=0
    while i<len(table_list):
        table_row_dict=table_list[i]
        if "---" in table_row_dict[column_header_first]:
            if dash_on==True:
                del table_list[i]
                dash_on=False
            else:
                dash_on=True
                i+=1
        else:
            dash_on=False
            i+=1

    

    return ["1","Данные сформированы.",table_list, 
        column_header_list]



def getInfoAboutEmployee(filter_eqv_dic={}, date_actual_rec_filter="",
    workmode="эксплуатация", filename_key='Employeers.xlsx',
    sheet_name = 'EmployeersList'):

    _, _, file_name = getUserFilePath(filename_key, workmode=workmode)
    if file_name == "":
        return ["0", "Не удалось найти путь к файлу Employeers.xlsx",
            []]

    res=getDataFromXlsx(file_name_full=file_name, sheet_name=sheet_name,
        filter_eqv_dic=filter_eqv_dic,
        date_actual_rec_filter=date_actual_rec_filter)
    return res


def getEmployeeList(date_filter: str, filter_eqv_dic={}, print_err="1"):
    res = getInfoAboutEmployee(date_actual_rec_filter=date_filter,
                               filter_eqv_dic=filter_eqv_dic)
    if res[0] == "0":
        txt1_1 = f"Не удалось сформировать список сотрудников."
        txt1_2 = f"{bcolors.OKBLUE}Нажмите любую клавишу.{bcolors.ENDC}"
        txt1 = f"{bcolors.WARNING}{txt1_1}{bcolors.ENDC}\n" \
            f"{txt1_2}"
        if print_err=="1":
            questionOneKey("", txt1)
        return ["0", txt1_1, []]
    employeer_list = res[2]
    if len(employeer_list) == 0:
        txt1_1 = f"Не удалось сформировать список сотрудников " \
            f"на {date_filter}."
        txt1 = f"\n{bcolors.WARNING}{txt1_1}{bcolors.ENDC}\n" \
            f"{bcolors.OKBLUE}Нажмите любую клавишу.{bcolors.ENDC}"
        if print_err=="1":
            questionOneKey("", txt1)
        return ["2", txt1_1, []]

    return ["1", "Список сформирован", employeer_list]



def getEmployerFIOList(employeer_list):
    employeer_filter_list = []
    for employeer in employeer_list:
        a_dic = {}
        a_dic['nameFull'] = employeer['nameFull']
        a_dic['nameShort'] = employeer['nameShort']
        employeer_filter_list.append(a_dic)
    return employeer_filter_list


def getShortFIO(full_FIO:str):
    employee_short=""
    employee_short_list = full_FIO.split(" ", -1)
    employee_short=employee_short_list[0]
    for i in range(1,len(employee_short_list)):
        initial=employee_short_list[i][0]
        txt_1=""
        if i==1:
            txt_1=" "
        employee_short=f"{employee_short}{txt_1}{initial}."
    return employee_short



def checkCorrectDate(dateIn: str, printErrMsg="yes"):
    date_1=re.sub("/|,",".",dateIn)
    date_1_list=date_1.split(".")
    if len(date_1_list) == 3 and len(date_1_list[2])<4:
        year_cur=toformatNow()[1][6:10]
        date_1_year=date_1_list[2]
        year_add = year_cur[0:2]
        if len(date_1_year)==1:
            year_add=f"{year_add}0"
        date_1_year = f"{year_add}{date_1_year}"
        date_1_list[2]=date_1_year
        date_1=".".join(date_1_list)

    try:
        a = datetime.strptime(date_1, '%d.%m.%Y')
        date_str=str(a)
        date_str = f"{date_str[8:10]}.{date_str[5:7]}.{date_str[0:4]}"
        return ["1", "Дата корректная.", date_str, a]
    except ValueError:
        if printErrMsg =="yes":
            err_txt1 = "Неверный формат даты. Дата должна быть в формате: " \
                "'ДД/ММ/ГГГГ' или 'ДД.ММ.ГГГГ'"
        print(f"{bcolors.WARNING}{err_txt1}{bcolors.ENDC}")
        return ["0", "Неверный формат даты.","",0]



def inputNumberDevice(number_len_list, product_group, 
        sheet_name, txt="", spec_keys=[], 
        print_number_big_font="откл", workmode="эксплуатация"):

    cicl = True
    while cicl:
        ret = ["0", "0"]
        txt1 = "Введите или отсканируйте номер устройства."
        if txt != "":
            txt1 = txt

        to_break_mark=False

        if not "/" in spec_keys:
            spec_keys.append("/")
            txt1=f"{txt1}\n{bcolors.OKBLUE}Чтобы прервать ввод " \
                f"- нажмите '/'.{bcolors.ENDC}"
            to_break_mark=True

        k=0
        for i in number_len_list:
            if k==0:
                txt1_1=str(i)
            else:
                txt1_1=txt1_1+" или "+str(i)
            k+=1
        txt2 = f"Вводимый номер должен состоять из {txt1_1} симв."
        
        oo1 = inputSpecifiedKey(bcolors.OKBLUE, txt1,
                txt2, number_len_list, spec_keys, 0, "",
                print_number_big_font, workmode)
        for i in spec_keys:
            if oo1 == i:
                if oo1=="/" and to_break_mark:
                    return ["0", "Отказ от ввода"]
                
                return ["2", oo1]
            
        device_number = oo1

        if print_number_big_font=="окно":
            a_dic={"text_2": device_number}
            saveConfigValue("print_big_font_line.json", a_dic,
                workmode, "заменить часть")

        if len(device_number)==9:
            a1=device_number[0]
            a_prod_dic={"счетчик э/э": ["1", "счетчика э/э"],
                "модуль связи": ["2", "модуля связи"],
                "клеммная колодка": ["3", "клеммной колодки"],
                "пульт управления ПУ": ["4", "пульта управления ПУ"]}
            a_dic={"1": "счетчика э/э", "2": "модуля связи",
                "3": "клеммной колодки", "4": "пульта управления ПУ"}
            a2_list=a_prod_dic.get(product_group, [])
            if len(a2_list)==0:

                if print_number_big_font=="окно":
                    a_dic={"text_2": device_number,
                        "text_3": "ERR  NUMBER"}
                    saveConfigValue("print_big_font_line.json", a_dic,
                        workmode, "заменить часть")

                a_txt=f"{bcolors.FAIL}В эталонном списке отсутствует " \
                    f"тип изделия '{product_group}'.{bcolors.ENDC}\n" \
                    f"{bcolors.OKBLUE}Нажмите Enter.{bcolors.ENDC}"
                questionSpecifiedKey("",a_txt, ["\r"], "", 1)
                print()
                return ["0", ""]
            a2=a2_list[0]
                
            if a1==a2:
                return ["1", device_number]

            txt1=f"{bcolors.WARNING}Ввели номер {a_dic[a1]}, " \
                f"а не {a_prod_dic[product_group][1]}.{bcolors.ENDC}\n"
            print(txt1)

            if print_number_big_font=="окно":
                a_dic={"text_2": device_number,"text_3": "ERR  NUMBER"}
                saveConfigValue("print_big_font_line.json", a_dic,
                    workmode, "заменить часть")

        else:
            check1 = checkDeviceType(device_number, 3, product_group, sheet_name,
                serial_number_len_list=number_len_list, serial_number_type=["decimal"], 
                print_msg="1")
            if check1[0]=="1":
                return ["1", device_number]
            
            elif check1[0] in ["2", "5"]:
                if print_number_big_font=="окно":
                    a_dic={"text_2": device_number,"text_3": "ERR  PROGRAM"}
                    saveConfigValue("print_big_font_line.json", a_dic,
                        workmode, "заменить часть")
                    
                return ["3"]
            
            if print_number_big_font=="окно":
                a_dic={"text_2": device_number,"text_3": "ERR  NUMBER"}
                saveConfigValue("print_big_font_line.json", a_dic,
                    workmode, "заменить часть")
                
                time.sleep(1)

                a_dic={"text_2": device_number}
                saveConfigValue("print_big_font_line.json", a_dic,
                    workmode, "заменить часть")
                
            printWARNING("Повторите ввод.")


def checkDeviceType(serial_number: str, id_property_device: int, type_filter: str, \
                    sheet_name: str, serial_number_len_list=[15,13], 
                    serial_number_type=["decimal"], print_msg="1",
                    workmode="эксплуатация"):

    _,res2, product_file= getUserFilePath('ProductNumber.xlsx',workmode=workmode)
    if product_file=="":
        return ["5", res2]
    if serial_number_len_list[0]==0 and len(serial_number)<1:
        txt1 = f"Число символов в номере должно не менее 1."
        if print_msg == "1":
            print(f"{bcolors.WARNING}{txt1}{bcolors.ENDC}")
        return ["3", txt1]
    else:
        txt_len_list=""
        n=0
        for i in serial_number_len_list:
            if n>0:
                txt_len_list=txt_len_list+" или " +str(i)
            else:
                txt_len_list=str(i)
            n+=1
            if i==len(serial_number):
                break
        else:
            txt1=f"Число символов в номере должно быть {txt_len_list}."
            if print_msg=="1":
                print(f"{bcolors.WARNING}{txt1}{bcolors.ENDC}")
            return ["3", txt1]

    for i in serial_number_type:
        if i=="decimal":
            txt1="Номер изделия должен состоять только из десятичных чисел."
            if serial_number.isdecimal():
                break
    else:
        if print_msg == "1":
            print(f"{bcolors.WARNING}{txt1}{bcolors.ENDC}")
        return ["4", txt1]

    res = toGetProductInfo2(serial_number, sheet_name)
    if res[0]=="1":
        val = res[id_property_device]
        if val == type_filter:
            return ["1","свойство изделия соответствует контрольному значению"]
        else:
            txt1 = f"По введенному серийному номеру найдено изделие " \
                f"'{val}', а не '{type_filter}'."
            if print_msg=="1":
                print(f"{bcolors.WARNING}{txt1}{bcolors.ENDC}")
            return ["0",txt1]
    else:
        txt1="Не удалось определить изделие по серийному номеру."
        if print_msg=="1":
            print(f"{bcolors.WARNING}{txt1}{bcolors.ENDC}")
        return ["0", txt1]


def inputEmployeeIdName(msg_txt="", employee_id="0", data_exchange_sutp="1"):
    cicl = True
    while cicl:
        if employee_id == "0":
            txt1 = "Введите свой табельный номер."
            if msg_txt != "":
                txt1 = msg_txt
            spec_keys = ["8", "*"]
            txt1 = txt1+"\nДля отказа от ввода нажмите 0.\nЧтобы выйти из программы нажмите *.\n"
            txt2 = ""
            oo = inputSpecifiedKey(
                bcolors.OKBLUE, txt1, txt2,len_input_list=[0], specified_keys_list=spec_keys)
            if oo == "0":
                return ["8", employee_id]
            elif oo == "*":
                return ["9", employee_id]
            employee_id = oo
        if data_exchange_sutp == "1":
            oo = getNameEmployee(employee_id)
            if oo[0] != "1":
                txt1_1 = f"Ошибка. Сотрудник с табельным номером {employee_id} в БД СУТП не найден.\n" + \
                    f"Попробуете еще раз? 0-нет, 1- да, 2-изменить табельный номер.\n" + \
                    "Чтобы выйти из программы нажмите 9."
                oo1 = questionSpecifiedKey(
                    bcolors.WARNING, txt1_1, ["0", "1", "2", "9"])
                if oo1 == "0":
                    return ["8", employee_id]
                elif oo1 == "9":
                    return ["9", employee_id]
                elif oo1 == "2":
                    employee_id = "0"
            else:
                employee_name = f"{oo[1]} {oo[2]} {oo[3]}"
                txt1_1 = f"Ф.И.О. сотрудника: {employee_name}. Верно? 0 - нет, 1 - да:"
                oo = questionSpecifiedKey(bcolors.OKBLUE, txt1_1, ["0", "1"])
                if oo == "1":
                    return ["1", employee_id, employee_name]
                else:
                    employee_id = "0"
        else:
            return ["2", employee_id, ""]



def questionFromListCicl(colortxt:str, explanation:str, variant_list_in:list, 
    variant_id_list_in: list, variant_id_def="", spec_list = [], 
    spec_keys = [], spec_id = [], list_keys_only = 1,
    one_column = 0, start_list_num = 0, spec_keys_hidden = [],
    cls_mode="каждый раз"):

    variant_list=variant_list_in.copy()
    variant_id_list=variant_id_list_in.copy()

    variant_id_cur = variant_id_list[0]
    if variant_id_def!=0:
        variant_id_cur = variant_id_def

    if not "\r" in spec_keys or not "\r" in spec_keys_hidden:
        spec_list.append("выбрать")
        spec_id=spec_list.copy()
        spec_keys.append("\r")

    spec_allid=spec_id+spec_keys_hidden

    cicl_count=0
    while True:
        cicl_count+=1
        if cls_mode=="каждый раз" or \
            (cls_mode=="кроме старта" and cicl_count>1):
            os.system("CLS")
        if cls_mode=="без очистки" and cicl_count>1:
            pass
        oo = questionFromList(colortxt, explanation, variant_list, 
            variant_id_list, variant_id_cur, spec_list, spec_keys, 
            spec_id, list_keys_only, one_column, 
            start_list_num, spec_keys_hidden)
        if oo=="выбрать":
            return variant_id_cur
        elif oo in spec_allid:
            return oo
        variant_id_cur=oo



def questionFromList(colortxt, txt1, list_txt, list_id, cur_id="",
    spec_list=[], spec_keys=[], spec_id=[], list_keys_only=1,
    one_column=0, start_list_num=0, spec_keys_hidden=[],
    cicl_end_key=""):

    cursor_select_list=["#вниз", "#вверх", "#вправо", "#влево"]

    exception_list = spec_list
    if cicl_end_key != "" and len(list_txt)>0:
        if cur_id==None or cur_id == "":
            cur_id = list_id[0]
        spec_keys_hidden = spec_keys_hidden+cursor_select_list
    len_list_txt = len(list_txt)
    len_spec_list = len(spec_list)
    len_all_list = len_list_txt+len_spec_list
    if len_all_list > 112:
        print(f"{bcolors.WARNING}Число элементов списка составляет: " 
            f'{len_all_list} и будет ограничено до 112.{bcolors.ENDC}')
        len_all_list = 112
        len_list_txt = 112-len_spec_list
    all_list_id = list_id[0:len_list_txt]+spec_id+spec_keys_hidden
    limit_count_symbol = 200  
    count_col = 1
    if len_all_list > 19 and one_column == 0:
        count_col = 2
        limit_count_symbol = 50  # предельное число символов в строке
    list_txt_f = list_txt[0:len_list_txt]
    spec_list_f = spec_list
    for i in range(len_list_txt):
        b = list_txt_f[i]
        a = b[0:limit_count_symbol]
        if count_col != 1:
            l1 = limit_count_symbol-len(a)
            if l1 > 0:
                a = a+" "*l1
        list_txt_f[i] = a
    for i in range(len_spec_list):
        b = spec_list_f[i]
        a = b[0:limit_count_symbol]
        if count_col != 1:
            l1 = limit_count_symbol-len(a)
            if l1 > 0:
                a = a+" "*l1
        spec_list_f[i] = a
    all_list_f = list_txt_f+spec_list_f
    count_row = int(len_all_list/count_col)
    if count_row != len_all_list/count_col:
        count_row += 1
    m1 = []
    for i in range(len_list_txt):
        m1.append(str(start_list_num+i))
    m2 = m1
    m1 = m1+spec_keys+spec_keys_hidden
    for i in spec_keys:
        ascii_rep_dic = {"\r": "\'Enter\'", "\x1b": "'ESC'"}
        if i in ascii_rep_dic:
            i = ascii_rep_dic[i]
        m2.append(i)
    printColor(txt1, colortxt)
    cicl = True
    while cicl:
        for i in range(count_row):
            colortxt1 = colortxt
            if cur_id != "" and all_list_id[i] == cur_id:
                colortxt1 = bcolors.OKRESULT  # текущее зн-е выведем черными буквами на зеленом фоне
            if count_col == 1:
                print(
                    f"{colortxt1}{m2[i]} - {all_list_f[i]}{bcolors.ENDC}", end="")
            else:
                a1 = m2[i]+" - "+all_list_f[i]+" "*4
                a = colortxt1+a1+bcolors.ENDC
                colortxt1 = colortxt
                if i+count_row <= len_all_list-1:
                    if cur_id != "" and all_list_id[i+count_row] == cur_id:
                        colortxt1 = bcolors.OKRESULT  # текущее зн-е выведем черными буквами на зеленом фоне
                    a = a+colortxt1+m2[i+count_row]+" - " + \
                        all_list_f[i+count_row]+bcolors.ENDC
                print(f"{a}", end="")
            if i < (count_row-1):
                print()
        oo = inputSpecifiedKey(colortxt, "", "", len_input_list=[0],
            specified_keys_list=m1, specified_keys_only=list_keys_only)
        cur_id_old = cur_id
        cur_id=all_list_id[m1.index(oo,0)]
        if cicl_end_key=="": 
            return cur_id
        
        cur_old_index=all_list_id.index(cur_id_old, 0)
        
        if oo == cicl_end_key:
            return cur_id_old
        elif oo in cursor_select_list:
            if oo=="#вверх":
                if cur_old_index==0:
                    cur_id=list_id[len_list_txt-1]
                else:
                    cur_id = all_list_id[cur_old_index-1]
            
            elif oo=="#вниз":
                if cur_old_index == len_list_txt-1:
                    cur_id=all_list_id[0]
                else:
                    cur_id = all_list_id[cur_old_index+1]

            elif count_col==2 and oo=="#вправо":
                a_1 = cur_old_index+count_row
                if a_1>len_list_txt:
                    a_1=len_list_txt
                cur_id = all_list_id[a_1]

            elif count_col==2 and oo=="#влево":
                a_1 = cur_old_index-count_row
                if a_1>0:
                    a_1=0
                cur_id = all_list_id[a_1]

            else:
                cur_id = cur_id_old
                
        print(f"\r \r\033[{count_row}A", end="", flush=True)



def questionOneKeyPause(colortxt,txt1,txt2,time_pause):
    if colortxt=='':
        print(f"{txt1}\nПауза {time_pause} сек...")
    else:
        print(f"{colortxt}{txt1}\n{bcolors.ENDC}Пауза {time_pause} сек...")
    time.sleep(time_pause)
    oo1=questionOneKey(colortxt, txt2)
    return oo1


def questionRetry(txt=""):
    if txt!="":
        print(f"{bcolors.WARNING}{txt} {bcolors.ENDC}")
    ret=False
    txt1="    Повторить попытку? 0-нет, 1-да"
    oo1=questionSpecifiedKey(bcolors.OKBLUE,txt1,["0","1"])
    if oo1=="1":
        ret=True
    print()
    return ret


def pause_ui (time_pause):
    print(f"Пауза {time_pause} сек...")
    time.sleep(time_pause)


def disconnectControlStatus():
    disconnect_control = GXDLMSDisconnectControl("0.0.96.3.10.255")
    a1=reader.read(disconnect_control, 1)
    a2=reader.read(disconnect_control, 2)
    a3 = reader.read(disconnect_control, 3)
    a4=reader.read(disconnect_control, 4)
    print ("физ.состояние реле (1): "+str(a1))
    print("программное состояние реле (2): "+str(a2))
    print("режим работы реле (3): "+str(a3))
    print(" (4): "+str(a4))


def print_default(filename, workmode_1="тест",output_mode="принтер"):
    txt="на печать"
    if output_mode=="экран":
        txt="для просмотра на экране"
    print(f"\n{bcolors.OKGREEN}Отправляю файл {txt}.{bcolors.ENDC}")
    printer = win32print.GetDefaultPrinter()
    PRINTER_DEFAULTS = {"DesiredAccess": win32print.PRINTER_ALL_ACCESS}
    try:
        pHandle = win32print.OpenPrinter(printer, PRINTER_DEFAULTS)
    except Exception as e:
        txt1_1 = f"При подключении к принтеру по умолчанию возникла ошибка: '{e.args[1]}'.\n"
        if e.args[0] ==5:
            txt1_1 =f"{bcolors.FAIL}Отказано в доступе к принтеру по умолчанию {printer}{bcolors.ENDC}"
        txt1_1=txt1_1+f"\n{bcolors.OKBLUE}Нажмите любую клавишу.{bcolors.ENDC}"
        print(f"{txt1_1}")
        waitOneKey()
        return 0
    properties = win32print.GetPrinter(pHandle, 2)
    properties['pDevMode'].Color = 1  # if str(_color.get()) == "Color" else 2
    properties['pDevMode'].Copies = 1
    properties['pDevMode'].Duplex = 3
    win32print.SetPrinter(pHandle, 2, properties, 0)

    if workmode_1 == "эксплуатация" or output_mode=="принтер":
        try:
            win32api.ShellExecute(0, "print", filename, None,  ".", 0)
            pause_ui(5)
            win32print.ClosePrinter(pHandle)
            return 1
        except:
            print(f"Ошибка печати")
            return 0
    else:
        txt1_1 = "Файл открыт. После окончания его проверки нажмите Enter."
        try:
            win32api.ShellExecute(0, "open", filename, None,  ".", 0)
            questionSpecifiedKey(bcolors.OKBLUE, txt1_1, ["\r"])
            return 1
        except:
            print(f"{bcolors.FAIL}Ошибка открытия файла {filename}.{bcolors.ENDC}")
            return 0


def qrmaker(qrtxt, qrfilename,scale=1):
    qrcode = segno.make_qr(qrtxt)
    qrcode.save(qrfilename,border=2, scale=scale) 


def toFillListProductModel(product_filter: list, sheet_name: str,
    workmode="эксплуатация"):
    model_list = []

    _, _, product_file = getUserFilePath('ProductNumber.xlsx',workmode=workmode)
    if product_file == "":
        return model_list
    try:
        wb = load_workbook(product_file, data_only=True)
    except Exception as e:
        txt1_1 = f"При чтении файла '{product_file}' возникла ошибка: '{e.args[1]}'.\n"
        if e.args[1] == "No such file or directory":
            txt1_1 = f"Не найден файл '{product_file}'.\n"
        txt1_1 = txt1_1+"Нажмите любую клавишу."
        print(f"{bcolors.FAIL}{txt1_1}{bcolors.ENDC}")
        waitOneKey()
        return model_list
    sheet = wb[sheet_name]
    row_start = 4
    col_model=2
    a_num_rows = sheet.max_row
    num_rows=a_num_rows
    for i in range(a_num_rows, row_start-1, -1):
        a_id = sheet.cell(i, 2).value
        a_model = sheet.cell(i, 2).value
        if (a_id==None or a_id=="") and (a_model==None or a_model==""):
            num_rows=i-1
    
    for i in range(row_start, num_rows+1):
        model1 = sheet.cell(i, col_model).value
        for j in product_filter:
            if j==model1[0:len(j):1] and (not model1 in model_list):
                model_list.append(model1)

    model_list.sort()
    if len(model_list) == 0:
        print(f"{bcolors.WARNING}  Не найдены модели изделия, содержащие " \
            f"{product_filter} в файле {product_file} лист {sheet_name} " \
            f"{bcolors.ENDC}")
    return model_list



def checkConcession(test_id: str, meter_ser_num_str: str, formula_name: str, \
        condition_in_list=[], var_in_list=[]):
    res = checkConcessionSub(test_id=test_id, meter_ser_num_str=meter_ser_num_str,
            formula_name=formula_name, condition_in_list=condition_in_list, \
            var_in_list=var_in_list)
    if res[0]=="0":
        txt1_1=f"{bcolors.FAIL}При проверке разрешения несоответствия типа ПУ произошла " \
            f"ошибка {res[1]}{bcolors.ENDC}\n"
        txt1_3=txt1_1+f"{bcolors.OKBLUE}Нажмите Enter{bcolors.ENDC}"
        spec_keys=["\r"]
        oo=questionSpecifiedKey(colortxt="",txt=txt1_3, specified_keys_in=spec_keys, file_name_mp3="", \
            specified_keys_only=1)
        return res
    else:
        return res


def checkConcessionSub(test_id: str, meter_ser_num_str: str,
    formula_name: str, condition_in_list=[], var_in_list=[],
    workmode="эксплуатация"):
    
    _, res2, checkConcession_file = getUserFilePath('CheckConcession.xlsx',
        workmode=workmode)
    if checkConcession_file == "":
        return ["0",res2]
    var_format_data_list=[]
    var_name_list=[]
    var_val_in_list=[]
    for i in var_in_list:
        var_format_data_list.append(i[0])
        var_name_list.append(i[1])
        var_val_in_list.append(i[2])
    meter_ser_num=int(meter_ser_num_str)
    sheet_name = "РазрешЗначения"
    try:
        wb = load_workbook(checkConcession_file, data_only=True)
    except Exception as e:
        txt1_1 = f"При чтении файла '{checkConcession_file}' возникла ошибка: '{e.args[1]}'.\n"
        if e.args[1] == "No such file or directory":
            txt1_1 = f"Не найден файл '{checkConcession_file}'.\n"
        print(f"{bcolors.FAIL}{txt1_1}{bcolors.ENDC}")
        return ["0", "Не удалось получить доступ к ф 'CheckResult.xlsx"]
    sheet = wb[sheet_name]
    row_start = sheet.cell(row=4, column=1).value
    if row_start == "" or row_start == "None" or row_start == None:
        txt1_1 = f"{bcolors.WARNING}Не указана стартовая строка в табл. " \
            f"файла '{checkConcession_file}' лист '{sheet_name}' для поиска " \
            f"разрешенных значений{bcolors.ENDC}"
        print(txt1_1)
        return ["0", "Не указана стартовая строка в табл. ф.CheckResult.xlsx"]
    num_rows = sheet.max_row
    starting_moment=""
    finishing_moment=""
    test_id_1=""
    type_permit=""
    starting_num_meter=0
    finishing_num_meter=0
    dt_now=datetime.now()
    for i in range(row_start, num_rows+1):
        a=sheet.cell(row=i, column=2).value
        if a==None or a=="None"or a=="":
            return ["1","Подходящее условие не найдено","4"]
        test_id_1=sheet.cell(row=i, column=9).value
        if test_id!=test_id_1:
            continue
        starting_moment=sheet.cell(row=i, column=4).value
        finishing_moment=sheet.cell(row=i, column=5).value
        a=sheet.cell(row=i, column=6).value
        if a!=None and a!="None" and a!="":
            starting_num_meter=int(a)
        a=sheet.cell(row=i, column=7).value
        if a!=None and a!="None" and a!="":
            finishing_num_meter=int(a)
        type_permit=sheet.cell(row=i, column=8).value
        var_val_list=[]
        col_start=11
        for j in range(0,len(var_in_list)):
            nam_tbl = sheet.cell(row=i, column=col_start+j*2).value
            val_tbl = sheet.cell(row=i, column=col_start+j*2+1).value
            for j1 in range(0,len(var_name_list)):
                if nam_tbl==var_name_list[j1]:
                    if var_format_data_list[j]=="целое":
                        val_tbl=int(val_tbl)
                    var_val_list.append(val_tbl)
                    break
            else:
                return ["0", "'Ошибка 1' при чтении данных из табл."]
        ret=["1","успешное сравнение","1"]
        if type_permit=="временное":
            ret=["1","успешное сравнение","2"]
        if formula_name=="ф.1":
            cond=condition_in_list[0]
            if (cond == "==" and var_val_in_list[0]==var_val_in_list[1]) or \
                (cond == "!=" and var_val_in_list[0]!=var_val_in_list[1]) or \
                (cond == "<" and var_val_in_list[0]<var_val_in_list[1]) or \
                (cond == ">" and var_val_in_list[0] > var_val_in_list[1]):
                if starting_moment > dt_now or finishing_moment < dt_now:
                    if meter_ser_num >= starting_num_meter and meter_ser_num <= finishing_num_meter:
                        return ret
                    else:
                        return ["1", "успешное сравнение", "3"]
                return ret
        elif formula_name=="ф.2":
            cond = condition_in_list[0]
            if (cond == "==" and var_val_in_list[0] == var_val_list[0] and \
                var_val_in_list[1] == var_val_list[1]) or \
                (cond == "!=" and var_val_in_list[0] != var_val_list[0] and \
                var_val_in_list[1] != var_val_list[1]) or \
                (cond == "<" and var_val_in_list[0] < var_val_list[0] and \
                var_val_in_list[1] < var_val_list[1])  or \
                (cond == ">" and var_val_in_list[0] > var_val_list[0] and \
                var_val_in_list[1] > var_val_list[1]):
                if starting_moment > dt_now or finishing_moment < dt_now:
                    if meter_ser_num >= starting_num_meter and meter_ser_num <= finishing_num_meter:
                        return ret
                    else:
                        return ["1", "успешное сравнение", "3"]
                return ret 
        else:
            return ["0", "'Ошибка 2' при чтении данных из табл."]
    ret [2]="0"
    return ret

        
def questionConcession(msg_txt:str, res_id: str, num_space: int ):
    txt_color=bcolors.WARNING
    txt1 = " "*num_space+msg_txt
    if res_id == "0" or res_id == "4":
        print(f"{bcolors.FAIL}{txt1}{bcolors.ENDC}")
        return "0"
    elif res_id=="2":
        txt1_1 = txt1+"\n"+" "*num_space+"Временно разрешенное несоответствие."
        print(f"{txt_color}{txt1_1}{bcolors.ENDC}")
        return "1"
    elif res_id=="3":
        txt1_1 = txt1+"\n"+" "*num_space+"Срок действия разрешения истек."
        print(f"{txt_color}{txt1_1}{bcolors.ENDC}")
        txt1_3=f"{bcolors.OKBLUE}Выберите действие:\n" \
            "0 - после проведения проверки отправить ПУ на ремонт\n" \
            "1 - разрешить данное несоответствие\n" \
            f"9 - сейчас прекратить проверку{bcolors.ENDC}"
        oo=questionSpecifiedKey(colortxt="",txt=txt1_3, specified_keys_in=["0","1","9"], \
            file_name_mp3="", specified_keys_only=1)
        print()
        if oo=="9":
            return "9"
        elif oo=="0":
            txt1_2 =" "*num_space+"ПУ будет отправлен в ремонт."
            print(f"{txt_color}{txt1_2}{bcolors.ENDC}")
            return "0"
        else:
            txt1_2 = " "*num_space+"Пользователь согласовал несоответствие."
            print(f"{txt_color}{txt1_2}{bcolors.ENDC}")
            txt1_1 = txt1+"\n"+txt1_2
            return "2"


def toGetProductInfo2(serial_number: str, sheet_name="Product1", 
    workmode="эксплуатация"):

    _,_, product_file = getUserFilePath('ProductNumber.xlsx',workmode=workmode)
    if product_file=="":
        return ["0"]
    if len(serial_number)==15:
        product_code = serial_number[4:8:1]
    elif len(serial_number)==13:
        product_code = serial_number[2:6:1]
    else:
        txt1_1=f"{bcolors.WARNING}Неизвестная структура серийного номера " \
            f"изделия '{serial_number}'.{bcolors.ENDC}"
        print(f"{txt1_1}")
        return ["0"]
    ret= ["0", product_code]
    try:
        wb = load_workbook(product_file, data_only=True)
    except Exception as e:
        txt1_1 = f"При чтении файла '{product_file}' возникла ошибка: '{e.args[1]}'.\n"
        if e.args[1] == "No such file or directory":
            txt1_1 = f"Не найден файл '{product_file}'.\n"
        txt1_1 = txt1_1+"Нажмите любую клавишу."
        print(f"{bcolors.FAIL}{txt1_1}{bcolors.ENDC}")
        waitOneKey()
        ret[0] = "0"
        return ret
    sheet = wb[sheet_name]
    row_start = 4
    num_rows = sheet.max_row
    for i in range(row_start, num_rows+1):
        val = sheet.cell(i, 1).value
        if val == product_code:
            number_start = sheet.cell(i, 16).value
            number_finish = sheet.cell(i, 17).value
            if len(serial_number)==13:
                number_start = f"{number_start[0:2]}{number_start[4:15]}"
                number_finish = f"{number_finish[0:2]}{number_finish[4:15]}"
            if number_start != "-" and number_finish!="-" and \
                (int(serial_number) < int(number_start) or \
                int(serial_number) > int(number_finish)):
                continue
            ret[0]="1"
            column_list=[2, 3, 4, 7, 12, 13, 14, 15, 16, 17, 11, 19, 25, 26,
                27, 28, 29, 30, 31, 32, 33, 34, 35, 36, 18, 37, 38, 39, 40, 41,
                42, 43, 44]
            i1=2
            for j in column_list:
                ret.append(sheet.cell(i, j).value)
            return ret
    if ret[0] != "1":
        printWARNING(f"\nНе найдено изделие с кодом '{product_code}' в файле " \
            f"'{product_file}' лист '{sheet_name}'.\n"
            f"Проверьте правильность серийного номера изделия: {serial_number}.")
        ret[0] = "2"
    return ret


def findDescriptionQr(qr_id: str,workmode="эксплуатация"):

    _, _, list_defects_file = getUserFilePath('QR_list.xlsx',workmode=workmode)
    if list_defects_file == "":
        return ["0"]
    workbook = load_workbook(list_defects_file)
    worksheet = workbook["QR"]
    qr_descript=""      # описание QR-кода
    qr_var_name=""      # имя перемнной
    qr_var_value=""     # зн-е для переменной
    nrows=worksheet.max_row
    row_start = worksheet.cell(row=2,column=1).value
    if row_start=="" or row_start=="None" or row_start==None:
        txt1_1=f"{bcolors.WARNING}Не указана стартовая строка в табл. " \
            f"файла '{list_defects_file}' лист 'QR' для поиска " \
            f"QR-кода{bcolors.ENDC}\n" \
            f"{bcolors.OKBLUE}Нажмите Enter{bcolors.ENDC}"
        questionOneKey(colortxt="",txt=txt1_1)
        return ["0",""]
    row_start=int(row_start)
    for i in range(row_start, nrows):
        val_eng = worksheet.cell(i, 2).value
        if val_eng=="" or val_eng=="None" or val_eng==None:
            qr_descript==""
            break
        val_rus=worksheet.cell(i, 3).value
        if val_eng==qr_id or val_rus==qr_id :
            qr_descript = worksheet.cell(i, 4).value
            qr_var_name=worksheet.cell(i, 5).value
            if qr_var_name=="None" or qr_var_name==None:
                qr_var_name=""
            qr_var_value = worksheet.cell(i, 6).value
            return ["1",qr_descript, qr_var_name, qr_var_value]
    if qr_descript=="":
        txt1_1=f"{bcolors.WARNING}Не найдено описание qr-кода {qr_id} " \
            f"в файле {list_defects_file}{bcolors.ENDC}\n" \
            f"{bcolors.OKBLUE}Нажмите Enter{bcolors.ENDC}"
        questionOneKey(colortxt="",txt=txt1_1)
        return ["0",""]



def readWriteFile(file_path: str, operation:str, content, encoding="utf-8",
    err_msg_set="1", reserve="1"):

    
    def innerFileReplacement(txt: str):

        txt1=f"{bcolors.WARNING}{txt}\n" \
            "Использовать резервный файл? (0-нет, 1-да)"
        
        oo = questionSpecifiedKey("", txt1, ["0", "1"], "", 1)
        print()
        if oo=="0":
            return "0"
        
        try:
            shutil.copy2(file_path_reserve, file_path)
            return "1"
        
        except Exception:
            txt1 = f"Ошибка при копировании файла {file_path_reserve}."
            if err_msg_set == "1":
                printWARNING(txt1)
            return "0"
    
    
    file_path_reserve=file_path+"_res"

    dir_path=os.path.split(file_path)[0]

    operation_permission_list=["w-json","r-json"]
    if operation not in operation_permission_list:
        return ["2","Такой операции нет.",""]

    if operation=="w-json" and (content=="" or \
        content==None):
        a_err_txt=f"Нет данных для записи в ф.{file_path}."
        if err_msg_set == "1":
            printFAIL(a_err_txt)
            keystrokeEnter()
            
        return ["0",a_err_txt,""]

    
    if operation=="r-json":
        attemt_count=1

        while True:
            if not os.path.isfile(file_path):
                if attemt_count<4:
                    attemt_count+=1

                    time.sleep(0.1)

                    continue

                a_err_txt=f"Не найден файл {file_path}."
                if err_msg_set == "1":
                    printFAIL(a_err_txt)
                    keystrokeEnter()

                return ["0", a_err_txt]
            
            break
        
        attemt_count=1

        while True:
            if os.path.getsize(file_path)==0:
                
                if attemt_count<4:
                    attemt_count+=1

                    time.sleep(0.1)

                    continue

                if os.path.isfile(file_path_reserve):
                    a_err_txt=f"Размер файла {file_path} =0."
                    res=innerFileReplacement(a_err_txt)
                    if res=="0":
                        return ["0", "Файл поврежден."]
                
                else:
                    return ["0", "Файл поврежден."]
            
            break

        
        while True:
            attemt_count=1

            try:
                with open(file_path, "r", errors="ignore", encoding="utf-8") as file:
                    content = json.load(file)
                
                return ["1","Операция выполнена успешно.", content]

            except Exception as e:
                if attemt_count<4:
                    attemt_count+=1

                    time.sleep(0.1)

                    continue

                err=e.args[0]
                if len(e.args)>1:
                    err=e.args[1]
                
                if err == "Permission denied" or \
                    err == "Процесс не может получить доступ к файлу, " \
                            f"так как этот файл занят другим процессом":
                    txt1_1 = f"{bcolors.FAIL}Нет доступа к файлу '{file_path}'. " \
                        f"Вероятно файл открыт. Закройте его.{bcolors.ENDC}" \
                        f"\n{bcolors.OKBLUE}Для повтора операции нажмите Enter.\n" \
                            f"Для прекращения чтения данных - нажмите 9.{bcolors.ENDC}"
                    oo = questionSpecifiedKey("", txt1_1, ["\r", "9"], "", 1)
                    if oo == "9":
                        print(f"\n{bcolors.WARNING}Операция прервана пользователем.{bcolors.ENDC}")
                        return ["0","Операция прервана пользователем",""]
                    
                elif os.path.isfile(file_path_reserve):
                    a_err_txt=f"Файл {file_path} поврежден."
                    res=innerFileReplacement(a_err_txt)
                    if res=="0":
                        return ["0", "Файл поврежден."]
                    
                else:
                    if err_msg_set == "1":
                        printFAIL(f"При чтении файла '{file_path}' возникла ошибка: "
                            f"'{err}'.")
                        keystrokeEnter()

                    return ["0", "Файл поврежден."]
                

    elif operation=="w-json":
        if not os.path.isdir(dir_path):
            res=createFolder(dir_path, "1")
            if res[0]=="0":
                return ["0", res[1]]



        attemt_count=1

        if reserve=="1":
            while True:
                try:
                    shutil.copy2(file_path, file_path_reserve)

                    break

                except Exception:
                    if attemt_count<4:
                        attemt_count+=1

                        time.sleep(0.1)

                        continue

                    if err_msg_set == "1":
                        printFAIL(f"При резервировании файла '{file_path}' возникла ошибка.")
                        keystrokeEnter()

                    return ["0", "Ошибка при резервировании файла.", ""]


        attemt_count=1

        while True:
            try:
                with open(file_path, "w", errors="ignore", encoding="utf-8") as file:
                    json.dump(content, file, ensure_ascii=False)
            
                if os.path.getsize(file_path)==0:
                    if attemt_count<4:
                        attemt_count+=1
    
                        time.sleep(0.1)

                        continue

                    else:
                        a_err_txt = "Ошибка при записи данных в конфигурационный " \
                            f" файл '{file_path}'."
                        if err_msg_set=="1":
                            printFAIL(a_err_txt)
                            keystrokeEnter()
                            
                        return ["0", a_err_txt, ""]
                
                return ["1","Операция выполнена успешно.", content] 

            except Exception as e:
                err=e.args[0]
                if len(e.args)>1:
                    err=e.args[1]
                
                if err == "Permission denied" or \
                    err == "Процесс не может получить доступ к файлу, " \
                            f"так как этот файл занят другим процессом":
                    txt1_1 = f"{bcolors.FAIL}Нет доступа к файлу '{file_path}'. " \
                        f"Вероятно файл открыт. Закройте его.{bcolors.ENDC}" \
                        f"\n{bcolors.OKBLUE}Для повтора операции нажмите Enter.\n" \
                            f"Для прекращения записи данных - нажмите 9.{bcolors.ENDC}"
                    oo = questionSpecifiedKey("", txt1_1, ["\r", "9"], "", 1)
                    if oo == "9":
                        print(f"\n{bcolors.WARNING}Операция прервана пользователем.{bcolors.ENDC}")
                        return ["0","Операция прервана пользователем",""]
                    
                txt1_1 = f"При записи в файл '{file_path}' возникла ошибка: " \
                    f"'{err}'."
                if err_msg_set=="1":
                    printFAIL(txt1_1)
                    keystrokeEnter()

                return ["0",f"Ошибка при записи файла: {err}",""]




def fileWriter(file_name, mode_write, encoding_write: str, content: str, \
               err_msg="", err_msg_set="no", dirname_sos="", join="off",
               print_on_screen="off"):
    
    if print_on_screen=="on":
        a_content=content
        l=len(a_content)
        a=a_content[l-1]
        if a=="\n":
            a_content=a_content[0:l-1]
        print (a_content)

    t1=threading.Thread(target=fileWriterThread, args=(file_name, mode_write, \
        encoding_write, content, err_msg, err_msg_set, dirname_sos,), daemon=False)
    t1.start()
    if join=="on":
        t1.join()


def fileWriterThread(file_name_full: str, mode_write: str, encoding_write, content, \
                    err_msg="", err_msg_set="on",dirname_sos=""):
    
    global bcolors_list     #список цветов у класса bcolors

    ret="0"
    
    for b_color in bcolors_list:
        content=content.replace(b_color, "")

    end_index = file_name_full.rfind("\\")
    file_name=file_name_full[end_index+1:len(file_name_full)]
    dirname_work = file_name_full[0:end_index:1]
    file_name_sos=""
    if dirname_sos!="":
        file_name_sos = os.path.join(dirname_sos, file_name)
    if encoding_write=="":
        encoding_write=None
    a = dirname_work[0:2]
    if not os.path.isdir(dirname_work) and dirname_work[0:2]=="\\\\" and dirname_sos!="":
        if not os.path.isdir(dirname_sos) and dirname_sos[0:5]=="\\\\":
            return ["0","Нет доступа к рабочей сетевой и резервной сетевой папке"]
        elif (not os.path.isdir(dirname_sos) and dirname_sos[0:5]!="\\\\") or \
           os.path.isdir(dirname_sos):
            try:
                with open(file_name_sos, mode=mode_write,errors="ignore", encoding=encoding_write) as file:
                    file.write(f"{content}")
                return ["2","Запись выполнена в резервную папку"]
            except FileNotFoundError:
                if err_msg_set=="on":
                    print(f"{bcolors.WARNING}\n{err_msg}: нет доступа к резервной папке "\
                          f"{dirname_sos}{bcolors.ENDC}")
                return ["0","Нет доступа к резервной локальной папке"]
            except Exception as e:
                if err_msg_set=="on":
                    print(f"{bcolors.WARNING}\n{err_msg}: {e.args[0]}{bcolors.ENDC}")
                return ["0","Прочие ошибки при сохранении файла в резервной локальной папке"]
    elif not os.path.isdir(dirname_work) and dirname_work[0:2] != "\\\\":
        os.makedirs(dirname_work)
    try:
        with open(file_name_full, mode=mode_write, errors="ignore", 
                  encoding=encoding_write) as file:
            file.write(f"{content}")
        return ["1","Запись выполнена в рабочую папку"]
    except Exception as e:
        if err_msg_set=="on":
            print(f"{bcolors.WARNING}\n{err_msg}: {e.args[0]}{bcolors.ENDC}")
        return ["0","Прочие ошибки при сохранении файла в рабочей локальной папке"]


def joinTestResult(join_symbol, result_txt,txt1):
    if result_txt=="":
        result_txt=result_txt+txt1
    else:
        result_txt=result_txt+join_symbol+txt1
    return result_txt


def toCopyReportFile(default_dirname, work_dirname, sos_dirname, filename):
    t1=threading.Thread(target=toCopyReportFileTread, args=(default_dirname, \
                        work_dirname, sos_dirname, filename,), daemon=False)
    t1.start()


def toCopyReportFileTread(default_dirname, work_dirname, sos_dirname, filename):
    if os.path.isdir(f"{work_dirname}\\Report"):
        try:
            shutil.copy2(f"{default_dirname}\\{filename}",
                        f"{work_dirname}\\Report\\{filename}")
            return

        except FileNotFoundError:
            if sos_dirname!="":
                if not os.path.isdir(sos_dirname):
                    try:
                        os.makedirs(sos_dirname) 
                    except Exception as e:
                        printWARNING(f"Ошибка при создании папки {sos_dirname}: {e.args[0]}")
                        return
                try:
                    shutil.copy2(f"{default_dirname}\\{filename}",
                                    f"{sos_dirname}\\{filename}")
                except Exception as e:
                    printWARNING(f"Ошибка при копировании отчета в папку {sos_dirname}: {e.args[0]}")
                    return

    if sos_dirname!="":
        if not os.path.isdir(sos_dirname):
            try:
                os.makedirs(sos_dirname) 
            except Exception as e:
                printWARNING(f"Ошибка при создании папки {sos_dirname}: {e.args[0]}")
                return
        try:
            shutil.copy2(f"{default_dirname}\\{filename}",
                            f"{sos_dirname}\\{filename}")
        
        except Exception as e:
            printWARNING(f"Ошибка при копировании отчета в папку {sos_dirname}: {e.args[0]}")

    return


def toCheckDeviceStatus(meter_serial_number,chek_status_id=7):
    res = getInfoAboutDevice(meter_serial_number)
    if res[0] == "1":
        deviceStatusId = res[6]
        if deviceStatusId != "" or deviceStatusId!=0 :
            meter_status_name = findNameMeterStatus(deviceStatusId)
            if meter_status_name[0] == "0":
                print(
                    f"{bcolors.WARNING}Ошибка при поиске имени статуса {str(deviceStatusId)} в СУТП.{bcolors.ENDC}")
            if deviceStatusId == chek_status_id:
                return ["1",meter_status_name[1],int(deviceStatusId)]
            else:
                return ["2", meter_status_name[1],int(deviceStatusId)]
        else:
            print(
                f"{bcolors.WARNING}Сервер вернул некорректный id статус ПУ из СУТП.{bcolors.ENDC}")
            return ["0", "",0]
    else:
        print(
            f"{bcolors.WARNING}Не удалось получить информацию о текущем статусе ПУ из СУТП.{bcolors.ENDC}")
        return ["0", "",0]
 


def cryptString(operation:str, str_in: str, key=""):

    from cryptography.fernet import Fernet

    if key=="":
        key = Fernet.generate_key()
    
    
    fernet = Fernet(key)

    if operation=="зашифровать":
        str_out=fernet.encrypt(str_in.encode())
        str_out=str_out.decode()
    else:
        str_out=fernet.decrypt(str_in).decode()

    return ["1", "Строка обработана.", str_out, key]



def cryptStringSec(operation:str, str_in:str):

    from libs.otkLib import toformatNow

    key_base_crypt="t3cX0Rr0-GsKohSy4Fxef-R5sgvSVJbjbWrLhtDD-oc="

    sec=str(toformatNow()[3])
    sec6=sec[-6:-1]+sec[-1]
    sec6=sec6.replace(".","5")


    if operation=="зашифровать":
        key_crypt=sec6+key_base_crypt[6:len(key_base_crypt)]
        res=cryptString("зашифровать", str_in, key_crypt)
        str_out=res[2]
        str_out=sec6+str_out
    elif operation=="расшифровать" and len(str_in)>16:
        sec6=str_in[0:6]
        str_encrypt=str_in[6:len(str_in)]
        key_crypt=sec6+key_base_crypt[6:len(key_base_crypt)]
        res=cryptString("расшифровать", str_encrypt, key_crypt)
        str_out=res[2]
    else:
        return["0", "Неизвестная операция или длина строки " \
               "меньше 17 символов.", str_in]
    return ["1", "Операция выполнена.", str_out]



def getAutoCOMPortOne(device_name: str, get_mode="0", comment_txt=""):

    cicl2=True
    while cicl2:
        if get_mode=="0":
            os.system("CLS")
        txt1=f"{bcolors.OKGREEN}Автоматическое определение COM-порта.\n" \
            f"{bcolors.OKBLUE}Проверьте, что {device_name} отключен " \
            f"от компьютера.\nЕсли {device_name} подключен, то отключите " \
            f"его.\nНажмите Enter."
        if get_mode=="1":
            txt1=f"{bcolors.OKBLUE}Отсоедините {device_name} от компьютера и " \
                f"нажмите Enter."
        if comment_txt!="":
            txt1=f"{bcolors.WARNING}{comment_txt}\n{txt1}"
        txt1=txt1+ f"\n{bcolors.OKBLUE}Для выхода - нажмите '/'."
        spec_keys=["\r","/"]
        oo=questionSpecifiedKey(colortxt="", txt=txt1, specified_keys_in=spec_keys, 
            file_name_mp3="",specified_keys_only=1)
        print()
        if oo=="/":
            return ["3", "Отказ от операции.", ""]
        pause_ui(3)

        comport_list_opto_off = comPortList()
        txt1 = f"{bcolors.OKBLUE}Подключите {device_name} к компьютеру и " \
            f"нажмите Enter.\nДля выхода нажмите '/'."
        spec_keys=["\r","/"]
        oo=questionSpecifiedKey(
            colortxt="", txt=txt1, specified_keys_in=spec_keys, 
            file_name_mp3="", specified_keys_only=1)
        if oo=="/":
            return ["3", "Отказ от операции.", ""]
        pause_ui(3)

        comport_list_opto_on = comPortList(print_list="1")

        for i in comport_list_opto_off:
            for j in comport_list_opto_on:
                if j==i:
                    comport_list_opto_on.remove(j)
                    break
        if len(comport_list_opto_on)==0 or len(comport_list_opto_on)>1:
            txt1 = f"{bcolors.WARNING}Не удалось определить COM-порт " \
                f"для оборудования.\n" \
                f"{bcolors.OKBLUE}Повторить попытку? 0-нет, 1-да."
            spec_keys = ["0", "1"]
            oo = questionSpecifiedKey(
                colortxt="", txt=txt1, specified_keys_in=spec_keys, 
                file_name_mp3="", specified_keys_only=1)
            if oo == "0":
                return ["3", "Отказ от операции.", ""]
            else:
                continue
        com_port = comport_list_opto_on[0]
        txt1 = f"{bcolors.OKGREEN}Считаем, что {device_name} подключен к " \
            f"{com_port}.\nНажмите Enter."
        spec_keys=["\r"]
        questionSpecifiedKey(
            colortxt="", txt=txt1, specified_keys_in=spec_keys, 
            file_name_mp3="", specified_keys_only=1)
        return ["1", "COM-порт определен.", com_port]
    


def getAutoCOMPort(device_name: str, get_mode="0", comment_txt="",
    window_title_list=[]):

    import keyboard     #для имитации нажатия клавиш

    
    if len(window_title_list)==0:
        a_dic={"0":"0", "1":"1", "2":"1"}
        a_get_mode=a_dic.get(get_mode, "0")
        res=getAutoCOMPortOne(device_name, a_get_mode, comment_txt)
        return res

    key_press_name=None
    
    def innerOnPressAnyKey(key):
        nonlocal key_press_name 
        
        foregroundWindow = GetWindowText(GetForegroundWindow())
    
        for window_title in window_title_list:
            if window_title in foregroundWindow:
                key_press_name = key.name
            


        return


    spec_keys=["\r","/"]

    txt1=""

    if get_mode=="0":
        os.system("CLS")
        txt1=f"{bcolors.OKGREEN}Автоматическое определение COM-порта " \
            "для одного ПУ.\n" \
            f"{bcolors.OKBLUE}Проверьте, что {device_name} отключен " \
            f"от компьютера.\nЕсли {device_name} подключен, то отключите " \
            f"его.\nПосле окончания подготовки нажмите Enter."
        
        
    elif get_mode=="1":
        txt1=f"{bcolors.OKBLUE}Отсоедините {device_name} от компьютера и " \
            f"нажмите Enter."
        

    if comment_txt!="":
        txt1=f"{bcolors.WARNING}{comment_txt}\n{txt1}"
    if txt1!="":
        txt1=txt1+ f"\n{bcolors.OKBLUE}Для выхода - нажмите '/'."
    else:
        txt1=f"{bcolors.OKBLUE}Для выхода - нажмите '/'."

    print_wait_connect=True

    if get_mode!="2":
        oo=questionSpecifiedKey(colortxt="", txt=txt1, specified_keys_in=spec_keys, 
            file_name_mp3="",specified_keys_only=1)
        print()
        if oo=="/":
            keyboard.unhook_all()
            return "3", "Отказ от операции.", ""

        pause_ui(2)
    

    comport_list_opto_off = comPortList("0")
    number_of_comport=len(comport_list_opto_off)


    if get_mode!="2":
        a_txt=f"Подключите {device_name} к компьютеру."
        printBLUE(a_txt)

    comport_list_opto_off_txt=""
    if get_mode!="1" and len(comport_list_opto_off)>1 \
        and len(comport_list_opto_off)<10:
        for i in range(0,len(comport_list_opto_off)):
            comport_list_opto_off_txt=f"{comport_list_opto_off_txt}\n" \
                f"{i+1}. {comport_list_opto_off[i]}"
        
        a_txt=f"Также можете выбрать порт из списка:" \
            f"{comport_list_opto_off_txt}"
        printBLUE(a_txt)


    keyboard.on_press(innerOnPressAnyKey)

    cicl2=True
    while cicl2:
        if key_press_name=="/":
            keyboard.unhook_all()
            return "3", "Отказ от операции.", ""

        elif get_mode!="1" and key_press_name in ["1", "2", "3", 
            "4", "5", "6", "7", "8", "9"]:
            if int(key_press_name)<=len(comport_list_opto_off):
                com_port = comport_list_opto_off[int(key_press_name)-1]

                return ["1", "COM-порт определен.", com_port]

        if print_wait_connect:
            printBLUE(f"Для выхода - нажмите '/'.")
            print_wait_connect=False
            print(f"Ожидаю подключение...")
        
        time.sleep(0.5)

        comport_list_opto_on = comPortList(print_list="0")

        if len(comport_list_opto_on)<number_of_comport:
            comport_list_opto_off=comport_list_opto_on.copy()
            number_of_comport=len(comport_list_opto_on)
            if comport_list_opto_off_txt!="":
                comport_list_opto_off_txt=""
                for i in range(0,len(comport_list_opto_off)):
                    comport_list_opto_off_txt=f"{comport_list_opto_off_txt}\n" \
                        f"{i+1}. {comport_list_opto_off[i]}"  
                a_txt=f"Список доступных портов изменился:" \
                    f"{comport_list_opto_off_txt}"
                printBLUE(a_txt)
                print(f"Ожидаю подключение...")
            continue

        for i in comport_list_opto_off:
            for j in comport_list_opto_on:
                if j==i:
                    comport_list_opto_on.remove(j)
                    break
            
        
        if len(comport_list_opto_on)==1:
            com_port = comport_list_opto_on[0]

            return ["1", "COM-порт определен.", com_port]
    


def findStrInList(txt: str, txt_list_in: list, print_msg="1", 
    mode="1", start=0, end=None):

    ret_descript={"0": "При поиске подстроки возникла ошибка.",
                  "1": "Искомая подстрока найдена.", 
                  "2": "Подстрока отсутствует"}
    
    txt_list=txt_list_in.copy()

    if mode in ["0", "1"]:
        txt=txt.replace(" ","").upper()
    if end==None:
        end = len(txt_list)-1

    for i in range(start, end+1):
        t_cur=txt_list[i]
        if mode in ["0", "1"]:
            t_cur= t_cur.replace(" ","").upper()
        if (mode in ["1", "3"] and t_cur==txt) or \
            (mode in ["0", "2"] and txt in t_cur):
            if print_msg=="1":
                print (f"\n{bcolors.WARNING}Такое значение уже " \
                    f"имеется в списке.\n"
                    f"{bcolors.OKBLUE}Нажмите любую клавишу.")
                waitOneKey()
            return ["1", ret_descript["1"], txt_list[i], i]
    return ["2", ret_descript["2"], "", -1]



def findListInStr(txt: str, txt_list_in: list, print_msg="1", 
    mode="1", start=0, end=None):

    txt_list=txt_list_in.copy()
    
    ret_descript={"0": "При поиске подстроки возникла ошибка.",
                  "1": "Подстрока найдена.", 
                  "2": "Подстрока отсутствует"}
    
    if mode in ["0", "1"]:
        txt=txt.replace(" ","").upper()

    if end==None:
        end = len(txt)

    for i in range(0,len(txt_list)):
        t_cur=txt_list[i] 
        if mode in ["0", "1"]:
            t_cur= t_cur.replace(" ","").upper()
        if (mode in ["1", "3"] and t_cur==txt[start:end]) or \
            (mode in ["0", "2"] and t_cur in txt[start:end]):
            if print_msg=="1":
                print (f"\n{bcolors.WARNING}Значение {t_cur} " \
                    f"имеется в строке.\n"
                    f"{bcolors.OKBLUE}Нажмите любую клавишу.")
                waitOneKey()
            return ["1", ret_descript["1"], txt_list[i], i]
    return ["2", ret_descript["2"], "", -1]



def appendTxtList(txt="", txt_list=[]):
    
    res=findStrInList(txt, txt_list, "0")
    ind=res[3]
    if ind==-1:
        txt_list.append(txt)
        ind=len(txt_list)-1
        return ["1","Список дополнен.", txt_list, ind]
    return ["0","Такое значение имеется в списке.", txt_list, ind]



def delTxtList(txt="", txt_list=[]):
    
    res=findStrInList(txt, txt_list, "0")
    ind=res[3]
    if ind!=-1:
        del txt_list[ind]
        return ["1","Значение удалено из списка.", txt_list]
    return ["0","Такое значение отсутствует в списке.", txt_list]
    


def delItemList(in_list: list):

    if len(in_list)>0:
        i=0
        for a in in_list:
            if a=="":
                del in_list[i]
            i+=1
    return in_list
    


def changeUserAfterWaiting(default_value_dict_in: dict, time_wait_start:float, 
    time_wait_ctrl=5, workmode="эксплуатация"):

    default_value_dict=default_value_dict_in.copy()

    a_dt=toformatNow()[3]-time_wait_start
    if a_dt<=time_wait_ctrl*60:
        return ["3", "Время ожидания было меньше или равно "
                "установленному значению.", default_value_dict]

    print(f"{bcolors.WARNING}Программа заблокирована.{bcolors.ENDC}")
    res=changeUser(workmode)
    if res[0]!="1":
        print(f"\n{bcolors.WARNING}Проверка прервана.{bcolors.ENDC}")
        return ["2", "Проверка прервана пользователем.", default_value_dict]
    
    res = readGonfigValue("opto_run.json", [], {}, workmode, "1")
    if res[0] != "1":
        return ["0", "Ошибка при чтении данных из ф.opto_run.json."]

    default_value_dict = res[2]

    return ["1", "Информация о пользователе получена успешно.",
            default_value_dict]



def changeUser(workmode="эксплуатация"):

    
    res = readGonfigValue("opto_run.json", [], {}, workmode, "1")
    if res[0] != "1":
        return ["0", "Ошибка при чтении данных из ф.opto_run.json."]

    opto_config_dic = res[2]

    data_exchange_sutp=opto_config_dic["data_exchange_sutp"]

    res = readGonfigValue("var_all_value.json", [], {}, workmode, "1")
    if res[0] != "1":
        return ["0", f"Ошибка при чтении данных из ф.var_all_value.json."]

    var_all_value_dic = res[2]

    cicl=True
    while cicl:
        txt1="Введите свой табельный номер.\n" \
            "Чтобы вернуться - нажмите '/'."
        spec_keys = ["#num>0", "/"]
        oo= inputSpecifiedKey(bcolors.OKBLUE, txt1,"",[0],spec_keys)
        if oo=="/":
            return ["2", "Отказ от ввода."]
        employee_id_new = oo
        date_cur=toformatNow()[1]
        filter_eqv_dic = {"subdivision": ["ОТК"],"employeeId":[employee_id_new]}
        res=getInfoAboutEmployee(filter_eqv_dic=filter_eqv_dic, 
            date_actual_rec_filter=date_cur, workmode=workmode)
        if res[0]=="0":
            return ["0", "Не удалось получить информацию из ф.Employeers.xlsx"]

        elif res[0]=="2":
            print (f"{bcolors.WARNING}Сотрудник с табельным номером {employee_id_new}" \
                    f" в списке 'ОТК' не найден.{bcolors.ENDC}")
            continue

        dict=res[2][0]
        pw_encrypt=dict.get("SUTPPassCrypt","")
        if pw_encrypt=="" or pw_encrypt==None:
            printWARNING(f"В таблице не указан Ваш пароль " 
                f"доступа к СУТП.\nОбратитесь к администратору "
                f"программы.")
            continue

        employees_name_new=dict.get("nameFull","")
        txt1_1 = f"Ф.И.О пользователя: {employees_name_new}. Верно? 0 - нет, 1 - да:"
        oo= questionSpecifiedKey(bcolors.OKBLUE,txt1_1,["0","1"],1)
        print()
        if oo=="0":
            continue
        
        if data_exchange_sutp=="1":
            res=getNameEmployee(employee_id_new)
            if res[0]=="0":
                a_err_txt="Не удалось получить из СУТП информацию " \
                    f"о сотруднике с табельным номером {employee_id_new}."
                printWARNING(a_err_txt)
                txt1="Нажмите Enter"
                oo1 = questionSpecifiedKey(bcolors.OKBLUE, txt1,["\r"], 1)
                return ["3", a_err_txt]

        employee_id=employee_id_new
        employees_name=employees_name_new
        employee_pw_encrypt=pw_encrypt

        config_user_new_dic={"employee_id": employee_id_new, 
            "employees_name": employees_name_new,
            "employee_pw_encrypt": pw_encrypt}

        a_filename='Employeers.xlsx'
        a_sheet_name='EmployeersList'
        
        ret_var_dic={}
        
        var_list=list(var_all_value_dic.keys())
        for var_cur in var_list:
            file_name=var_all_value_dic[var_cur]["file_name"]

            column_name = var_all_value_dic[var_cur]["column_name"]

            all_descript_dic=var_all_value_dic[var_cur]["all_value"]
            all_value_dic={}
            
            if column_name=="" or len(all_descript_dic)==0:
                continue

            a_keys_list=list(all_descript_dic.keys())
            for a_key in a_keys_list:
                all_value_dic[all_descript_dic[a_key]]=a_key

            res = readGonfigValue(file_name, [], {}, workmode, "1")
            if res[0] != "1":
                return ["0", f"Ошибка при чтении данных из ф.{file_name}."]

            act_check_dic = {}
            val_cur = res[2][var_cur]
            descript_cur=all_value_dic[val_cur]

            ret_var_dic[var_cur] = val_cur

            if column_name!=None and column_name!="":
                act_check_dic[column_name]=descript_cur

                res=checkAvailableAct(employee_id, act_check_dic, a_filename, 
                    a_sheet_name, workmode, "1")
                if res[0] == "0":
                    return ["0", res[1]]
                
                if res[0]=="2":
                    a_val=all_descript_dic[res[2][var_cur]]
                    a_dic={var_cur: a_val}
                    ret_var_dic[var_cur] = a_val
                    saveConfigValue(file_name, a_dic)
            
        saveConfigValue("opto_run.json", config_user_new_dic)

        config_new_dic={**config_user_new_dic, **ret_var_dic}
        
        return ["1", "Успешно", config_new_dic]




def getMeterConfigFilePath(meter_tech_number: str, print_err="1",
    workmode="эксплуатация", default_config="0", meter_sn=None):

    
    from libs.sutpLib import downloadFileURL
    from libs.sutpLib import getMeterConfigFileName


    def innerCmpConfigFileName(file_name_order: str):

        res=getMeterConfigFileName(meter_tech_number, print_err,
            workmode="эксплуатация")
                                   
        if res[0]=="0":
            return ["0", res[1], ""]
        
        if res[2]!=file_name_order:
            a_err_txt="Имя файла, примененного для конфигурирования ПУ " \
                f"'{res[2]}' отличается от имени файла, указанного в заказе " \
                f"'{file_name_order}'."
            return ["3", a_err_txt, res[2]]
        
        return ["1", "Имена файлов совпадают.", res[2]]


    ret_file_name_full=""
    file_name=""

    if default_config=="1":
        if meter_sn==None:
            return ["0","Для загрузки конфигурационного файла по умолчанию " 
                    "отсутствует серийный номер ПУ.", "", "", ""]
        res=toGetProductInfo2(meter_sn, "Product1", workmode)
        a_err_txt="Не удалось получить из ф.'ProductNumber.xlsx' имя " \
            "файла с конфигурацией по умолчанию."
        if res[0] in ["0", "2"]:
            return ["0", a_err_txt, "", "", ""]
        file_name=res[25]
        if file_name in [None, "None", ""]:
            return ["0", a_err_txt, "", "", ""]
        
    else:
        res=getUrlMeterConfigFile(meter_tech_number, "1", workmode)
        if res[0]=="0":
            txt1="Ошибка при получении из СУТП ссылки на файл конфигураии ПУ."
            return ["0",txt1, "", "", ""]
        url=res[2]

        sha256=res[3]
        
        file_name=os.path.split(url)[1]



    res=innerCmpConfigFileName(file_name)
    if res[0]!="1":
        return [res[0], res[1], "", file_name, res[2]]

    key_name_dic={"0": ["meterConfigLocal", "meterConfigPublic"],
                  "1": ["meterConfigDefLocal", "meterConfigDefPublic"]}
    key_name_local=key_name_dic.get(default_config, None)[0]
    key_name_public=key_name_dic.get(default_config, None)[1]

    _,_,dir_name_local = getUserFilePath(key_name_local, "1", workmode=workmode)
    if dir_name_local=="":
        return ["0",f"Ошибка при формировании пути до папки '{key_name_local}'.", 
                "", "", ""]

    file_name_local=os.path.join(dir_name_local, file_name)

    if os.path.exists(file_name_local)==True:
        res=checksumFile(file_name_local, sha256, "0")
        if res[0]=="1":
            return ["1", "Ссылка на файл конфигурации сформирована.", 
                file_name_local, file_name, file_name]
        
        elif res[0] in ["0", "4"] and print_err == "1":
            a_err_txt="При подсчете контрольной суммы файла конфигурации " \
                f"{file_name_local} возникла ошибка: {res[1]}"
            printWARNING(a_err_txt)

        os.remove(file_name_local)


    _,_,dir_name_pub = getUserFilePath(key_name_public, "1", workmode=workmode)
    if dir_name_pub=="":
        return ["0",f"Ошибка при формировании пути до папки '{key_name_public}'.",
                "", "", ""]

    file_name_pub=os.path.join(dir_name_pub, file_name)

    if os.path.exists(file_name_pub)==True:
        try:
            shutil.copy2(file_name_pub, file_name_local)
        except Exception:
            txt1 = f"Ошибка при копировании файла " \
                f"{file_name_pub}"
            if print_err == "1":
                printWARNING(txt1)
            return ["0", txt1, "", "", ""]
        
        res=checksumFile(file_name_local, sha256, "0")
        if res[0]=="1":
            return ["1", "Ссылка на файл конфигурации сформирована.", 
                file_name_local, file_name, file_name]

        elif res[0] in ["0", "4"] and print_err == "1":
            a_err_txt="При подсчете контрольной суммы файла конфигурации " \
                f"{file_name_local} возникла ошибка: {res[1]}"
            printWARNING(a_err_txt)

        os.remove(file_name_local)

    
    file_Path=file_name_local+"_tmp"

    print ("Загружаю файл конфигурации для ПУ с сервера СУТП...")
    res=downloadFileURL(url, file_Path, print_err, sha256, 3)
    if res[0]=="2":
        txt1="Ошибка контрольной суммы загруженного файла с конфигурацией для ПУ."
        if print_err == "1":
            print(f"{bcolors.FAIL}{txt1}{bcolors.ENDC}")
        os.remove(file_Path)
        return ["0",txt1, "", "", ""]
    elif res[0]!="1":
        txt1="Ошибка при загрузке с сервера СУТП файла с конфигурацией " \
            f"для ПУ."
        if print_err == "1":
            print(f"{bcolors.FAIL}{txt1}{bcolors.ENDC}")
        return ["0",txt1, "", "", ""]

    os.rename(file_Path, file_name_local)
    print('\nФайл конфигурации загружен.')

    ret_file_name_full=file_name_local
    _,_,dir_name_pub_new = getUserFilePath("meterConfigNew", "1", workmode=workmode)
    if dir_name_pub_new=="":
        txt1="Ошибка при формировании пути до папки 'meterConfigNew'."
        if print_err == "1":
            print(f"{bcolors.WARNING}{txt1}{bcolors.ENDC}")
        return ["2","Ошибка при формировании пути до папки 'meterConfigNew'.", 
                ret_file_name_full, file_name, ""]

    file_name_pub_new=os.path.join(dir_name_pub_new, file_name)

    a_index=0
    a_exten=""

    a_pos=file_name_pub_new.rfind(".")
    if a_pos!=-1:
        a_exten=f".{file_name_pub_new[a_pos+1:]}"

    while True:
        if os.path.exists(file_name_pub_new)==True:
            res=checksumFile(file_name_pub_new, sha256, "0")
            if res[0]=="1":
                return ["1", "Ссылка на файл конфигурации сформирована.", 
                    ret_file_name_full, file_name, file_name]
            
            elif res[0] in ["0", "4"] and print_err == "1":
                a_err_txt="При подсчете контрольной суммы файла конфигурации " \
                    f"{file_name_pub_new} возникла ошибка: {res[1]}"
                
                if print_err == "1":
                    printWARNING(a_err_txt)
                
                return ["2", a_err_txt, ret_file_name_full, file_name, ""]
            
            a_index+=1
            file_name_pub_new=f"{file_name_pub_new[0:a_pos]}_" \
                f"{a_index}{a_exten}"
        
        else:
            break
                

    try:
        shutil.copy2(file_name_local, file_name_pub_new)
    except Exception:
        txt1 = f"{bcolors.WARNING}Ошибка при копировании файла " \
            f"{file_name_local} в папку 'meterConfigNew'.{bcolors.ENDC}"
        if print_err == "1":
            print(txt1)
        return ["2", txt1, ret_file_name_full, file_name , ""]

    return ["1", "Ссылка на файл конфигурации сформирована.", 
        ret_file_name_full, file_name, file_name]



def getDefaultValue(var_name_list=[]):
    
    default_value_dict={
    "workmode": "эксплуатация",
    "employees_name": "",
    "employee_id": "",
    "employee_pw_encrypt": "",
    "rep_copy_public":"0",
    "speaker": "0",
    "sutp_to_save": "2",
    "modem_status": "1",
    "SIMcard_status": "0",
    "meter_color_body": "серый",
    "meter_adjusting_clock": "1",
    "meter_type_def": "i-prom.3-3-1-1/2-P-RG-Y-N",
    "modem_type_def": "MC.3-P-F",
    "actions_no_mc":"0",
    "res_ext_at_begin_test":"1",
    "order_control": "0",
    "order_control_descript": "",
    "order_num": "",
    "order_descript": "",
    "order_ev": "0",
    "data_exchange_sutp": "1",
    "meter_config_check": "0",
    "meter_config_res_list": [],
    "meter_pw_default": "1234567898765432",
    "meter_pw_default_descript": "Стандартный высокого уровня",
    "com_opto": "",
    "print_number_big_font": "откл",
    "number_of_meters": 1,
    "meter_position_cur": 0,
    "multi_com_opto_dic": {
        "caption": ["позиция № 1"," позиция № 2"],
        "com_name": ["COM3", "COM12"]
    },
    "com_rs485": "",
    "multi_com_rs485_dic": {
        "caption": [
            "позиция №1",
        ],
        "com_name": [
            "COM3",
        ]
    },
    "com_current": "com_opto",
    "com_config_opto": "",
    "multi_com_config_opto_dic": {
        "caption": [
            "позиция №1",
            "позиция №2"
        ],
        "com_name": [
            "COM3",
            "COM12"
        ]
    },
    "com_config_rs485": "",
    "multi_com_config_rs485_dic": {
        "caption": [],
        "com_name": []
    },
    "com_config_current": "com_config_opto",
    "com_config_current_select": "1",
    "com_config_user": "com_config_opto",
    "com_config_eqv_com":"1",
    "config_send_mail":"1",
    "rep_err_send_mail": "1",
    "no_data_in_SUTP_send_mail":"1",
    "get_statistic_SUTP": "1",
    "filename_rep": "",
    "meter_pw_encrypt": "",
    "meter_pw_descript": "",
    "meter_pw_level": "High",
    "meter_pw_low_encrypt": "",
    "meter_pw_low_descript": "",
    "meter_pw_high_encrypt":"",
    "meter_pw_high_descript":"",
    "meter_tech_number": "",
    "meter_tech_number_start_list":[],
    "meter_tech_number_list":[],
    "meter_status_test_list": [],
    "meter_tn_source": "",
    "meter_tn_lbl": "",
    "meter_serial_number": "",
    "meter_serial_number_list":[],
    "meter_serial_number_start_list": [],
    "meter_soft":"",
    "meter_soft_list":[],
    "meter_sn_source": "",
    "meter_sn_lbl": "",
    "meter_sn_ep":"",
    "meter_phase": "",
    "meter_voltage_dic": {},
    "meter_amperage_dic": {},
    "meter_voltage_str": "",
    "meter_amperage_str": "",
    "electrical_test_circuit":"1-0",
    "ctrl_current_electr_test": 1,
    "test_start_time":"",
    "duration_test":0,
    "rep_err_list":[],
    "clipboard_err_list":[],
    "rep_remark_list": [],
    "meter_config_res_list":[],
    "gsm_serial_number":"",
    "gsm_serial_number_list":[],
    "gsm_SIM_number":"0",
    "rc_serial_number":"",
    "rc_serial_number_list":[]
    }
     
    val_dict=default_value_dict.copy()
    
    if len(var_name_list)>0:
        for var_name in var_name_list:
            if var_name in default_value_dict:
                val_dict[var_name] = default_value_dict[var_name]
    
    return val_dict



def getAvailableActDic(filter_dic_in: dict, filter_date:str, 
    all_action_dic_in: dict, column_name: str, 
    filename_key='Employeers.xlsx', sheet_name='EmployeersList', 
    workmode="эксплуатация", msg_err_print="1"):

    filter_dic=filter_dic_in.copy()
    all_action_dic=all_action_dic_in.copy()

    if filter_date=="" or filter_date==None:
        filter_date = toformatNow()[1]

    res = getInfoAboutEmployee(filter_dic, filter_date, workmode,
        filename_key, sheet_name)
    if res[0] in ["0", "2"]:
        ret_txt='Не удалось получить данные для ' \
            f'формирования списка доступных действий.'
        if msg_err_print=="1":
            a_txt = f'{bcolors.WARNING}{ret_txt}{bcolors.ENDC}\n' \
                f'{bcolors.OKBLUE}Нажмите Enter.{bcolors.ENDC}'
            oo=questionSpecifiedKey("", a_txt, ["\r"], "", 1)
        return ["0", ret_txt, {}]

    if len(res[2])>1:
        ret_txt = 'По результату выборки число действующих записей в ' \
            f'ф.{filename_key} более 1.'
        if msg_err_print=="1":
            a_txt=f'{bcolors.WARNING}{ret_txt}{bcolors.ENDC}\n' \
                f'{bcolors.OKBLUE}Нажмите Enter.{bcolors.ENDC}'
            oo=questionSpecifiedKey("", a_txt, ["\r"], "", 1)
        return ["0", ret_txt, {}]
    a_items = res[2][0][column_name]

    if a_items=="" or a_items=="None" or a_items==None:
        ret_txt = f"Для параметра '{column_name}' в таблице " \
            f"ф.{filename_key}\nотсутствует " \
            f'перечень доступных пользователю действий.'
        if msg_err_print == "1":
            a_txt = f'{bcolors.WARNING}{ret_txt}{bcolors.ENDC}\n' \
                f'{bcolors.OKBLUE}Нажмите Enter.{bcolors.ENDC}'
            oo = questionSpecifiedKey("", a_txt, ["\r"], "", 1)
        return ["2", ret_txt, {}]
    
    action_list = a_items.split(",")

    for i in range(0, len(action_list)):
        action_list[i]=action_list[i].strip(" ")

    menu_item_dic={}
    for a_item in action_list:
        a_id=all_action_dic.get(a_item,None)
        if a_id==None:
            ret_txt = 'В общем списке доступных ' \
                f'действий отсутствует "{a_item}".'
            if msg_err_print == "1":
                a_txt = f'{bcolors.WARNING}{ret_txt}{bcolors.ENDC}\n' \
                    f'{bcolors.OKBLUE}Нажмите Enter.{bcolors.ENDC}'
                oo = questionSpecifiedKey("", a_txt, ["\r"], "", 1)
            return ["2", ret_txt, {}]
        menu_item_dic[a_item]=a_id
    
    return ["1", "Словарь с действиями сформирован.", menu_item_dic]



def checkAvailableAct(employee_id:str, act_dic_in: dict,
    filename_key='Employeers.xlsx', sheet_name='EmployeersList', 
    workmode="эксплуатация", msg_err_print="1"):

    act_dic=dictCopy(act_dic_in)

    date_cur = toformatNow()[1]

    a_filter_dic={"employeeId": [employee_id]}

    res = getInfoAboutEmployee(a_filter_dic, date_cur, workmode,
        filename_key, sheet_name)
    if res[0] in ["0", "2"]:
        ret_txt='Не удалось получить данные для ' \
            f'формирования списка доступных пользователю действий.'
        if msg_err_print=="1":
            a_txt = f'{bcolors.WARNING}{ret_txt}{bcolors.ENDC}\n' \
                f'{bcolors.OKBLUE}Нажмите Enter.{bcolors.ENDC}'
            oo=questionSpecifiedKey("", a_txt, ["\r"], "", 1)
        return ["0", ret_txt, {}]

    if len(res[2])>1:
        ret_txt = 'По результату выборки для пользователя число ' \
            f'действующих записей в ф.{filename_key} более 1.'
        if msg_err_print=="1":
            a_txt=f'{bcolors.WARNING}{ret_txt}{bcolors.ENDC}\n' \
                f'{bcolors.OKBLUE}Нажмите Enter.{bcolors.ENDC}'
            oo=questionSpecifiedKey("", a_txt, ["\r"], "", 1)
        return ["0", ret_txt, {}]
    
    ret_id="1"
    ret_txt="Словарь с допустимыми действиями остался без изменений."

    user_xls_dic=res[2][0]
    column_name_list=list(act_dic.keys())
    for column_name in column_name_list:
        user_act=user_xls_dic.get(column_name, None) 
        if user_act==None:
            ret_txt = f'Для пользователя с табельным номером ' \
                f'{employee_id} у параметра "{column_name}" в ' \
                f'ф.{filename_key} отсутствуют допустимые значения.'
            if msg_err_print == "1":
                a_txt = f'{bcolors.WARNING}{ret_txt}{bcolors.ENDC}\n' \
                    f'{bcolors.OKBLUE}Нажмите Enter.{bcolors.ENDC}'
                oo = questionSpecifiedKey("", a_txt, ["\r"], "", 1)
            return ["0", ret_txt, {}]
        
        user_act_list = user_act.split(",")

        act_list=act_dic[column_name].split(",")
        for act_cur in act_list:
            if act_cur in user_act_list:
                return [ret_id, ret_txt, act_dic]
            
        a_default=user_act_list[0]
        
        act_dic[column_name] = a_default
        ret_id="2"
        ret_txt="В словарь с допустимыми действиями внесли изменения."

        return [ret_id, ret_txt, act_dic]



def checksumFile(file_path: str, sha256: str, print_err_msg="1"):


    import hashlib

    if not os.path.exists(file_path):
        a_txt=f"Файл {file_path} отсутствует."
        if print_err_msg=="1":
            printFAIL(a_txt)
        return ["4", a_txt, None]

    try:
        block_size=4096
        with open(file_path, 'rb') as rf:
            h = hashlib.sha256()
            for chunk in iter(lambda: rf.read(block_size), b''):
                h.update(chunk)
        a_checksum=h.hexdigest()

    except Exception as e:
        a_txt=f"При подсчете контрольной суммы файла '{file_path}' " \
            f"возникла ошибка: {e}"
        if print_err_msg=="1":
            print(f"{bcolors.FAIL}{a_txt}{bcolors.ENDC}")
        return ["0", a_txt, None]  


    if sha256!=None and sha256!="":
        if a_checksum!=sha256:
            a_file_name=os.path.split(file_path)[1]
            
            a_txt=f"Контрольная сумма файла '{a_file_name}' равна\n" \
                f"{h.hexdigest()}\n" \
                f"и отличается от эталонного значения\n{sha256}."
            if print_err_msg=="1":
                printFAIL(a_txt)
            return["2", a_txt, a_checksum]
    
    return["1", "Контрольная сумма файла совпала.", a_checksum]


def openFileWaitKey(file_path: str, window_title="", txt="", check_close_file="0", wait_key="\r", print_err_msg="1",
                    time_wait_exec=2):
    """
    Функция принимает путь к файлу и флаги настроек,

    функция открывает указанную программу и передает фокус ввода на нее.
    """
    if txt!="":
        printColor(txt)

    file_keys=""
    file_path_start=file_path

    a_path = os.path.split(file_path)[0]# путь к файлу
    filename = os.path.split(file_path)[1]# название файла

    if a_path=="":
        a_file_name=filename

        if " " in a_file_name:
            a_pos=a_file_name.find(" ")
            filename=a_file_name[0:a_pos]
            file_keys=a_file_name[a_pos:len(a_file_name)]# что это такое - я не знаю

        _, ans2, file_path = getUserFilePath(filename)
        if file_path == "":
            return ["0", ans2]
        
        file_path_start=f"{file_path}{file_keys}"

    key_descript = wait_key
    if wait_key!="" and wait_key!=None:
        a_dic={"\r":"Enter", "\t":"Tab", "\b":"Backspace",
            "\x1b":"ESC"}
        if wait_key in a_dic:
            key_descript=a_dic[wait_key]

        txt1 = f"Затем перейдите в настоящую " \
            f"программу и нажмите '{key_descript}'."
        printColor(txt1, bcolors.OKBLUE)
    
        txt1="Сейчас для запуска программы нажмите " \
            "Enter.\nПо окончании выполнения программы " \
            "- вернитесь в текущую программу и " \
            f"нажмите '{key_descript}'."
        questionSpecifiedKey(bcolors.OKBLUE, txt1, ["\r"], "", 1)
    
    txt1=f"Пожалуйста подождите, идет " \
        f"запуск программы..."
    printGREEN(txt1)


    subproc_hwnd = None
    pid=None
    progr_start_ok=False

    if window_title!="":
        res = searchTitleWindow(window_title)
        if res[0] == "1":
            subproc_hwnd = res[2]

            res=actionsSelectedtWindow([], subproc_hwnd,
                "показать")
            if res[0] == "0":
                a_err_txt=f"Ошибка при отображении " \
                    f"окна ф.'{filename}'."
                printMsgWait(a_err_txt, bcolors.WARNING, 
                    print_err_msg )
                
                res=actionsSelectedtWindow([], subproc_hwnd,
                    "закрыть")
                if res[0]=="0":
                    a_err_txt=f"Ошибка при закрытии окна " \
                        f"окна ф.'{filename}'."
                    printMsgWait(a_err_txt, bcolors.WARNING, 
                        print_err_msg )

                    return ["0", a_err_txt]

            progr_start_ok=True


    if not progr_start_ok:
        try:
            cmd_txt=f"start {file_path_start}" # Запускает массовый конфигуратор через командную строку с параметрами: логин+пароль+техномер
            if filename.rfind(".py")!=-1:
                cmd_txt=f"python {file_path_start}"
            subprocess.Popen(cmd_txt, shell=True)
            time.sleep(1)

        except Exception as e:
            txt1_1 = f"{bcolors.FAIL}Ошибка открытия файла {file_path}: "\
                f"{e.strerror}.{bcolors.ENDC}\n" \
                f"{bcolors.OKBLUE}Нажмите Enter.{bcolors.ENDC}"
            questionSpecifiedKey("", txt1_1, ["\r"])
            return [0, f"Ошибка открытия файла {file_path}.", 
                None, None]
        
        if time_wait_exec!=-1:
            pid = getPidProcess(filename, time_wait_exec)
            if pid==None:
                txt1_1 = f"{bcolors.FAIL}Ошибка открытия файла {file_path}. "\
                    f"{bcolors.ENDC}\n" \
                    f"{bcolors.OKBLUE}Нажмите Enter.{bcolors.ENDC}"
                questionSpecifiedKey("", txt1_1, ["\r"])
                return [0, f"Ошибка открытия файла {filename}.", 
                None, None]

            res = getHwndPid(pid)
            subproc_hwnd = res[2]
        
    if wait_key!="" and wait_key!=None:
        txt1_1 = f"\nОжидаю нажатия клавиши {key_descript}."
        questionSpecifiedKey(bcolors.OKBLUE, txt1_1, wait_key,
            "", 1)

    if check_close_file=="1":
        res = checkFileIsClosed(file_path)
        if res[0] == "0":
            return ["0", f"Ошибка при проверке доступности "
                f"ф.'{file_path}'", None,  None]
        
    return ["1","Операция выполнена успешно.", pid, subproc_hwnd]


def checkFileIsClosed(file_path: str):

    cicl=True
    while cicl:
        try:
            os.rename(file_path, file_path)
            return ["1", "Файл доступен."]
        except Exception as e:
            a_err_txt = f"При чтении файла '{file_path}' возникла ошибка: '{e.args[1]}'."
            if e.args[1] == "No such file or directory":
                a_err_txt = f"В текущей папке отсутствует файл '{file_path}'."

            elif e.args[1] == "Permission denied" or \
                    e.args[1] == "Процесс не может получить доступ к файлу, " \
                        "так как этот файл занят другим процессом":
                txt1_1 = f'{bcolors.FAIL}Нет доступа к файлу "{file_path}".{bcolors.ENDC}\n' \
                    f'{bcolors.FAIL}Вероятно файл открыт. Закройте его.{bcolors.ENDC}\n' \
                    f'{bcolors.OKBLUE}Для повтора операции нажмите Enter.{bcolors.ENDC}\n' \
                    f'{bcolors.OKBLUE}Для отказа - нажмите "/".{bcolors.ENDC}'
                oo = questionSpecifiedKey("", txt1_1, ["\r", "/"])
                if oo == "/":
                    return ["9", "Пользователь отказался закрыть файл."]
                continue

            print(f"{bcolors.FAIL}{a_err_txt}{bcolors.ENDC}\n" 
                  f"{bcolors.OKBLUE}Нажмите любую клавишу.{bcolors.ENDC}")
            waitOneKey()
            return ["0", a_err_txt]
        


def checkVarProgrAvailable(var_name: str, var_value, msg_err_print="1"):

    available_dic={"sutp_to_save":  
        ["==", ["0", "01", "02", "03", "1", "2", "3", None]]}

    if var_name not in available_dic:
        err_txt=f"Переменная '{var_name}' отсутствует в списке " \
            f"контролируемых значений."
        if msg_err_print=="1":
            a_err_screen_txt=f"{bcolors.FAIL}Ошибка при проверке " \
                f"корректности значения переменной. {err_txt}" \
                f"{bcolors.ENDC}\n{bcolors.OKBLUE}Нажмите Enter." \
                f"{bcolors.ENDC}"
            questionSpecifiedKey("", a_err_screen_txt, ["\r"], "", 1)
        return ["0", err_txt]

    verification_method= available_dic[var_name][0]
    available_value_list = available_dic[var_name][1]

    err_txt=None
    if verification_method=="==" and (var_value not in 
        available_value_list):
        a_var_value=var_value
        if isinstance(var_value,str):
            a_var_value=f"'{var_value}'"
        
        a_value_txt=None
        for a_value_cur in available_value_list:
            a_value=a_value_cur
            if isinstance(a_value_cur, str):
                a_value=f"'{a_value}'"
        
            if a_value_txt==None:
                a_value_txt=a_value
            
            else:
                a_value_txt=f"{a_value_txt}, {a_value}"

        err_txt = f"Текущее значение переменной {var_name}=" \
            f"{a_var_value} отсутствует в списке допустимых " \
            f"значений: {a_value_txt}."
    
    if err_txt!=None:
        if msg_err_print == "1":
            a_err_screen_txt=f"{bcolors.FAIL}Ошибка при проверке " \
                f"корректности значения переменной. {err_txt}" \
                f"{bcolors.ENDC}\n{bcolors.OKBLUE}Нажмите Enter." \
                f"{bcolors.ENDC}"
            questionSpecifiedKey("", a_err_screen_txt, ["\r"], "", 1)
        return ["2", err_txt]
    
    else:
        return ["1", "Значение переменной имеется в списке."]



def findStrInFile(sub_str:str, file_path:str, mode="0", 
    msg_err_print="1", encoding='utf-8'):

    ret_descript={"0": "При поиске подстроки в файле возникла ошибка.",
                  "1": "Искомая подстрока найдена.", 
                  "2": "Подстрока отсутствует"}

    line_list=[]
    err_txt=""

    size = os.path.getsize(file_path)
    
    if size > 102400:
        err_txt = f"Размер файла '{file_path}' более 100 кБ."
    elif size==0:
        err_txt = f"Файл '{file_path}' пуст."
        
    if err_txt!="":
        if msg_err_print=="1":
            print (f"{bcolors.WARNING}{err_txt}{bcolors.ENDC}")
        return["0", err_txt, "", -1]
    
    try:
        with open(file_path, 'r', errors="ignore", encoding=encoding) as fp:
            for line in fp:
                line = line.rstrip('\n')
                line_list.append(line)
    except Exception as e:
        err_txt=f"При открытии ф.'{file_path}' для поиска подстроки " \
            f"возникла ошибка: {e}."
        if msg_err_print=="1":
            print(f"{bcolors.WARNING}{err_txt}{bcolors.ENDC}")
        return["0", err_txt, "", -1]
    
    res=findStrInList(sub_str, line_list, print_msg="0", 
        mode=mode, start=0, end=None)
    if res[0]=="2":
        return ["2", ret_descript["2"], "", -1]
    
    elif res[0]=="1":
        a_txt=res[2]
        a_index=res[3]
        return ["1", ret_descript["1"], a_txt, a_index]
    
    else:
        err_txt=f"Ошибка при поиске подстроки в файле " \
            f"'{file_path}'."
        if msg_err_print=="1":
            print(f"{bcolors.WARNING}{err_txt}{bcolors.ENDC}")
        return["0", err_txt, "", -1]
    


def cmpVers(vers1: str, vers2:str):

    vers1_list=vers1.split(".")
    vers2_list=vers2.split(".")

    len_max=len(vers1_list)
    if len(vers2_list)>len_max:
        len_max=len(vers2_list)
        for i in range(len(vers1_list), len_max+1):
            vers1_list.append('0')
    
    if len(vers2_list)<len_max:
        for i in range(len(vers2_list), len_max+1):
            vers2_list.append('0')

    for i in range(0, len(vers1_list)):
        if int(vers1_list[i])==int(vers2_list[i]):
            continue

        elif int(vers1_list[i])<int(vers2_list[i]):
            return "<"
        
        else:
            return ">"
        
    return "="



def insHiphenColor(data_list_in: list, ins_str="- ", 
    color_str=bcolors.FAIL):

    data_list=listCopy(data_list_in)

    for i in range(0, len(data_list)):
        if color_str!=None and color_str!="":
            data_list[i] = f"{color_str}{ins_str}" \
                f"{data_list[i]}{bcolors.ENDC}"
        
        else:
            data_list[i] = f"{ins_str}{data_list[i]}"

    return ["1", "Список успешно обработан.", data_list]



def setWinreg(reg_path, name, value):

    import winreg

    try:
        winreg.CreateKey(winreg.HKEY_CURRENT_USER, reg_path)
        registry_key = winreg.OpenKey(winreg.HKEY_CURRENT_USER, reg_path, 0, 
                                       winreg.KEY_WRITE)
        winreg.SetValueEx(registry_key, name, 0, winreg.REG_SZ, value)
        winreg.CloseKey(registry_key)
        return True
    except WindowsError:
        return False


def getWinReg(reg_path, name):

    import winreg

    try:
        registry_key = winreg.OpenKey(winreg.HKEY_CURRENT_USER, reg_path, 0,
                                       winreg.KEY_READ)
        value, regtype = winreg.QueryValueEx(registry_key, name)
        winreg.CloseKey(registry_key)
        return value
    except WindowsError:
        return None



def delKeyWinReg(reg_path, name):
    """
       Что пытается сделать код:
    - Открыть раздел реестра по пути reg_path в HKEY_CURRENT_USER.

    - Удалить в нём подраздел с именем name.

    - При успехе вернуть True, при ошибке — False.
    """

    import winreg

    try:
        registry_key = winreg.OpenKey(winreg.HKEY_CURRENT_USER, reg_path, 0,
                                       winreg.KEY_READ)
        winreg.DeleteKeyEx(registry_key, name)
        winreg.CloseKey(registry_key)
        return True
    except WindowsError:
        return False



def subKeysWinReg(reg_path):
    """
    Инициализирует пустой список win_key_list для хранения имён подключей.

    Пытается открыть раздел реестра по пути reg_path в ветке HKEY_CURRENT_USER.

    Получает количество подключей через winreg.QueryInfoKey(h_apps)[0].

    В цикле перебирает индексы от 0 до количество_подключей − 1 и:

    вызывает winreg.EnumKey(h_apps, idx) для получения имени под‑ключа по индексу;

    добавляет имя в список win_key_list.

    При ошибке (WindowsError) формирует сообщение об ошибке и возвращает кортеж с кодом "0" и текстом ошибки.

    При успешном выполнении возвращает кортеж с кодом "1", сообщением и списком подключей.
    """

    import winreg
    
    win_key_list=[]
    try:
        with winreg.OpenKey(winreg.HKEY_CURRENT_USER, f"{reg_path}\\") as h_apps:
            for idx in range(winreg.QueryInfoKey(h_apps)[0]):
                win_key_list.append(winreg.EnumKey(h_apps, idx))# возвращает имя под‑ключа по заданному числовому индексу (idx) в открытом разделе реестра (h_apps)

    except WindowsError:
        a_err_txt="Ошибка при получении списка вложенных ключей " \
            "из реестра Windows."
        return ["0", a_err_txt]

    return ["1", "Список сформирован.", win_key_list]



def checkChangeKayboardLayout(control_layout_id="00000409",
    print_err_msg="1"):
    
    REG_PATH="Keyboard Layout\\Preload"
    reg_name="1"
    layout_id_cur=getWinReg(REG_PATH, reg_name)
    if layout_id_cur==None:
        a_err_txt="При получении данных об установленной раскладке " \
            "клавиатуры по умолчанию возникла ошибка."
        a_txt=a_err_txt+f"{bcolors.OKBLUE}Нажмите клавишу Enter."
        printMsgWait(a_txt, bcolors.WARNING, print_err_msg, 
            ["\r"])
        return ["0", a_err_txt, None]
    
    if layout_id_cur!=control_layout_id:
        win32api.LoadKeyboardLayout(control_layout_id,1)

    return ["1","Операция выполнена успешно", layout_id_cur]



def printMsgWait(msg_txt: str, msg_color: str, 
    print_msg="1", wait_key_list=None):

    if msg_txt==None or msg_txt=="":
        return

    printColor(msg_txt, msg_color)

    if wait_key_list!=None and len(wait_key_list)>0:
        questionSpecifiedKey(colortxt="", txt="", 
            specified_keys_in=wait_key_list, specified_keys_only=1)
    
    return



def insColorCodeStr(txt:str, color_def=None):

    global bcolors_list     #список цветов у класса bcolors

    txt_list = txt.split("\n")

    color_cur=None

    if color_def!=None:
        color_cur=color_def

    for i in range(0, len(txt_list)):
        txt_cur = txt_list[i]

        txt_cur = txt_cur.replace("\n", "")
        
        if txt_cur=="":
            continue

        j=0
        
        for b_color in bcolors_list:
            if txt_cur[0:(len(b_color))]==b_color:
                break
        
        else:
            if color_cur!=None:
                txt_cur=color_cur+txt_cur
        
        while j< len(txt_cur):
            for b_color in bcolors_list:
                if txt_cur[j:(j+len(b_color))]==b_color:
                    if b_color==bcolors.ENDC:
                        color_cur=None
                        if color_def!=None:
                            color_cur=color_def
                            if j+len(b_color)<len(txt_cur):
                                txt_cur=txt_cur[0:(j+len(b_color))]+color_cur+ \
                                    txt_cur[(j+len(b_color)):]
                                j=j+len(color_cur)-1
                        
                        j=j+len(b_color)-1
                        break

                    if color_cur!=None and color_cur!=b_color:
                        txt_cur=txt_cur[0:j+1]+bcolors.ENDC+txt_cur[j:]
                        color_cur=b_color
                        j=j+len(bcolors.ENDC)+len(b_color)-1
                        break

                    else:
                        color_cur=b_color
                        j=j+len(b_color)-1
                        break


            j+=1
        
        if txt_cur[-(len(bcolors.ENDC))]!=len(bcolors.ENDC):
            txt_cur=txt_cur+bcolors.ENDC

        txt_list[i]=txt_cur

    txt = "\n".join(txt_list) 

    return txt         




def printColor(txt:str, color_def=None):

    txt=insColorCodeStr(txt, color_def)
    
    print (txt)



def printWARNING(txt:str):

    printColor(txt, bcolors.WARNING)

    return



def printGREEN(txt:str):

    printColor(txt, bcolors.OKGREEN)

    return



def printBLUE(txt:str):

    printColor(txt, bcolors.OKBLUE)

    return



def printFAIL(txt:str):

    printColor(txt, bcolors.FAIL)

    return



def searchImageWindow(image_path:str, window_title:str, 
    print_msg_err="1", search_mode="0"):



    res=searchTitleWindow(window_title)
    if res[0]=="0":
        a_err_txt = f"При поиске окна с заголовком " \
            f"{window_title} возникла ошибка."
        printMsgWait(a_err_txt, bcolors.WARNING, 
            print_msg_err)
        return ["0", a_err_txt]
    
    elif res[0]=="2":
        return["2", res[1]]
    
    hwnd=res[2]

    try:
        windowPosition = win32gui.GetWindowRect(hwnd)
    except Exception:
        a_err_txt = f"При поиске окна с заголовком " \
            f"{window_title} возникла ошибка."
        printMsgWait(a_err_txt, bcolors.WARNING, 
            print_msg_err)
        return ["0", a_err_txt]

    window_pos=(windowPosition[0], windowPosition[1])
    window_wh=(windowPosition[2], windowPosition[3])

    if search_mode=="1":
        return ["1", "Информация об окне сформирована." ,
            window_pos, None, None, 
            window_wh]

    bullseye = PIL.Image.open(image_path)
    image = PIL.ImageGrab.grab(windowPosition)

    try:
        location=pyscreeze.locate(bullseye, image, grayscale=True)

    except Exception:
        a_err_txt = f"Изображение не найдено."
        printMsgWait(a_err_txt, bcolors.WARNING, 
            print_msg_err)
        return ["3", a_err_txt]

    location_pos=(location[0], location[1])
    bullseye_wh=(bullseye.width, bullseye.height)

    return ["1", "Изображение найдено.", window_pos,
            location_pos, bullseye_wh, window_wh]

        

def mouseClickGraf(x:int, y:int, print_msg_err="1"):

    import pyautogui as pyautogui

    try:
        pyautogui.click(x, y)
    except Exception as e:
        a_err_txt = f"При имитации нажатия клавиши мыши " \
            f"возникла ошибка."
        printMsgWait(a_err_txt, bcolors.WARNING, 
            print_msg_err)
        return ["0", a_err_txt]

    return ["1", "Операция выполнена успешно."]



def menuSelectActions(menu_item_main_list_in: list, 
    menu_id_main_list_in: list, header_user=None, 
    menu_item_add_list=[], menu_id_add_list=[], 
    interrupt_list=["Прервать проверку", "/"]):
    
    menu_item_main_list=menu_item_main_list_in.copy()
    menu_id_main_list=menu_id_main_list_in.copy()

    while True:
        header = "Выберите дальнейшее действие:"
        if header_user!=None:
            header = header_user

        menu_item_list=menu_item_main_list.copy()
        menu_id_list=menu_id_main_list.copy()

        if len(menu_item_add_list) > 0 and len(menu_id_add_list) > 0:
            menu_item_list.extend(menu_item_add_list)
            for i in range(0, len(menu_id_add_list)):
                menu_id_add_list[i] = f"#add-{menu_id_add_list[i]}"
            menu_id_list.extend(menu_id_add_list)
        spec_list=[]
        spec_keys=[]
        spec_id_list=[]
        if len(interrupt_list)>0:
            spec_list = [interrupt_list[0]]
            spec_keys = [interrupt_list[1]]
            spec_id_list = ["прервать"]
        oo = questionFromList(bcolors.OKBLUE, header, menu_item_list,
            menu_id_list, "", spec_list, spec_keys, spec_id_list, 
            1, start_list_num=1)
        print()
        if oo == "прервать":
            return ["9", interrupt_list[0]]

        elif "#add-" in oo:
            a_index = menu_id_add_list.index(oo)
            oo = oo.replace("#add-", "")
            return ["2",menu_item_add_list[a_index], oo]

        elif oo in menu_id_list:
            a_index = menu_id_list.index(oo)
            return ["1", menu_item_list[a_index], oo]

        

def menuSimple(color: str, header: str, 
    menu_item_list_in: list, menu_id_list_in: list,
    id_cur=None, menu_spec_item_list=[], menu_spec_id_list=[],
    menu_spec_keys_list=[]):

    spec_list = ["ok", "отмена"]
    spec_keys = ["\r", "/"]
    spec_id = spec_list

    menu_item_list=menu_item_list_in.copy()
    menu_id_list=menu_id_list_in.copy()

    if len(menu_spec_item_list)>0:
        spec_list.extend(menu_spec_item_list)
        spec_keys.extend(menu_spec_keys_list)
        spec_id.extend(menu_spec_id_list)

    id_cur_old=id_cur

    while True:
        os.system("CLS")
        oo = questionFromList(color, header, menu_item_list, 
            menu_id_list, id_cur, spec_list, spec_keys, spec_id, 
            1, 1, 1, [], "")
        if oo == "ok":
            if id_cur!=id_cur_old:
                return ["1", "Значение изменилось.", id_cur]
            
            else:
                return ["2", "Значение осталось прежним.", id_cur_old]
            
        elif oo == "отмена":
            return ["2", "Значение осталось прежним.", id_cur_old]
        
        elif oo in menu_spec_keys_list:
            return ["1", "Нажали спец. клавишу.", oo]

        else:
            id_cur=oo

    

def menuChangeValue(var_name_list_in: list, employee_id: str, 
    all_value_file_name="var_all_value.json", header=None, 
    color=None, workmode="эксплуатация"):

    var_name_list=var_name_list_in.copy()
    
    res = readGonfigValue(all_value_file_name, [], {}, workmode, "1")
    if res[0] != "1":
        return ["0", f"При чтении ф.{all_value_file_name} "
            "возникла ошибка."]

    var_all_value_dic = res[2]
    
    val_cur_dic={}

    descript_dic={}

    val_def_dic={}

    val_standart_dic={}

    column_name_dic={}

    config_file_name_dic={}

    var_type_dic={}

    for a_name_cur in var_name_list:
        a_val_cur_dic=var_all_value_dic[a_name_cur]

        config_file_name_dic[a_name_cur]=a_val_cur_dic["file_name"]
        
        res = readGonfigValue(config_file_name_dic[a_name_cur], 
            [], {}, workmode, "1")
        if res[0] != "1":
            return ["0", f"При чтении ф.{config_file_name_dic[a_name_cur]} "
                "возникла ошибка."]

        val_cur_dic[a_name_cur]=res[2][a_name_cur]

        descript_dic[a_name_cur] = a_val_cur_dic["descript"]

        val_def_dic[a_name_cur]=a_val_cur_dic["all_value"]

        val_standart_dic[a_name_cur]=a_val_cur_dic["value_standart"]

        column_name_dic[a_name_cur]=a_val_cur_dic["column_name"]

        var_type_dic[a_name_cur]=a_val_cur_dic.get("type","")


    ret_dic=val_cur_dic.copy()
    
    menu_item_list=[]
    menu_id_list=[]

    vars_list=list(val_cur_dic.keys())

    date_cur = toformatNow()[1]

    val_def_rev_dic={}
    for var_cur in vars_list:
        a_dic=val_def_dic[var_cur]
        a_list=list(a_dic.keys())
        a_rev_dic={}
        for a_descr in a_list:
            a_rev_dic[a_dic[a_descr]]=a_descr
        val_def_rev_dic[var_cur]=a_rev_dic


    while True:
        os.system("CLS")
        a_dic={}
        menu_item_list=[]
        menu_id_list=[]
        for var_cur in vars_list:
            menu_id_list.append(var_cur)
            a_dic = val_def_rev_dic[var_cur]
            if len(a_dic)==0:
                a_item=f"{descript_dic[var_cur]}:  {bcolors.OKGREEN}{ret_dic[var_cur]}"
            
            else:
                a_color=bcolors.OKGREEN

                if len(val_standart_dic[var_cur])>0:
                    if not ret_dic[var_cur] in val_standart_dic[var_cur]:
                        a_color=bcolors.ATTENTIONWARNING

                a_item=f"{descript_dic[var_cur]}: {a_color} {a_dic[ret_dic[var_cur]]} "
            
            menu_item_list.append(a_item)


        spec_list = ["ok", "отмена"]
        spec_keys = ["\r", "/"]
        spec_id = spec_list

        header1=header+"\nВыберите параметр для изменения:"

        oo = questionFromList(color, header1, menu_item_list, 
            menu_id_list, "", spec_list, spec_keys, spec_id, 
            1, 1, 1, [], "")
            
        if oo == "ok":
            ret_id="2"
            for var_cur in vars_list:
                if ret_dic[var_cur]!=val_cur_dic[var_cur]:
                    ret_id="1"
                    break

            break
            
        elif oo == "отмена":
            return ["9", "Отменили изменение значений."]
            
        else:
            var_cur=oo
            val_old=ret_dic[var_cur]
            val_cur=val_old
            val_cur_def_dic=val_def_dic[var_cur]
            var_cur_type=var_type_dic[var_cur]

            if len(val_cur_def_dic)==0:
                os.system("CLS")
                a_txt="Введите значение для параметра " \
                    f"'{descript_dic[var_cur]}'.\n" \
                    "По окончании нажмите Enter.\n" \
                    "Чтобы прервать ввод нажмите '/'."
                spec_keys_list=["/", var_cur_type]
                if var_cur_type=="":
                    spec_keys_list=["/"]
                oo=inputSpecifiedKey(bcolors.OKBLUE, a_txt, "", [0],
                    spec_keys_list, 0, val_cur)
                if oo=="/":
                    continue

                ret_dic[var_cur]=oo
                val_cur=oo
                continue

                

            column_name=column_name_dic.get(var_cur, None)
            if column_name!=None and column_name!="":
                if employee_id==None or employee_id=="":
                    a_err_txt="Отсутствует табельный номер пользователя."
                    printWARNING(a_err_txt)
                    keystrokeEnter()
                    return ["0", a_err_txt]
                
                a_filter_dic = {"employeeId": [employee_id]}
                res = getAvailableActDic(a_filter_dic, date_cur, val_cur_def_dic,
                    column_name, 'Employeers.xlsx', 'EmployeersList', workmode, "1")
                if res[0] in ["0", "2"]:
                    return ["0", "При получении информации о доступных пользователю "
                        "значений параметра возникла ошибка."]

                val_cur_def_dic=res[2]

            menu_id_list=[]
            header2=f"{header}\nВыберите значение для " \
                f"параметра '{descript_dic[var_cur]}':"
            
            menu_item_list=list(val_cur_def_dic.keys())
            for a_descript in menu_item_list:
                menu_id_list.append(val_cur_def_dic[a_descript])

            res=menuSimple(color, header2, menu_item_list, 
                menu_id_list, val_cur)
            if res[0]=="2":
                ret_id="2"
                break

            else:
                ret_id="1"
                ret_dic[var_cur]=res[2]
                val_cur=res[2]


    if ret_id=="1":
        for var_cur in vars_list:
            a_dic={var_cur:ret_dic[var_cur]}
            res = saveConfigValue(
                config_file_name_dic[var_cur], a_dic, workmode, "заменить часть")
            if res[0]=="0":
                return ["0", "При сохранении новых значений в "
                    f"файле {config_file_name_dic[var_cur]} "
                    "возникла ошибка."]
        
    a_dic={"1": "Значение изменилось.",
        "2": "Значение осталось прежним.",
        "9": "Выбрали 'отмена'."}
    
    return [ret_id, a_dic[ret_id], ret_dic]




def cycleCheckResizeFile(file_path:str, control_size: int, 
    time_interval=1, time_interval_limit=10):

    time_sec_start=toformatNow()[3]
    file_size_cur=0
    while True:
        time_sec_cur=toformatNow()[3]
        if time_sec_cur-time_sec_start>time_interval:
            try:
                file_size_cur=os.path.getsize(file_path)
            except Exception:
                a_err_txt="Ошибка при получении информации о " \
                    f"размере файла '{file_path}'."
                a_txt=a_err_txt+f"\n{bcolors.OKBLUE}Нажмите Enter."
                printMsgWait(a_txt, bcolors.WARNING, "1", ["\r"])
                return ["0", a_err_txt]
            
            if file_size_cur>control_size:
                return ["1", "Размер файла увеличился.", file_size_cur]

            elif file_size_cur<control_size:
                return ["2", "Размер файла уменьшился.", file_size_cur]

            else:
                if time_sec_cur-time_sec_start>time_interval_limit:
                    return ["3", "Размер файла остался прежним.", 
                            file_size_cur]
                


def checkResizeFile(file_path:str, control_size: int):

    file_size_cur=0

    try:
        file_size_cur=os.path.getsize(file_path)
    except Exception:
        a_err_txt="Ошибка при получении информации о " \
            f"размере файла '{file_path}'."
        a_txt=a_err_txt+f"\n{bcolors.OKBLUE}Нажмите Enter."
        printMsgWait(a_txt, bcolors.WARNING, "1", ["\r"])
        return ["0", a_err_txt]
            
    if file_size_cur>control_size:
        return ["1", "Размер файла увеличился.", file_size_cur]

    elif file_size_cur<control_size:
        return ["2", "Размер файла уменьшился.", file_size_cur]

    else:
        return ["3", "Размер файла остался прежним.", file_size_cur]
    


def closeProgram(filename:str):

    try:
        os.system(f'taskkill /f /t /im {filename}')
    except Exception:
        pass

    return



def sendMail(file_name_config:str, subject: str, message_txt: str,
    attach_file_list=[], workmode="эксплуатация", 
    print_msg_err="1", rec_block_name="конфигурация"):

    res=readGonfigValue(file_name_config,[],{}, workmode, "1")
    if res[0]!="1":
        return ["0", "Ошибка при чтении конфигурационного файла."]

    if rec_block_name=="":
        rec_block_name="Васильеву"
        # rec_block_name="конфигурация"

    mail_config_dic=res[2]
    mailbox_username=mail_config_dic["mailbox_username"]
    mailbox_pw_encrypt=mail_config_dic["mailbox_pw_encrypt"]
    mail_server_adr=mail_config_dic["mail_server_adr"]
    mail_server_smtp_port=mail_config_dic["mail_server_smtp_port"]
    from_mail=mail_config_dic["from_mail"]
    
    recipients_dic=mail_config_dic[rec_block_name]

    to_mail_list=recipients_dic["to_mail_list"]
    cc_mail_list=recipients_dic["cc_mail_list"]
    
    res=cryptStringSec("расшифровать", mailbox_pw_encrypt)
    if res[0]=="0":
        a_err_txt = f"Не удалось получить пароль от почтового " \
            "ящика пользователя."
        printMsgWait(a_err_txt, bcolors.WARNING, 
            print_msg_err)
        return ["0", a_err_txt]

    mailbox_pw=res[2]

    to_mail=", ".join(to_mail_list)
    cc_mail=", ".join(cc_mail_list)

    msg = MIMEMultipart()
    msg["From"] = from_mail
    msg['To'] =to_mail

    if len(cc_mail_list)>0:
        msg['Cc'] =cc_mail

    msg["Subject"] = Header(f'{subject}', 'utf-8')
    msg["Date"] = formatdate(localtime=True)
    msg.attach(MIMEText(f"{message_txt}", 'html', 'utf-8'))

    if len(attach_file_list)>0:
        _,_, attach_file_dir = getUserFilePath("file_attach_mail", "1",
            workmode=workmode)
        if attach_file_dir=="":
            return ["0","Ошибка при формировании пути до папки с " 
                    "вкладываемыми файлами."]
        
        for file_name in attach_file_list:
            file_path = os.path.join(attach_file_dir, file_name)
            if not os.path.exists(file_path):
                a_err_txt = f"Отсутствует файл '{file_path}' " \
                    "для вложения в письмо."
                printMsgWait(a_err_txt, bcolors.WARNING, 
                    print_msg_err)
                return ["0", a_err_txt]
            
            part = MIMEBase('application', "octet-stream")
            part.set_payload(open(file_path,"rb").read())
            encoders.encode_base64(part)
            msg.attach(part)
            part.add_header('Content-Disposition', 'attachment', 
                            filename=file_name)
    
    content = msg.as_string()
    smtp = smtplib.SMTP_SSL(mail_server_adr, mail_server_smtp_port)

    t1=threading.Thread(target=sendMailThread, args=(smtp, mailbox_username, \
        mailbox_pw, from_mail, to_mail_list, cc_mail_list, 
        content), daemon=False)
    res=t1.start()

    return ["1", "Письмо успешно отправлено.", to_mail]
    


def sendMailThread(smtp_obj, mailbox_username: str,
    mailbox_pw: str, from_mail: str, to_mail_list_in: list, 
    cc_mail_list_in: list, content: str):
    
    to_mail_list=to_mail_list_in.copy()
    cc_mail_list=cc_mail_list_in.copy()

    try:
        smtp_obj.login(mailbox_username, mailbox_pw)
        smtp_obj.sendmail(from_mail, to_mail_list+ 
            cc_mail_list, content)
        smtp_obj.quit()
    except Exception:
        return

    return



def inputReportPeriod():

    dt_start=""

    date_cur_list = toformatNow()
    
    i=0
    while True:
        txt=f"Введите дату начала периода в формате " \
            "'ДД/ММ/ГГГГ'.\nДля ввода текущей даты можно " \
            "нажать 0.\nЧтобы отменить ввод нажмите '/'."
        
        spec_keys_list=["0", "/", "#date"]

        if i==1:
            if dt_start==date_cur_list[1]:
                return [dt_start, dt_start]

            txt=f"\nВведите дату окончания периода в формате " \
            "'ДД/ММ/ГГГГ'.\n" \
            "Для ввода текущей даты можно нажать 0.\n" \
            "Если дата окончания периода совпадает " \
            "с датой начала периода, то можно нажать Enter " \
            "с пустой строкой.\n" \
            "Чтобы отменить ввод - нажмите '/'."

            spec_keys_list=["0", "/", "\r", "#date"]

        oo = inputSpecifiedKey(bcolors.OKBLUE, txt, "", [0],
                specified_keys_list=spec_keys_list)
        if oo == "/":
            return ["", ""]
        
        if oo == "\r":
            return [dt_start, dt_start]
        
        elif oo=="0":
            oo=date_cur_list[1]
            printGREEN (f"\nУстановлена дата {oo}.")

        else:
            a = checkCorrectDate(oo)[3]
            
            if a > date_cur_list[2]:
                printWARNING(f"Введена дата {oo} позднее текущей " \
                        f"даты {date_cur_list[1]}.")
                continue
            
            if i==1:
                b = checkCorrectDate(dt_start)[3]
                if a < b:
                    printWARNING(f"Введена дата {oo} ранее даты начала периода " \
                        f"{dt_start}.")
                    continue

        if i==0:
            dt_start=oo
            i+=1
            continue

        return [dt_start, oo]
    


def inputReportDate(txt=None):

    date_cur_list = toformatNow()

    while True:
        if txt==None:
            txt = "\nВведите отчетную дату в формате 'ДД/ММ/ГГГГ'.\n" \
                "Для ввода текущей даты можно нажать 0.\n" \
                "Чтобы отменить ввод нажмите '/'."
        oo = inputSpecifiedKey(bcolors.OKBLUE, txt, "", [0],
                specified_keys_list=["0", "/", "#date"])
        if oo == "/":
            return ""
        
        elif oo=="0":
            printGREEN (f"\nУстановлена дата {date_cur_list[1]}")
            return date_cur_list[1]

        else:
            a = checkCorrectDate(oo)[3]
            if a > date_cur_list[2]:
                printWARNING(f"Введена дата {oo} позднее текущей " \
                        f"даты {date_cur_list[1]}.")
                continue
            return oo
        


def getPathOptoRun(workmode="эксплуатация"):

        res=getUserFilePath("opto_run.json", "0", workmode)
        opto_run_path=res[2]
        if opto_run_path == "":
            return ["0", res[1]]
        
        
        res = getUserFilePath("multi_config_dir",
            only_dir="1", workmode=workmode)
        multi_config_dir=res[2]
        if multi_config_dir == "":
            return ["0", res[1]]
        
        return ["1", "Пути сформированы.", opto_run_path, multi_config_dir]



def getOneSimbFile(var_name: str, workmode="эксплуатация"):

    def innerRetSimb():
        nonlocal val
        nonlocal ind

        while True:
            if val!=None and val!="" and ind<len(val):
                simb=val[ind:ind+1]
                ind+=1
                return simb

            else:
                val=innerReadDataFile(val)
    
    
    def innerReadDataFile(val_old: str):
        time_start=toformatNow()[3]
        
        attemt_count=1

        while True:

            res=readGonfigValue("print_big_font_line.json",[],{}, workmode, "0")
            if res[0]!="1":
                attemt_count+=1
                if attemt_count>3:
                    return None
                
                else:
                    continue

            attemt_count=1

            val=res[2][var_name]

            if val=="" or val==None or val==val_old:
                time_now=toformatNow()[3]
                
                while (time_start+time_interval)>time_now:
                    time_now=toformatNow()[3]
                    pass

                time_start=toformatNow()[3]

                continue

            return val

    
    res=readGonfigValue("print_big_font.json",[],{}, workmode, "1")
    if res[0]!="1":
        return None

    time_interval=res[2]["big_font_time_interval"]
    if time_interval<1:
        time_interval=1
    
    res=readGonfigValue("print_big_font_line.json",[],{}, workmode, "1")
    if res[0]!="1":
        return None
    
    val=res[2][var_name]

    ind=0

    return innerRetSimb

  
def getLastCellExcel(sheet, col_start=1, row_start=1):

    num_rows = sheet.max_row

    num_cols = sheet.max_col

    for i in range(row_start, num_rows+1):
        a_value = sheet.cell(row=i, column=col_start).value
        if a_value != None and a_value != "":
            continue
        
        i+=1
        break
    
    i-=1
    for j in range(i, num_cols+1):
        a_value = sheet.cell(row=i, column=j).value
        if a_value != None and a_value != "":
            continue

        j+=1
        break

    j-=1
    return ["1", "Координаты последней заполненной " \
        "ячейки определены.", j, i]
        


def replaceStrInFile(sub_str:str, new_str: str, file_path:str,
    mode="0", msg_err_print="1", encoding='utf-8'):

    
    dir_path=os.path.split(file_path)[0]

    a=toformatNow()[4]
    file_name_index=0

    replace_counts=0

    if mode=="0":
        sub_str=f"{sub_str}\n"
    
    while True:
        if file_name_index==0:
            file_2_name=f"{a}.txt"
        
        else:
            file_2_name=f"{a}_{file_name_index}.txt"

        file_2_path=os.path.join(dir_path, file_2_name)
        if not os.path.exists(file_2_path):
            break
    
        file_name_index+=1
    
    try:
        with open(file_path, 'r', errors="ignore", encoding=encoding) as f1, \
            open(file_2_path, 'w', encoding=encoding) as f2:
                        
            for line in f1:
                if mode=="0":
                    if line==sub_str:
                        line=f"{new_str}\n"
                        
                        replace_counts+=1
                    
                else:
                    if sub_str in line:
                        line=line.replace(sub_str, new_str)

                        replace_counts+=1
            
                f2.write(line)

    except Exception as e:
        err_txt="При выполнении операции по замене строки в файле " \
            f"{file_path} возникла ошибка: {e}."
        if msg_err_print=="1":
            printWARNING(err_txt)

        return["0", err_txt]
    
    if replace_counts>0:
        try:
            os.remove(file_path)

        except Exception:
            txt_err = f"Ошибка при удалении файла {file_path}"
            if msg_err_print == "1":
                printWARNING(txt_err)
            return ["0", txt_err]

        
        try:
            os.rename(file_2_path, file_path)
        except Exception:
            txt_err = f"Ошибка при переименовании файла '{file_2_path}'."
            if msg_err_print == "1":
                printWARNING(txt_err)
            return ["0", txt_err]
        
        return ["1", "Замена произведена успешно."]
    
    else:
        return ["2", "Искомая строка (фраза) в файле ненайдена."]



def findFileNameInDir(dir_path: str, sub_str:str, msg_err_print="1"):

    
    file_name_list=[]

    try:
        dir_list = os.listdir(dir_path)

    except Exception as e:
        err_txt="При получении списка файлов из папки " \
            f"{dir_path} возникла ошибка: {e}."
        if msg_err_print=="1":
            printWARNING(err_txt)

        return["0", err_txt, file_name_list]    


    for file_name in dir_list:
        if sub_str in file_name:
            file_name_list.append(file_name)
        
    if len(file_name_list)>0:
        return ["1", "Имя файла найдено", file_name_list]
    
    return ["2", "Файл в директории не найден", file_name_list]



def sorteFileNameReport(file_name_in_list: list):

    file_name_list=file_name_in_list.copy()



def dictCopy(in_dic: dict):

    keys_list=list(in_dic.keys())

    out_dict={}

    for key in keys_list:
        val=in_dic[key]
        val_type=type(val)

        if val_type==list:
            val=listCopy(val)

        elif val_type==dict:
            val=dictCopy(val)

        out_dict[key]=val

    return out_dict



def listCopy(in_list: list):

    out_list=[]

    for val in in_list:
        val1=val
        val_type=type(val)
        if val_type==list:
            val1=listCopy(val)
        
        elif val_type==dict:
            val1=dictCopy(val)

        out_list.append(val1)

    return out_list
