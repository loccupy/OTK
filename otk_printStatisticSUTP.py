# выводим на экран статистику из СУТП

from prettytable import PrettyTable
from otk_opto import writeDefaultValue, bcolors, optoRunVarRead

from libs.otkLib import *
from libs.sutpLib import request_sutp

import time
from datetime import datetime, timedelta
import os
import sys
from win32gui import GetWindowText, GetForegroundWindow
import keyboard     #pip install keyboard для отслеживания нажатия клавиш
# для записи ссылки на тек. каталог
from otk_saveLinkCurFolder import *

from openpyxl import load_workbook  # pip install openpyxl для работы с Excell файлами



# 060624
# выводим на экран в виде таблицы список выполненных работ всех сотрудников за указанную дату
# date_filter-запрашиваем данные по указанной дате,
#   если дата не указана, то берем текущий день
# workmode-режим работы программы:"тест"-тест,"эксплуатация"-эксплуатация (для ПП request_sutp)
# enable_change_get_statistic-разрешить отключение получения статистики из СУТП при
#   возникновении ошибки во время обмена данными с СУТП
# employeer_filter_list - список сотрудников, по которым выводим информацию.
#   Если список пуст, то выводим информацию по всем сотрудникам.
#   Формат списка:
#   employeer_filter_list=[{"nameFull":"Кузьмин Сергей Юрьевич", "nameShort":"Кузьмин С.Ю."},
#                    {"nameFull":"Васильев Владислав Валерьевич", "nameShort":"Васильев В.В."}]
# print_mode-режим вывода таблиц:"ОТК+прочие"- выводит на экран только
#       значения статусов ОТК: "ОТК пройден","В ремонте на линии", "Дефект".
#       Другие статусы будут сведены в общий столбец "Прочие статусы".
#   "все статусы" - формирует в одну таблицу все найденные статусы;
#   "ОТК,другие"- формирует две таблицы: только со стаусами ОТК и только с другими статусами

def getStatisticEmployeerSUTP(date_filter="", workmode="эксплуатация", output_device="display",
                              enable_change_get_statistic="yes", employeer_filter_list=[], grade_filter_list=[],
                              print_mode="ОТК+прочие"):
    # возвращает список:
    # [0]-статус выполнения операции ("0"-прочие ошибки,"1"-успешно,
    #   "2"-ошибка, возникшая при обмене данными с сервером,
    #   "3"-если имеются прочие статусы, не относящиеся к ОТК)
    # [1]-словесное описание статуса выполнения операции
    # [2]-итоговый сгруппированный список, из которого
    #   создается таблица PrettyTable

    # метка получения статистики из СУТП:"0"-откл.,"1"-вкл.
    global get_statistic_SUTP

    # получим тек. дату в формате строки
    # "ДД.ММ.ГГГГ ЧЧ:ММ:СС"
    date_cur_full = toformatNow()[0]
    # возьмем только дату
    date_cur = date_cur_full[0:10]
    # возьмем только время "ЧЧ:ММ"
    time_cur = f"{date_cur_full[11:13]} час {date_cur_full[14:16]} мин"
    # если не указан фильтр по дате
    if date_filter == "":
        # установим фильтр по текущей дате (сегодня)
        date_filter = date_cur
    # преобразуем дату в формат "ГГГГ.ММ.ДД"
    date_req = f"{date_filter[6:10]}.{date_filter[3:5]}.{date_filter[0:2]}"
    # запросим в СУТП список список выполненных работ всех сотрудников за указанную дату
    url1 = f"/api/Employee/EmployeesHistory/Date={date_req}"
    response = request_sutp(
        "GET", url1, [], err_txt="", workmode=workmode)
    # если вернулась ошибка
    if response[0] != "1":
        # если пользователю разрешено изменение метки
        # получения статистики из СУТП
        if enable_change_get_statistic == "yes":
            colortxt = ""
            txt1 = f"{bcolors.WARNING}При получении данных из СУТП произошла ошибка.{bcolors.ENDC}\n" \
                f"{bcolors.WARNING}Отключить отображение статистики из СУТП? (0-нет, 1-да){bcolors.ENDC}"
            specified_keys = ["0", "1"]
            specified_keys_only = 1
            oo = questionSpecifiedKey(colortxt=colortxt, txt=txt1, specified_keys_in=specified_keys,
                                      file_name_mp3="", specified_keys_only=specified_keys_only)
            if oo == "1":
                # сбросим метку отображения статистики из СУТП
                get_statistic_SUTP = "0"
                # запишем новые зн-я в словарь значений по умолчанию
                default_value_dict = writeDefaultValue(default_value_dict)
                # сохраним изменения зн-ий по умолчанию в файле opto_run.json
                saveConfigValue('opto_run.json', default_value_dict)
                print(
                    f"\n{bcolors.OKGREEN}Отображение статистики из СУТП отключено.{bcolors.ENDC}")
        # выйдем с ошибкой
        return ["2", f"Ошибка при обмене данными с СУТП:{response[0]}", []]
    # загрузим json данные в переменную для дальнейшей обработки
    # val_sutp = response[1].json()
    # получим данные, преобразованные из json формата
    val_sutp = response[2]

    # для ТЕСТА
    # print(val_sutp)
    # file_name="C:\\Vasilev\\Programm\\gurux_otk\\1userData\\2.json"
    # with open(file_name, "w", errors="ignore", encoding='utf-8') as file:
    #                 json.dump(val_sutp, file)

    # если переменная пустая
    if len(val_sutp) == 0:
        txt1 = f"{bcolors.WARNING}Статистика по сотрудникам за {date_filter} в СУТП отсутствует.{bcolors.ENDC}"
        print(txt1)
        # выйдем из ПП
        return ["1", f"Статистика по сотрудникам отсутствует", []]

    # список сотрудников
    employees_list = employeer_filter_list

    # сформируем словарь сотрудников с полным и кратким ФИО
    employee_FIO_dic = {}
    # подготовим список с полным ФИО сотрудников
    employee_full_FIO_list = []
    # если имеются данные в фильтре-списке по сотрудникам
    if len(employeer_filter_list) > 0:
        # переберем список сотрудников из фильтра
        for employee in employeer_filter_list:
            # добавим полное ФИО в список
            employee_full_FIO_list.append(employee['nameFull'])
            # добавим в новый словарь значения полного и краткого ФИО
            employee_FIO_dic[employee['nameFull']] = employee['nameShort']
    # фильтр-список пуст
    else:
        # Из полученных данных выберем полные Ф.И.О.
        # переберем полученные данные
        for val_sutp_dict in val_sutp:
            # получим полное ФИО
            employee_full_FIO = val_sutp_dict["name"]
            # если это ФИО отсут. в списке
            if employee_full_FIO not in employee_full_FIO_list:
                # добавим полное ФИО в список
                employee_full_FIO_list.append(employee_full_FIO)
                # Сформируем краткое ФИО
                employee_short = getShortFIO(employee_full_FIO)
                # добавим кр.форму ФИО в словарь
                employee_FIO_dic[employee_full_FIO] = employee_short

    # создадим метку вывода на экран второй таблицы
    # для режима "ОТК,другие"
    cicl = 1
    # если режим вывода "ОТК,другие" (2 таблицы)
    if print_mode == "ОТК,другие":
        # установим метку вывода на экран второй таблицы
        # для режима "ОТК,другие"
        cicl = 3

    # подготовим переменную для списка с прочими столбцами
    table_other_column_list = []
    # пока не вывели на экран вторую таблицу для
    # режима вывода "ОТК,другие"
    # или для других режимов вывода
    while cicl > 0:
        # список для группировки значений в виде
        grouping_values_list = [{"keyNameSearch": "operation",
                                 "keyNameNew": "grade", "ОТК пройден": "На поверку",
                                 "В ремонте на линии": "В ремонт", "Дефект": "В ремонт",
                                 "Прочие статусы": "*"}]

        # если режим вывода "все статусы"
        if print_mode == "все статусы":
            # изменим список для группировки значений
            grouping_values_list = [{"keyNameSearch": "operation",
                                     "keyNameNew": "grade", "ОТК пройден": "На поверку",
                                     "В ремонте на линии": "В ремонт", "Дефект": "В ремонт"}
                                    ]
        # если режим вывода "ОТК,другие" (2 таблицы)
        # и делаем второй или третий проход цикла (получили
        # столбец с прочими статусами)
        elif print_mode == "ОТК,другие" and cicl < 3:
            # изменим список для группировки значений
            grouping_values_list = [{"keyNameSearch": "operation",
                                     "keyNameNew": "grade"}]

        # получим из ф.Production.xlsx словарь соответствий кодов изделий
        # виду изделия ("0008": "1 ф.пр.","0114": "3 ф.пр.", 
        # "0203": "3 ф.тр.","1012":"МС 3 ф.сч."
        # и список видов изделий ("1 ф.пр.","3 ф.пр.","3 ф.тр.","МС 3ф.сч.")
        res=getDevicesTypeNameList(workmode=workmode)
        # если вернулась ошибка
        if res[0]!="1":
            # выведем сообщение об ошибке
            print(f"{bcolors.WARNING}Статистика по сотрудникам: {res[1]}{bcolors.ENDC}")
            # выйдем из ПП с ошибкой
            return["0",res[1],[]]
        # получим словарь соответствия кодов изделий виду изделий
        # id_device_type_name_dict = res[2]

        # подготовим словарь для группировки по кодам изделий
        group_dic_3={"keyNameSearch": "codeElementType","keyNameNew": "typeName"}
        # сделаем общий словарь добавив словарь соответствий кодов изделий и виду изделий
        combined_dic = {**group_dic_3, **res[2]}

        # добавим в список для группировки коды изделий
        # grouping_values_list.append({"keyNameSearch": "codeElementType",
        #                              "keyNameNew": "typeName", "0008": "1 ф.пр.",
        #                              "0114": "3 ф.пр.", "0203": "3 ф.тр."})

        # добавим в список для группировки коды изделий
        grouping_values_list.append(combined_dic)

        # сформируем словарь для добавления в grouping_values_list
        group_dic_1 = {"keyNameSearch": "name", "keyNameNew": "nameShot"}
        # объединим два словаря
        group_dic_2 = {**group_dic_1, **employee_FIO_dic}
        # добавим словарь в grouping_values_list
        grouping_values_list.append(group_dic_2)
        # # для ТЕСТА
        # print("список для группировки значений")
        # printDicInList(grouping_values_list)

        # сформируем список-фильтр с именами сотрудников
        filter_to_grouping_list = [{"key_name": "name",
                                   "item": employee_full_FIO_list}]

        # если находимся в режиме вывода "ОТК,другие"
        # и печатаем вторую таблицу
        if print_mode == "ОТК,другие" and cicl == 1:
            # добавим в список-фильтр список с именами
            # прочих статусов, который получили при
            # подготовке данных для первой таблицы
            filter_to_grouping_list.append({"key_name": "operation",
                                            "item": table_other_column_list})
        # если находимся в режиме вывода "ОТК,другие"
        # и печатаем первую таблицу
        elif print_mode == "ОТК,другие" and cicl == 2:
            # добавим в список-фильтр список со
            # статусами ОТК
            filter_to_grouping_list.append({"key_name": "operation",
                                            "item": ["ОТК пройден", "В ремонте на линии",
                                                     "Дефект"]})

        # # сформируем список-фильтр со статусами
        # filter_to_grouping_list=[{"key_name":"operation",
        #     "item":["ОТК пройден","В ремонте на линии",
        #             "Дефект"]}]

        # # для ТЕСТА
        # print("список-фильтр с именами сотрудников")
        # printDicInList(filter_to_grouping_list)

        # список с именами ключей, которые складываем
        # в сгруппированные строки
        key_name_merged_list = ['quantity']

        # список с ключами, которые остаются в
        # сгруппированном списке
        key_name_group_list = ["nameShot", "typeName",
                               "grade", "quantity"]

        content_list = val_sutp
        # получим сгруппированные данные по ключам в key_name_merged_list
        res = addGroupingValueToData(content_list, grouping_values_list,
                                     filter_to_grouping_list, key_name_merged_list,
                                     key_name_group_list)
        # если список пуст
        if res[0] == "2":
            txt2 = f"сотруднику {employees_list[0]['nameShort']}"
            txt_name = ""
            # если список сотрудников содержит более 1 записи
            if len(employees_list) > 1:
                txt2 = "сотрудникам:"
                # переберем список сотрудников
                for employeer in employees_list:
                    # сформируем строку из списка сотрудников
                    txt_name = txt_name+"\n"+employeer["nameShort"]
            txt1 = f"{bcolors.WARNING}В СУТП отсутствует статистика за {date_filter}{bcolors.ENDC}\n" \
                f"{bcolors.WARNING}по {txt2}{bcolors.ENDC}{bcolors.WARNING}{txt_name}{bcolors.ENDC}\n"
            print(txt1)
            # выйдем из ПП
            return ["1", f"Статистика по сотрудникам отсутствует", []]
        # если вернулась ошибка
        elif res[0] == "0":
            # выйдем из ПП с ошибкой
            return ["0", "ПП addGroupingValueToData() вернула ошибку.", []]
        # получим список, в котором строки объединены
        table_list = res[3]
        # если находимся в режиме вывода "ОТК,другие"
        # и делаем первый проход цикла или
        # находимся в другом режиме вывода
        if print_mode != "ОТК,другие" or \
                (print_mode == "ОТК,другие" and cicl == 3):
            # получим список с прочими столбцами
            table_other_column_list = res[4]

        # # для ТЕСТА
        # column_name_list=[
        #     {"columnName":"Ф.И.О",
        #      "columnValue":["shortName"],
        #      "sorting":"up"},
        #     {"columnName":"На поверку",
        #       "columnValue":["grade","На поверку","quantity"]},
        #     {"columnName": "На поверку",
        #      "columnValue": ["grade", "На поверку", "quantity"]},

        # итоговый сгруппированный объединенный список
        table_list_finish = table_list

        # Сформируем список статусов
        # подготовим переменную для списка "grade"
        grade_list = []
        for table_row in table_list_finish:
            # получим тек. зн-е для ключа "grade"
            grade = table_row["grade"]
            # если данного статуса нет в списке
            if grade not in grade_list:
                # добавим его в список
                grade_list.append(grade)

        column_name_list1 = [{"columnName": "Вид изделия",
                              "valueKeyList": ["typeName"],
                              "typeValue": "str", "sorting": "down"
                              },
                             {"columnName": "Ф.И.О.",
                              "valueKeyList": ["nameShot"],
                              "typeValue": "str", "sorting": "down"
                              }
                             ]
        # Добавим список статусов
        # подготовим пустой словарь
        column_name_dict = {}
        # переберем список статусов
        for grade in grade_list:
            # запишем заголовок столбца
            column_name_dict["columnName"] = grade
            # укажем из какого ключа нужно брать зн-е
            column_name_dict["valueKeyList"] = ["quantity"]
            # укажем тип данных "число"
            column_name_dict["typeValue"] = "num"
            # зададим фильтр
            column_name_dict["filter"] = {"grade": [grade]}
            # укажем,что данный столбец нужно учитывать
            # в сумме ВСЕГО по строкам
            column_name_dict["totalColumn"] = "yes"
            # добавим полученный словарь в список
            column_name_list1.append(column_name_dict.copy())

        # сформируем словарь для столбца "Итого"
        column_name_dict = {"columnName": "Итого",
                            "valueKeyList": ["quantity"],
                            "typeValue": "num",
                            "filter": {"grade": grade_list},
                            "totalColumn": "yes"
                            }
        # добавим словарь "Итого" в список для таблицы
        column_name_list1.append(column_name_dict.copy())
        # фильтровать исходные данные не требуется
        filter_dict1 = {}
        # отформатируем данные для таблицы
        res = formatToTable(table_list_finish,
                            column_name_list1, filter_dict1,
                            hide_zero_column=True, hide_zero_row=True,
                            subtotal=True, total=True)
        # получим данные для таблицы
        table_list = res[2]
        # получим заголовки столбцов
        column_name_list = res[3]
        # подготовим переменную для общего
        # списка зн-й столбцов
        val_table_list = []
        # подготовим переменную для зн-й строки
        val_row_table_list = []
        # переберем список с данными для таблицы
        for table_row_dict in table_list:
            # получим список зн-й столбцов
            val_row_table_list = list(table_row_dict.values())
            # добавим в общий список
            val_table_list.append(val_row_table_list)

        # если не находимся в режиме вывода "ОТК,другие"
        # или не делаем в нем первый проход цикла
        if print_mode != "ОТК,другие" or \
                (print_mode == "ОТК,другие" and cicl < 3):
            # выведем заголовок отчета по видам изделий на экран
            txt1 = f"за сегодня по состоянию на {time_cur}"
            # если был установлен фильтр по дате
            if date_filter != date_cur:
                txt1 = f"за {date_filter}"
            txt1_1 = "сотрудникам"
            # если список сотрудников содержит один элемент
            if len(employees_list) == 1:
                txt1_1 = "сотруднику"
            txt1_2 = ""
            # если находимся в режиме вывода "ОТК,другие"
            # и печатаем первую таблицу
            if print_mode == "ОТК,другие" and cicl == 2:
                # изменим надпись
                txt1_2 = " (статусы ОТК)"
            # если находимся в режиме вывода "ОТК,другие"
            # и печатаем вторую таблицу
            elif print_mode == "ОТК,другие" and cicl == 1:
                # изменим надпись
                txt1_2 = " (прочие статусы)"

            print(f"\n{bcolors.OKGREEN}Статистика из СУТП по {txt1_1}{bcolors.ENDC}\n"
                  f"{bcolors.OKGREEN}{txt1}{txt1_2}:{bcolors.ENDC}")

            # создадим переменную для таблицы
            table_statistic = PrettyTable()
            # зададим наименования столбцов таблицы
            table_statistic.field_names = column_name_list

            # выравним значения в столбце "Ф.И.О." по левому краю
            table_statistic.align["Ф.И.О."] = "l"

            # добавим полученный список в таблицу
            table_statistic.add_rows(val_table_list)
            # выведем на экран информацию
            print(f"{bcolors.OKGREEN}{table_statistic}{bcolors.ENDC}\n")

        # пойдем на следующий цикл
        cicl -= 1

    # если имеются прочие статусы, не относящиеся к ОТК
    if len(table_other_column_list) != 0:
        # выйдем с успехом
        return ["3", f"Статистика успешно показана. "
                "Имеются прочие статусы.", table_list_finish]
    # выйдем с успехом
    return ["1", f"Статистика успешно показана.", table_list_finish]



# 200824
# получаем из ф.Production.xlsx словарь соответствий кодов изделий
# виду изделия ("0008": "1 ф.пр.","0114": "3 ф.пр.", 
# "0203": "3 ф.тр.","1012":"МС 3 ф.сч."
# и список видов изделий ("1 ф.пр.","3 ф.пр.","3 ф.тр.","МС 3ф.сч.")
# workmode-режим работы программы:"тест","эксплуатация"
def getDevicesTypeNameList(workmode="эксплуатация"):
    # возвращает список:
    # [0]-код результата операции:"0"-ошибка, "1"-успешно
    # [1]-словесное описание результата операции
    # [2]-словарь соответствий кодов изделий виду изделий
    # [3]-список видов изделий
    # сформируем полный путь до Production.xlsx
    _, _, file_name = getUserFilePath("Production.xlsx", workmode=workmode)
    # если вернулась ошибка
    if file_name == "":
        # выйдем из ПП с ошибкой
        return ["0", "Ошибка при формировании пути до файла.", {},[]]
    
    try:
        # подключимся к книге "Production.xlsx"
        wb = load_workbook(file_name)
        # откроем лист "Вид изделия" для чтения из файла "Production.xlsx"
        sheet = wb["Вид изделия"]
    except Exception as e:
        txt1=f"Не удалось подключиться к листу 'Вид изделия' ф.Production.xlsx: " \
            f"{e.args[0]}"
        # выйдем из ПП с ошибкой
        return ["0", txt1,{}, []]
    # определим число занятых строк на листе "Вид изделия"
    num_rows = sheet.max_row
    # получим номер стартовой строки
    row_start=sheet.cell(row=3, column=1).value
    # получим номер стартового столбца
    col_start = sheet.cell(row=4, column=1).value
    # если номер стартовой строки или столбца пусты
    if row_start==None or row_start=="" or \
        col_start==None or col_start=="":
        # выйдем из ПП с ошибкой
        return ["0", "Не указана стартовая строка и (или) столбец \
            в ф.Production.xlsx. на листе 'Вид ПУ'.",{}, []]

    # Сформируем словарь соответствия кодов изделий виду ПУ
    # подготовим переменную для словаря соответствия, чтобы
    # потом группировать данные
    id_device_type_name_dict = {}
    # подготовим переменную для списка видов ПУ
    device_type_name_list=[]

    # переберем ячейки в табл.Excel
    for i in range(row_start,num_rows+1):
        # прочитаем код изделия
        id_device_type = sheet.cell(row=i, column=col_start).value
        # прочитаем вид ПУ
        device_type_name = sheet.cell(row=i, column=col_start+1).value
        # если код изделия и вид ПУ имеют значения
        if id_device_type!=None and id_device_type!="" and \
            device_type_name != None and device_type_name != "":
            # сохраним значения в словаре
            id_device_type_name_dict[id_device_type] = device_type_name
            # если данного вида ПУ нет в списке видов ПУ
            if device_type_name not in device_type_name_list:
                # добавим вид ПУ в список видов ПУ
                device_type_name_list.append(device_type_name)
        # если ячейки пустые
        else:
            # выйдем из for i
            break

    # выйдем из ПП с успехом
    return ["1", "Данные сформированы успешно.",id_device_type_name_dict, 
            device_type_name_list]



# 030325
# получаем из СУТП  информацию о количестве прошедших ПУ через
# этап ОТК и отображаем на экране
# date_filter-запрашиваем данные по указанной дате
#   если дата не указана, то берем текущий день
# workmode-режим работы программы:"тест"-тест,
#   "эксплуатация"-эксплуатация (для ПП request_sutp)
# enable_change_get_statistic-разрешить отключение получения
#   статистики из СУТП при возникновении ошибки во время
#   обмена данными с СУТП
# product_filter_user_list-список устройств, по которым выводим
#  данные:
#  ["ИПУ 1Ф", "ИПУ 1Ф S", "ИПУ 3Ф", "ИПУ 3Ф ТТ", "ИПУ 3Ф S"]
def printStatisticSUTP(date_filter="", workmode="эксплуатация",
    enable_change_get_statistic="yes", product_filter_user_list=[]):
    # возвращает список:
    # [0]-статус выполнения операции ("0"-прочие ошибки,"1"-успешно,
    #   "2"-ошибка, возникшая при обмене данными с сервером)
    # [1]-словесное описание статуса выполнения операции


    # метка получения статистики из СУТП:"0"-откл.,"1"-вкл.
    global get_statistic_SUTP

    # получим тек. дату в формате строки
    # "ДД.ММ.ГГГГ ЧЧ:ММ:СС"
    date_cur_full = toformatNow()[0]
    # возьмем только дату
    date_cur = date_cur_full[0:10]
    # возьмем только время "ЧЧ:ММ"
    time_cur = f"{date_cur_full[11:13]} час {date_cur_full[14:16]} мин"
    # если не указан фильтр по дате
    if date_filter == "":
        # установим фильтр по текущей дате (сегодня)
        date_filter = date_cur
    # преобразуем дату в формат "ГГГГ.ММ.ДД"
    date_req = f"{date_filter[6:10]}.{date_filter[3:5]}.{date_filter[0:2]}"

    # запросим в СУТП статистику за день по ОТК
    url1 = f"/api/Order/GetOrderPlanQcLineReportTableFor?startDate=" \
        f"{date_req}&endDate={date_req}"
    response = request_sutp(
        "GET", url1, [], err_txt="", workmode=workmode)
    # если вернулась ошибка
    if response[0] != "1":
        # если пользователю разрешено изменение метки
        # получения статистики из СУТП
        if enable_change_get_statistic == "yes":
            colortxt = ""
            txt1 = f"{bcolors.WARNING}При получении данных из СУТП произошла ошибка.{bcolors.ENDC}\n" \
                f"{bcolors.WARNING}Отключить отображение статистики из СУТП? (0-нет, 1-да){bcolors.ENDC}"
            specified_keys = ["0", "1"]
            specified_keys_only = 1
            oo = questionSpecifiedKey(colortxt=colortxt, txt=txt1, specified_keys_in=specified_keys,
                                      file_name_mp3="", specified_keys_only=specified_keys_only)
            if oo == "1":
                # сбросим метку отображения статистики из СУТП
                get_statistic_SUTP = "0"
                # запишем новые зн-я в словарь значений по умолчанию
                default_value_dict = writeDefaultValue(default_value_dict)
                # сохраним изменения зн-ий по умолчанию в файле opto_run.json
                saveConfigValue('opto_run.json', default_value_dict)
                print(
                    f"\n{bcolors.OKGREEN}Отображение статистики из СУТП "
                    f"отключено.{bcolors.ENDC}")
        # выйдем с ошибкой
        return ["2", f"Ошибка при обмене данными с СУТП:{response[0]}", []]
    # получим данные, преобразованные из json формата
    val_sutp = response[2]

        
    # подготовим переменную для статистики по линии сборки
    statistic_assembly_line_list=[]
    # запросим в СУТП статистику за день по линии сборки
    url1 = f"/api/Order/GetOrderPlanLineReportTableFor?startDate=" \
        f"{date_req}&endDate={date_req}"
    response = request_sutp(
        "GET", url1, [], err_txt="", workmode=workmode)

    # получим статистику линии сборки, преобразованные из json формата
    statistic_assembly_line_list = response[2]
    
    # для ТЕСТА
    # print(val_sutp)
    # file_name="C:\\Vasilev\\Programm\\gurux_otk\\1userData\\1.json"
    # with open(file_name, "w", errors="ignore", encoding='utf-8') as file:
    #                 json.dump(val_sutp, file)

    
    # если переменная пустая
    if len(val_sutp) == 0:
        txt1 = f"{bcolors.WARNING}Статистика за " \
            f"{date_filter} в СУТП отсутствует.{bcolors.ENDC}"
        print(txt1)
        # выйдем из ПП
        return ["1", f"Статистика отсутствует."]
    
    # установим метку, что за указанный день имеются только нулевые значения 
    # в статистике (выходной день)
    statistic_zero=True

    # переберем полученный от сервера список
    for a_rec in val_sutp:
        # если списки в текущем словаре имеют значения
        if a_rec["plannedDays"]!=None and  a_rec["doneDays"]!=None and \
            a_rec["doneDays"]!=None and a_rec["additionalDays"]!=None and \
            (len(a_rec["plannedDays"])>0 or len(a_rec["doneDays"])>0 or \
            len(a_rec["additionalDays"][0])>0 or len(a_rec["additionalDays"][1])>0):
            # сбросим метку, что в списке нулевые значения
            statistic_zero=False
            # выйдем из цикла перебора списка из СУТП
            #  (for a_rec in val_sutp)
            break


    # если установлена метка, что в списке нулевые значения
    if statistic_zero:
        txt1 = f"{bcolors.WARNING}Статистика за " \
            f"{date_filter} в СУТП отсутствует.{bcolors.ENDC}"
        print(txt1)
        # выйдем из ПП
        return ["1", f"Статистика отсутствует."]


    #сформируем список устройств, по которым выводим информацию
    product_filter_list=["ИПУ 1Ф", "ИПУ 1Ф S", "ИПУ 3Ф", 
        "ИПУ 3Ф ТТ", "ИПУ 3Ф S"]
    # если список устройств пользователя имеет записи и он отличается от
    # списка по умолчанию
    if len(product_filter_user_list)>0 and \
        product_filter_user_list!=product_filter_list:
        # изменим список устройств, по которым выводим информацию
        product_filter_list=product_filter_user_list.copy()

    # словарь соответствия вида изделия (1 ф., 3 ф.) и типа устройства 
    # из отчета
    device_type_name_dic={"ИПУ 1Ф": "1 ф.", "ИПУ 1Ф S": "1 ф.",
        "ИПУ 3Ф": "3 ф.","ИПУ 3Ф S": "3 ф.", "ИПУ 3Ф ТТ": "3 ф."}

   
    # подготовим список для вывода статистики
    statistic_list=[]

    # переберем полученный от сервера список
    for i in range(0, len(val_sutp)):
        # подготовим словарь для сохранения данных в итоговом списке
        statistic_line_dic={"вид устройства": "", "тип устройства": "",
            "факт ОТК": 0, "план ОТК": 0, "на поверку": 0,
            "в ремонт": 0, "баланс": 0, "план производства": 0,
            "факт производства": 0}
        # прочитаем тек. запись (словарь)
        rec_dict = val_sutp[i]
        # получим тип изделия из словаря
        a_type=rec_dict["deviceGeneralTypeName"]
        # если тип изделия отсутствует в списке, по которому выводим информацию
        if not a_type in product_filter_list:
            # вернемся в начало цикла for i in range(0, len(val_sutp)) 
            # за следующим элементом списка
            continue

        # текущий элемент имеется в списке-фильтре
        # добавим новый ключ для вида изделия в зависимости
        #  от типа изделия
        statistic_line_dic["вид устройства"] = device_type_name_dic.get(a_type,"неизвестный тип")
        # добавим тип устройства в новый словарь
        statistic_line_dic["тип устройства"]=a_type

        # если имеется запись для ключа "doneDays"
        if rec_dict["doneDays"]!=None and len(rec_dict["doneDays"])>0:
            # сохраним данные для ключа "факт ОТК"
            statistic_line_dic["факт ОТК"]=rec_dict["doneDays"][0]["count"]

        # если имеется запись для ключа "doneDays"
        if rec_dict["additionalDays"]!=None and len(rec_dict["additionalDays"])>0:
            if len(rec_dict["additionalDays"][0])>0:
                # сохраним данные для ключа "план ОТК"
                statistic_line_dic["план ОТК"]=rec_dict["additionalDays"][0][0]["count"]
            if len(rec_dict["additionalDays"][1])>0:
                # сохраним данные для ключа "на поверку"
                statistic_line_dic["на поверку"]=rec_dict["additionalDays"][1][0]["count"]

        # сделаем расчет для столбца "В ремонт"
        statistic_line_dic["в ремонт"]=statistic_line_dic["факт ОТК"]-statistic_line_dic["на поверку"]

        # сделаем расчет для столбца "Баланс"
        statistic_line_dic["баланс"]=statistic_line_dic["факт ОТК"]-statistic_line_dic["план ОТК"]

        # если имеется статистика линии сборки
        if len(statistic_assembly_line_list)>0:
            # переберем словари в списке статистики
            for a_dic in statistic_assembly_line_list:
                # если в текущем словаре указан текущий тип устройства
                if a_dic["deviceGeneralTypeName"]==a_type:
                    # если указан план
                    if len(a_dic["plannedDays"])>0:
                        # сохраним плановое значение линии сборки
                        statistic_line_dic["план производства"]=a_dic["plannedDays"][0]["count"]
                    # если указан факт
                    if len(a_dic["doneDays"])>0:
                        # сохраним плановое значение линии сборки
                        statistic_line_dic["факт производства"]=a_dic["doneDays"][0]["count"]
        
        # сохраним словарь в списке для статистики
        statistic_list.append(statistic_line_dic)

    
    # подготовим список столбцов для табл.6
    column_name_list1 = [{"columnName": "Вид ПУ",
                         "valueKeyList": ["вид устройства"],
                          "typeValue": "str", "sorting": "down"
                          },
                          {"columnName": "Тип ПУ",
                          "valueKeyList": ["тип устройства"],
                          "typeValue": "str", "sorting": "down"
                          },
                         {"columnName": "Баланс",
                          "valueKeyList": ["баланс"],
                          "typeValue": "num",
                          "totalColumn": "yes"
                          },
                         {"columnName": "План ОТК",
                          "valueKeyList": ["план ОТК"],
                          "typeValue": "num", "totalColumn": "yes"
                          },
                          {"columnName": "Факт ОТК",
                          "valueKeyList": ["факт ОТК"],
                          "typeValue": "num", "totalColumn": "yes"
                          },
                          {"columnName": "В ремонт",
                          "valueKeyList": ["в ремонт"],
                          "typeValue": "num", "totalColumn": "yes"
                          },
                          {"columnName": "На поверку",
                          "valueKeyList": ["на поверку"],
                          "typeValue": "num", "totalColumn": "yes"
                          },
                          {"columnName": "План сборки",
                          "valueKeyList": ["план производства"],
                          "typeValue": "num", "totalColumn": "yes"
                          },
                          {"columnName": "Факт сборки",
                          "valueKeyList": ["факт производства"],
                          "typeValue": "num", "totalColumn": "yes"
                          },
                         ]

    
    # фильтровать исходные данные не требуется
    filter_dict1 = {}
    # отформатируем данные для таблицы
    res = formatToTable(statistic_list, column_name_list1, 
        filter_dict1, False, hide_zero_row=True, subtotal=True,
        total=True)
    # получим данные для таблицы
    table_list = res[2]
    # получим заголовки столбцов
    column_name_list = res[3]
    # подготовим переменную для общего
    # списка зн-й столбцов
    val_table_list = []
    # подготовим переменную для зн-й строки
    val_row_table_list = []
    # переберем список с данными для таблицы
    for table_row_dict in table_list:
        # получим список зн-й столбцов
        val_row_table_list = list(table_row_dict.values())
        # добавим в общий список
        val_table_list.append(val_row_table_list)

    # выведем заголовок отчета по факт. данным на экран
    txt1 = f"за сегодня по состоянию на {time_cur}"
    # если был установлен фильтр по дате
    if date_filter != date_cur:
        txt1 = f"за {date_filter}"

    printGREEN(f"\nСтатистика из СУТП по проверке "
          f"ПУ сотрудниками ОТК\n{txt1}:")

    # создадим переменную для таблицы с факт. данными
    table_statistic_fact = PrettyTable()
    # зададим наименования столбцов таблицы
    table_statistic_fact.field_names = column_name_list

    # выравним значения в столбце "Вид устройства" по левому краю
    table_statistic_fact.align["Вид ПУ"] = "l"
    # выравним значения в столбце "Тип устройства" по левому краю
    table_statistic_fact.align["Тип ПУ"] = "l"

    # добавим полученный список в таблицу
    table_statistic_fact.add_rows(val_table_list)
    # выведем на экран таблицу с факт.данными
    print(f"{bcolors.OKGREEN}{table_statistic_fact}{bcolors.ENDC}\n")
    
    # выйдем с успехом
    return ["1", f"Статистика успешно показана."]



# 100624
# возвращает словарь со значениями переменных, установленных
# по умолчанию для данной программы
# var_name_list-список с именами переменных, для которых
#  возвращаем зн-я, если указаны
def getDefaultValueStatistic(var_name_list=[]):
    
    # словарь со значениями по умолчанию для переменных
    default_val_statistic_dict = {"programUser": "начальник",
        "indexRetryInterval": 5,
        "subdivision": "ОТК", "workmode": "тест",
        "printMode": "ОТК+прочие",
        "productionUpdateInterval": 60,
        "productionLastUpdate":0,
        "productionUpdateOn":True
        }
    
    # подготовим словарь для заполнения
    val_dict={}
    
    # если список с именами имеет значения
    if len(var_name_list)>0:
        # переберем список с запрашиваемыми переменными
        for var_name in var_name_list:
            # если имя имеется в словаре со зн-ями 
            # по умолчанию
            if var_name in default_val_statistic_dict:
                # присвоим переменной соотв. зн-е
                val_dict[var_name] = default_val_statistic_dict[var_name]
    
    # если словарь пуст
    if len(val_dict)==0:
        #заполним его значениями по умолчанию
        val_dict = default_val_statistic_dict
    
    return val_dict


# 300624
# обновим ф.Production.xlsx в локальной папке
# workmode-режим работы программы:"тест","эксплуатация"
# update_mode-режим обновления файла:"откл"-отключено обновление,
#   "изменен"-обновление в случаях, когда эталонный файл позднее текущего
#   "замена"-безуслованя замена файла
def updateProductionxlsx(workmode="эксплуатация", update_mode="изменен"):
    # возвращает список
    # [0]-статус выполнения: "0"-ошибка,"1"-успех
    # [1]-словесное описание статуса
    # [2]-возвращаем время обновления файла в секундах

    # подготовим текст для ошибки
    txt_err = "Ошибка при получении плановых значений:"
    # сформируем полный путь до эталонного ф.Production.xlsx
    _, _, dir_name = getUserFilePath("publicProduction",
        only_dir="1", workmode=workmode)
    # если вернулась ошибка
    if dir_name == "":
        txt_err = f"{txt_err} формирование пути до эталонного ф.Production.xlsx."
        # сообщим об ошибке
        print(f"{bcolors.WARNING}{txt_err}{bcolors.ENDC}")
        # выйдем из ПП с ошибкой
        return ["0", txt_err, 0]
    # добавим к полученному пути имя файла-источника
    file_name_source = os.path.join(dir_name, "Production.xlsx")

    # сформируем полный путь до локального файла Production.xlsx
    _, _, file_name_dest = getUserFilePath("Production.xlsx",
                                           only_dir="0", workmode=workmode)
    # если вернулась ошибка
    if file_name_dest == "":
        txt_err = f"{txt_err} формирование пути до локального ф.Production.xlsx."
        # сообщим об ошибке
        print(f"{bcolors.WARNING}{txt_err}{bcolors.ENDC}")
        # выйдем из ПП с ошибкой
        return ["0", txt_err, 0]

    # подготовим переменную для ф-ции
    file_list = [{"fileSourcePath": file_name_source,
                  "fileDestPath": file_name_dest,
                  "updateMode": update_mode}
                 ]
    # если эталонный файл был изменен, то заменим его в локальной папке
    res = updateFiles(file_list=file_list, err_msg_print="1")
    # если вернулась ошибка
    if res[0] == "0":
        txt_err = f"{txt_err} копирование эталонного ф.Production.xlsx."
        # сообщим об ошибке
        print(f"{bcolors.WARNING}{txt_err}{bcolors.ENDC}")
        # выйдем из ПП с ошибкой
        return ["0", txt_err, 0]

    # Изменим время, когда обновляли ф.Production.xlsx в конфиг.файле
    # создадим пустой конфиг.словарь
    config_dict={}
    # получим тек.время в сек
    production_last_update = toformatNow()[3]
    # изменим время последнего обновления ф.Production.xlsx
    # в конфиг. словаре переменных
    config_dict['productionLastUpdate'] = production_last_update
    # сохраним конфиг. словарь в конфиг. ф.print_statistic_run.json
    res = saveConfigValue("print_statistic_run.json", config_dict)
    # если вернулась ошибка
    if res[0] == "0":
        # выведем сообщение об ошибке
        print(f"\n{bcolors.WARNING}При записи изменений в "
              f"конфигурационный файл возникла ошибка.{bcolors.ENDC}")
        txt1 = f"{bcolors.OKBLUE}Нажмите Ввод.{bcolors.ENDC}"
        questionSpecifiedKey("", txt1, ["\r"], "", 1)
        # выйдем из ПП с ошибкой
        return ["0","Ошибка при обновлении записи в конфигурационном файле.",0]
    
    # выйдем с успехом
    return ["1", "Файл ф.Production.xlsx успешно обновлен.", production_last_update]



# обрабатываем отпускание клавиш
def onPress(key):
    # метка статуса работы программы
    global programm_status   
    # индекс времени обновления информации
    global index_retry_interval
    # словарь со зн-ями конфигурации программы
    global config_value_dict 
    # режим вывода статистики по персоналу
    global print_mode
    # метка, что в таблице статистики по сотрудникам
    # отсутствует столбец "Прочие статусы"
    global view_other_grade
    
    foregroundWindow = GetWindowText(GetForegroundWindow())
    if "Статистика по проверке ПУ" in foregroundWindow or \
            "otk_printStatisticSUTP" in foregroundWindow:
        key_num_interval_list=["1","2","3","4","5"]
        # если нажали Enter и статус работы программы "цикл"
        if key.name=="enter" and programm_status=="цикл":
            # установим статус работы программы: обновить
            programm_status = "обновить"    
        # если нажали ESC
        elif key.name=="esc":
            # установим статус работы программы: выйти
            programm_status = "выйти"
        # если нажали цфиру от 1 до 5 и статус работы программы "цикл"
        elif key.name in key_num_interval_list and \
            programm_status == "цикл":
            # установим индекс времени обновления информации
            # в соответствии с нажатой клавишей
            index_retry_interval_new = int(key.name)
            # если новое значение не равно существующему
            # индексу
            if index_retry_interval_new != index_retry_interval:
                # установим новое значение индекса
                index_retry_interval = index_retry_interval_new
                # обновим зн-е переменной в конфиг. словаре переменных
                config_value_dict["indexRetryInterval"]=index_retry_interval
                # сохраним конфиг. словарь в конфиг. ф.print_statistic_run.json
                res=saveConfigValue("print_statistic_run.json",config_value_dict)
                # если вернулась ошибка
                if res[0]=="0":
                    # выведем сообщение об ошибке
                    print(f"\n{bcolors.WARNING}При записи изменений в " \
                            f"конфигурационный файл возникла ошибка.{bcolors.ENDC}")
                    txt1=f"{bcolors.OKBLUE}Нажмите Ввод.{bcolors.ENDC}"
                    questionSpecifiedKey("",txt1,["\r"],"",1)
                    # выйдем из программы
                    sys.exit()
                # установим статус работы программы: обновить
                programm_status = "обновить"
        # нажали "6"
        elif key.name=="6":
            # установим статус программы "ввод даты"
            programm_status="ввод даты"
            # переведем режим вывода в "ОТК+прочие"
            # переведем режим вывода в "ОТК+прочие"
            print_mode = "ОТК+прочие"
        # нажали "-"
        elif key.name=="-":
            # установим статус программы "ввод даты -1 день"
            programm_status="ввод даты -1 день"
            # переведем режим вывода в "ОТК+прочие"
            print_mode = "ОТК+прочие"
        # нажали "7"
        elif key.name=="7":
            # если статус программы "цикл",
            # режим вывода "ОТК+прочие" и
            # установлена метка возможности
            # раскрытия табл. статистики по
            # сотрудникам
            if programm_status == "цикл" and \
                print_mode == "ОТК+прочие" and \
                view_other_grade:
                # установим режим вывода 2 таблиц для статистики по персоналу: 
                # первая таблица содержит только данные из статусов "ОТК пройден",
                # "В ремонте на линии", "Дефект",
                # другая таблица содержит значения по другим статусам
                print_mode = "ОТК,другие"
                # установим статус работы программы: обновить
                programm_status = "обновить"
            # иначе
            elif programm_status == "цикл" and \
                    print_mode == "ОТК,другие":
                # установим режим вывода "ОТК+прочие"
                print_mode="ОТК+прочие"
                # установим статус работы программы: обновить
                programm_status = "обновить"



# 030325
# выводим статистику из СУТП
# по проверке ПУ: всего и по сотрудникам ОТК
def printStatisticMain():
    # метка статуса работы программы:
    #  "выйти"-выйти, "цикл"-в цикле,
    #  "обновить"-обновить,
    #  "ввод даты" - ожидаем ввод даты
    #  "ввод даты -1 день" - переходим из цикл.
    #   запроса на вывод отчета за пред. день
    #  "цикл отключен"- циклический запрос из
    #       СУТП отключен при выведенном отчете за
    #       предыдущую дату
    global programm_status
    #  индекс времени обновления информации
    global index_retry_interval
    # словарь со зн-ями конфигурации программы
    global config_value_dict
    # режим вывода статистики по персоналу:
    # "ОТК+прочие","все статусы","ОТК,другие"
    global print_mode
    # метка, что в таблице статистики по сотрудникам
    # отсутствует столбец "Прочие статусы"
    global view_other_grade


    # получим словарь со значениями по умолчанию для программы
    default_val_statistic_dict=getDefaultValueStatistic()

    # список конфиг. переменных для присвоения им зн-я по умолчанию
    # из файла конфигурации
    var_local_list=["programUser","subdivision",
        "workmode", "indexRetryInterval", "printMode",
        "productionUpdateInterval","productionLastUpdate",
        "productionUpdateOn"]
    
    # получим словарь со зн-ями, которые записаны в конфиг. файле
    # print_statistic_run.json
    res=readGonfigValue("print_statistic_run.json",
        var_local_list, default_val_statistic_dict)
    # если вернулась ошибка
    if res[0]!="1":
        # выведем сообщение об ошибке
        print(f"\n{bcolors.WARNING}При формировании перменных " \
              f"конфигурации программы возникла ошибка.{bcolors.ENDC}")
        txt1=f"{bcolors.OKBLUE}Нажмите Ввод.{bcolors.ENDC}"
        questionSpecifiedKey("",txt1,["\r","",1])
        # выйдем из ПП
        return
    # сформируем словарь со зн-ями переменных для конфигурации программы
    config_value_dict=res[2]

    # Запишем зн-я в переменные
    # установим индекс времени обновления информации
    index_retry_interval = config_value_dict["indexRetryInterval"]
    # установим режим работы программы эксплуатация/тест
    workmode=config_value_dict["workmode"]
    # установим величину интервала 60 сек
    # для расчета времени обновления информации
    interval = 60
    # если режим работы программы "тест"
    if "тест" in workmode:
        # установим величину интервала 10 сек
        interval=10
    # установим метку пользователя программы
    # "начальник" или "инженер"
    program_user = config_value_dict["programUser"]
    # установим наименование подразделения, по которому
    #  получим список сотрудников
    subdivision=config_value_dict["subdivision"]
    # установим режим вывода статистики
    print_mode=config_value_dict["printMode"]
    # установим период обновления ф.Production.xlsx
    # в минутах
    production_update_interval = \
        config_value_dict["productionUpdateInterval"]
    # создадим переменную для хранения времени
    # последнего обновления ф.Production.xlsx
    # в секундах
    production_last_update=0
    # установим режим обновления ф.Production.xlsx
    # по времени последнего изменения
    production_update_mode = "изменен"
    # если в конфиг.файле отключено обновление ф.Production.xlsx
    if config_value_dict["productionUpdateOn"] == False:
        # отключим обновление этого ф.
        production_update_mode = "откл"
    
    # получим тек. дату
    date_cur_list = toformatNow()
    # установим дату, за которую получаем информацию
    # по умолчанию - текущая дата
    date_filter = date_cur_list[1]

    # установим фильтр по подразделению
    filter_eqv_dic = {"subdivision": [subdivision]}
    

    # назначаем обработчик события нажатия клавиши
    keyboard.on_press(onPress)
    # очистим экран
    os.system("CLS")
    # подготовим метку, что в таблице
    # статистики по сотрудникам 
    # отсутствует столбец "Прочие статусы"
    view_other_grade=False

    # обновим ф.Production.xlsx с плановыми значениями в
    # локальной папке
    res = updateProductionxlsx(
        workmode=workmode, update_mode=production_update_mode)
    # если успешно обновили файл
    # если вернулась ошибка
    if res[0] == "0":
        # Попробуем обновить файл через 5 минут
        production_last_update = toformatNow()[3]+5*60
    # успешно обновили файл
    else:
        # получим время последнего обновления ф.Production.xlsx
        production_last_update = res[2]

    # установим статус программы: в цикле
    programm_status = "цикл"

    # пока установлена метка работы программы
    while programm_status == "цикл" or  "ввод даты" in programm_status:
        # получим тек. дату
        date_cur_list = toformatNow()
        # получим тек. время в секундах
        time_sec=date_cur_list[3]
        # если истек период обновления ф.Production.xlsx
        if production_last_update+production_update_interval*60 < time_sec:
            # обновим ф.Production.xlsx с плановыми значениями в
            # локальной папке
            res = updateProductionxlsx(workmode=workmode)
            # если вернулась ошибка
            if res[0] == "0":
                # Попробуем обновить файл через 5 минут
                production_last_update=time_sec+5*60
            # успешно обновили файл
            else:
                # получим время последнего обновления ф.Production.xlsx
                production_last_update = res[2]

        # если текущий статус программы "ввод даты -1 день"
        if programm_status == "ввод даты -1 день":
            # Сдвинем дату в фильтре на один день назад 
            # от текущей отчетной даты в фильтре
            # преобразуем отчетную дату во внутр. формат datetime
            a = checkCorrectDate(date_filter)[3]
            # сдвинем на 1 день назад
            a=a-timedelta(days=1)
            # преобразуем полученную дату в строку
            a=str(a)
            # приведем строку в формат "ДД.ММ.ГГГГ"
            date_filter=f"{a[8:10]}.{a[5:7]}.{a[0:4]}"
            # переведем программу в статус "цикл"
            programm_status = "цикл" 
        
        # если текущий статус программы "ввод даты
        elif programm_status == "ввод даты":
            while True:
                # запросим ввод отчетной даты для фильтра
                txt1 = f"\n{bcolors.OKBLUE}Введите отчетную дату в формате " \
                    f"ДД/ММ/ГГГГ.{bcolors.ENDC}\n" \
                    f"{bcolors.OKBLUE}Для ввода текущей даты можно нажать 0." \
                    f"{bcolors.ENDC}"
                oo = inputSpecifiedKey(colortxt="", txt=txt1,
                        err_txt="", len_input_list=[0], specified_keys_list=["0", "#date"])
                # если нажали "0"
                if oo == "0":
                    # установим в качестве даты-фильтра текущую дату
                    date_filter = date_cur_list[1]
                # ввели корректную дату
                else:
                    # преобразуем введенную дату во внутр. формат datetime
                    a = checkCorrectDate(oo)[3]
                    # если введенная дата идет позднее тек. даты
                    if a > date_cur_list[2]:
                        # выведем сообщение о том, что это дата из будущего
                        print(f"{bcolors.WARNING}Введена дата {oo} позднее текущей " \
                              f"даты {date_cur_list[1]}.{bcolors.ENDC}")
                        # попросим ввести другую дату
                        # перейдем в начало цикла
                        continue
                    # установим фильтр по дате
                    date_filter = oo

                # обновим список сотрудников из ф.Employeers.xlsx,
                # по которым выводим статистику из СУТП
                res = getEmployeeList(date_filter, filter_eqv_dic)
                # если вернулась ошибка
                if res[0] == "0":
                    # выйдем из программы
                    sys.exit()
                # если список сотрудников заполнен
                elif res[0] == "1":
                    # выйдем из цикла
                    break

            # получим список-фильтр из полного и краткого
            #  ФИО сотрудников
            # employeer_filter_list = getEmployerFIOList(res[2])
            # переведем программу в статус "цикл"
            programm_status = "цикл"


        # установим таб. номер сотрудника по умолчанию
        employee_id = 0
        # создадим пустой список значений
        # по умолчанию для инженера
        opto_run_value_dict = []
        # если пользователем программы является "инженер",
        if program_user == "инженер":
            # прочитаем сохраненные в файле "opto_run.json" значения по умолчанию
            opto_run_value_dict = optoRunVarRead()
            # прочитаем таб. номер сотрудника
            employee_id = opto_run_value_dict['employee_id']
            # добавим в фильтр таб. номер инженер, т.к.
            # выводить данные будем только по нему
            filter_eqv_dic["employeeId"] = [employee_id]

        # подготовим переменную для фильтра списка сотрудников
        employeer_filter_list=[]

        # получим список сотрудников из ф.Employeers.xlsx,
        # по которым выводим статистику из СУТП
        # не будем выводить сообщения об ошибках print_err="0" при
        # выполнении функции
        res = getEmployeeList(date_filter, filter_eqv_dic,print_err="0")
        # если вернулась ошибка
        if res[0] == "0":
            # выведем сообщение об ошибке
            print(f"\n{bcolors.WARNING}При формировании списка "
                f"сотрудников {subdivision} возникла ошибка.{bcolors.ENDC}")
            txt1 = f"{bcolors.OKBLUE}Нажмите Ввод.{bcolors.ENDC}"
            questionSpecifiedKey("", txt1, ["\r", "", 1])
            # выйдем из программы
            sys.exit()
        # если список сотрудников заполнен
        elif res[0]=="1":
            # получим список-фильтр из полного и краткого
            #  ФИО сотрудников
            employeer_filter_list = getEmployerFIOList(res[2])
        
        # список-фильтр сотрудников будет пустым, чтобы
        # получить инф. по всем операциям
        # employeer_filter_list=[]

        # очистим экран
        os.system("CLS")
        # выведем статистику из СУТП о кол-ве ПУ,
        # прошедших через ОТК на экран за дату date_filter
        printStatisticSUTP(workmode=workmode, date_filter=date_filter)

        # выводим только данные из статусов "ОТК пройден",
        # "В ремонте на линии", "Дефект". Значения других статусов
        #  будут сведены в общий столбец "Прочие статусы"
        # print_mode = "ОТК+прочие"

        # выводим все найденные статусы для сотрудников
        # print_mode = "все статусы"

        # выводим 2 таблицы: 
        # первая таблица содержит только данные из статусов "ОТК пройден",
        # "В ремонте на линии", "Дефект",
        # другая таблица содержит значения по другим статусам
        # print_mode = "ОТК,другие"
        
        # подготовим переменную для вывода надписи об
        #  управляющей клавише "7"
        txt_7=""

        # если список-фильтр сотрудников заполнен
        if len(employeer_filter_list) > 0:
            # выведем статистику о кол-ве проверенных ПУ сотрудниками
            res=getStatisticEmployeerSUTP(employeer_filter_list=employeer_filter_list,
                date_filter=date_filter, workmode=workmode,
                print_mode=print_mode)
       
            # если в таблице имеется столбец "Прочие статусы"
            # и установлен режим вывода ""ОТК+прочие""
            if res[0]=="3" and print_mode == "ОТК+прочие":
                # сделаем надпись об управляющей клавише
                txt_7=f"{bcolors.OKGREEN}7 - раскрыть 'Прочие статусы' по сотрудникам{bcolors.ENDC}\n"
                # поставим метку, что можно раскрыть табл.
                # статистики по сотрудникам, т.к. имеется столбец
                # "Прочие статусы"
                view_other_grade=True
            # если режим вывода "ОТК,другие"
            elif print_mode == "ОТК,другие":
                # сделаем надпись об управляющей клавише
                txt_7 = f"{bcolors.OKGREEN}7 - свернуть 'Прочие статусы' по сотрудникам{bcolors.ENDC}\n"
        # если список сотрудников пуст
        else:
            # если пользователь "начальник"
            if program_user=="начальник":
                # подготовим сообщение
                txt1=f"{bcolors.WARNING}Не удалось сформировать список " \
                    f"сотрудников для {subdivision}.{bcolors.ENDC}\n"
            # если пользователь "инженер"
            else:
                # подготовим сообщение
                txt1=f"{bcolors.WARNING}Сотрудник с табельным номером " \
                    f"{employee_id} в списке подразделения {subdivision} " \
                    f"не найден.{bcolors.ENDC}\n"
            # выведем сообщение
            print (f"\n{txt1}")
                
        
        # если программы работает в режим ТЕСТ
        if "тест" in workmode:
            # выведем сообщение
            print (f"{bcolors.WARNING}Программа работает в режиме ТЕСТ.{bcolors.ENDC}")

        # Подготовим запись о периоде обновления информации
        txt1_2=f"{index_retry_interval} мин"
        # если interval=10
        if interval==10:
            # будем выводить "сек"
            txt1_2=f"{str(index_retry_interval*interval)} сек"
        txt1 = f"{bcolors.OKGREEN}1-5 - время обновления информации (1-5 мин.): " \
                f"{txt1_2}.{bcolors.ENDC}\n"
        txt1_1 = f"{bcolors.OKGREEN}Enter - обновить информацию немедленно{bcolors.ENDC}\n" \
        
        # если дата отображения информация отличается от текущей даты
        if date_filter!=date_cur_list[1]:
            # уберем из текста об управляющих клавишах информацию 
            # об изменении времени обновления сведений из СУТП
            # и о немедленном обновлении
            txt1=""
            txt1_1=f"{bcolors.OKGREEN}0 - вернуться в текущую дату{bcolors.ENDC}\n" \
                f"{bcolors.OKGREEN}- - сдвинуться на 1 день назад от указанной " \
                    f"отчетной даты{bcolors.ENDC}\n" \
                f"{bcolors.OKGREEN}+ - сдвинуться на 1 день вперед от указанной " \
                    f"отчетной даты{bcolors.ENDC}\n"
        
        txt_date = date_filter
        txt1_2=""
        # если дата в фильтре сегодняшняя
        if date_filter == date_cur_list[1]:
            txt_date = "сегодня"
            txt1_2=f"{bcolors.OKGREEN}- - показать отчет за предыдущий день{bcolors.ENDC}\n"

        # выведем список управляющих клавиш
        txt = f"{bcolors.OKGREEN}Список управляющих клавиш:{bcolors.ENDC}\n" \
            f"{txt1}" \
            f"{bcolors.OKGREEN}6 - изменить отчетную дату: {txt_date}{bcolors.ENDC}\n" \
            f"{txt_7}" \
            f"{txt1_2}" \
            f"{txt1_1}" \
            f"{bcolors.OKGREEN}ESC - выйти из программы{bcolors.ENDC}"
        print(txt)

        # если дата отображения информация является текущей датой
        if date_filter == date_cur_list[1]:
            # рассчитаем интервал обновления информации в сек.
            retry_interval = index_retry_interval*interval
            # осталось времени
            time_left = retry_interval
            print(f"\n{bcolors.OKGREEN}До обновления информации осталось: "
                f"{bcolors.ENDC}", end="")
            for t in range(0, retry_interval):
                # если статус программы "ввод даты"
                if "ввод даты" in programm_status:
                    # выйдем из цикла for t
                    break
                # получим текущее время в формате секунд
                minutes, seconds = divmod(time_left, 60)
                # преобразуем оставшееся время в формат ММ:СС
                time_left_format = "{:02d} мин {:02d} сек".format(minutes, seconds)
                # выведем на экран оставшееся время до обновления информации
                print(f"{bcolors.OKGREEN}{time_left_format}{bcolors.ENDC}",
                    end="", flush=True)
                # уменьшим оставшееся время ожидания
                time_left -= 1
                # если статус работы программы "выйти"
                if programm_status == "выйти":
                    # выйдем из цикла for t
                    # и цикла programm_status
                    break
                # если статус программы "обновить"
                elif programm_status == "обновить":
                    # установим статус программы: в цикле
                    programm_status = "цикл"
                    # выйдем из for t и вернемся
                    # в начало цикла programm_status
                    break
                # задержка 1 сек
                time.sleep(1)
                # сотрем выведенное значение оставшегося времени
                # (сдвинемся влево на 13 символов)
                print("\033[13D", end="", flush=True)
        # если в фильтре дата отличается от текущей
        else:
            # установим статус программы "цикл отключен"
            programm_status = "цикл отключен"
            # зададим список ожидаемых нажатие клавиш
            specified_keys=["\x1b","6","0","-","+"]
            # если установлена метка о возможности раскрытия
            # табл. по сотрудникам
            if view_other_grade:
                # дополним список ожидаемых клавиш
                # клавишей "7"
                specified_keys.append("7")
            # если установлен режим вывода
            # ожидаем нажатие одной из клавиш: "ESC","6", "0","-","+"
            oo=questionSpecifiedKey("","",specified_keys,"",
                specified_keys_only=1)
            # если нажали "0"
            if oo == "0":
                # установим в качестве даты-фильтра текущую дату
                date_filter = date_cur_list[1]
                # переведем программу в статус "цикл"
                programm_status = "цикл"
                # переведем режим вывода в "ОТК+прочие"
                print_mode = "ОТК+прочие"
                # сбросим метку возможности раскрытия
                # таблицы статистики по сотрудникам
                view_other_grade=False
            # если нажали "-"
            elif oo=="-":
                # Сдвинем дату в фильтре на один день назад 
                # от текущей отчетной даты в фильтре
                # преобразуем отчетную дату во внутр. формат datetime
                a = checkCorrectDate(date_filter)[3]
                # сдвинем на 1 день назад
                a=a-timedelta(days=1)
                # преобразуем полученную дату в строку
                a=str(a)
                # приведем строку в формат "ДД.ММ.ГГГГ"
                date_filter=f"{a[8:10]}.{a[5:7]}.{a[0:4]}"
                # переведем программу в статус "цикл"
                programm_status = "цикл"
                # переведем режим вывода в "ОТК+прочие"
                print_mode = "ОТК+прочие"
                # сбросим метку возможности раскрытия
                # таблицы статистики по сотрудникам
                view_other_grade = False
            # если нажали "+"
            elif oo=="+":
                # Сдвинем дату в фильтре на один день вперед 
                # от текущей отчетной даты в фильтре
                # преобразуем отчетную дату во внутр. формат datetime
                a = checkCorrectDate(date_filter)[3]
                # сдвинем на 1 день назад
                a=a+timedelta(days=1)
                # преобразуем полученную дату в строку
                a=str(a)
                # приведем строку в формат "ДД.ММ.ГГГГ"
                date_filter=f"{a[8:10]}.{a[5:7]}.{a[0:4]}"
                # переведем программу в статус "цикл"
                programm_status = "цикл"
                # переведем режим вывода в "ОТК+прочие"
                print_mode = "ОТК+прочие"
                # сбросим метку возможности раскрытия
                # таблицы статистики по сотрудникам
                view_other_grade = False
            # если нажали "7" (раскрытие статуса "прочие")
            elif oo=="7":
                # если режим вывода "ОТК+прочие"
                if print_mode == "ОТК+прочие":
                    # установим режим вывода 2 таблиц для статистики по персоналу:
                    # первая таблица содержит только данные из статусов "ОТК пройден",
                    # "В ремонте на линии", "Дефект",
                    # другая таблица содержит значения по другим статусам
                    print_mode = "ОТК,другие"
                # иначе
                else:
                    # установим режим вывода "ОТК+прочие"
                    print_mode = "ОТК+прочие"
                # переведем программу в статус "цикл"
                programm_status = "цикл"

    # # запускаем цикл обработки нажатия клавиш
    # keyboard.wait()




if __name__=="__main__":

    # сохраним в ф.link_current_folder.json ссылку на тек. рабочий каталог
    saveLinkCurFolder()
    
    # # ДЛЯ ТЕСТА
    # # сформируем список типов устройств, по которым выводим статистику
    # product_filter_list=["ИПУ 1Ф", "ИПУ 1Ф S", "ИПУ 3Ф", 
    #     "ИПУ 3Ф ТТ", "ИПУ 3Ф S"]
    # product_filter_list=["ИПУ 1Ф", "ИПУ 1Ф S"]
    # # получим статистику
    # res=printStatisticSUTP("27.02.2025", product_filter_user_list=product_filter_list)

    # sys.exit()
    
    # Изменим заголовок окна "otk_printStatisticSUTP"
    # на новый заголовок "Статистика по проверке ПУ"
    # заголовок окна, который ищем
    title_old = "otk_printStatisticSUTP"
    # новый заголовок окна
    title_new = "Статистика по проверке ПУ"
    # найдем окно с указанным заголовком и
    # заменим в нем заголовок
    res = replaceTitleWindow(title_old, title_new)
    
    # выводим статистику из СУТП
    # по проверке ПУ: всего и по сотрудникам ОТК
    printStatisticMain()
    